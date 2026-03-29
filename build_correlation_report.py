import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

#############################################
# 1. Load stock per jurisdiction per day
#############################################
stock_file = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\EXISTENCIAS SRP Y SR\EXISTENCIAS_SRP_SR_POR_DIA.xlsx"
df_stock = pd.read_excel(stock_file, sheet_name='Detalle por Jurisdicción')

def parse_date(d):
    try:
        return pd.to_datetime(d, dayfirst=True).strftime('%d/%m/%Y')
    except:
        return str(d).strip()

df_stock['FECHA_STR'] = df_stock['FECHA'].apply(parse_date)
df_stock['SECCIÓN'] = df_stock['SECCIÓN'].astype(str).str.strip()

# Normalize jurisdiction names to standard format
def normalize_jur(s):
    s = s.upper().strip()
    if 'NO. 1' in s or 'NO.1' in s or s.endswith('1'): return 'JURISDICCIÓN No. 1'
    if 'NO. 2' in s or 'NO.2' in s or s.endswith('2'): return 'JURISDICCIÓN No. 2'
    if 'NO. 3' in s or 'NO.3' in s or s.endswith('3'): return 'JURISDICCIÓN No. 3'
    if 'NO. 4' in s or 'NO.4' in s or s.endswith('4'): return 'JURISDICCIÓN No. 4'
    return s

df_stock['JUR_NORM'] = df_stock['SECCIÓN'].apply(normalize_jur)

# Only jurisdiction rows
jur_pattern = r'JURISDICCIÓN'
df_jur_stock = df_stock[df_stock['JUR_NORM'].str.contains('JURISDICCIÓN', na=False)].copy()
df_jur_stock = df_jur_stock.rename(columns={
    'SR PUNTOS': 'EXIST_SR_PUNTOS',
    'SRP PUNTOS': 'EXIST_SRP_PUNTOS',
    'TOTAL SR': 'EXIST_TOTAL_SR',
    'TOTAL SRP': 'EXIST_TOTAL_SRP'
})

# Sort by date and jurisdiction for delta computation
df_jur_stock['FECHA_PARSED'] = pd.to_datetime(df_jur_stock['FECHA_STR'], format='%d/%m/%Y', errors='coerce')
df_jur_stock = df_jur_stock.sort_values(['JUR_NORM', 'FECHA_PARSED'])

# Compute consumption = previous stock - current stock (per jurisdiction)
for col_prev, col_result in [('EXIST_SR_PUNTOS', 'CONSUMO_SR'), ('EXIST_SRP_PUNTOS', 'CONSUMO_SRP')]:
    df_jur_stock[col_result] = df_jur_stock.groupby('JUR_NORM')[col_prev].diff().mul(-1)

#############################################
# 2. Load applied doses per jurisdiction per day from CSV
#############################################
csv_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\REPORTE_SRP-SR-CENSIA\SRP-SR-2025_14-03-2026 06-41-16.csv"
df_doses = pd.read_csv(csv_path, usecols=[
    'JURISDICCION', 'Fecha de registro',
    'SR PRIMERA TOTAL', 'SR SEGUNDA TOTAL',
    'SRP  PRIMERA TOTAL', 'SRP SEGUNDA TOTAL'
])

df_doses['FECHA_PARSED'] = pd.to_datetime(df_doses['Fecha de registro'], errors='coerce')
df_doses = df_doses[df_doses['FECHA_PARSED'].dt.year >= 2026]
df_doses['FECHA_STR'] = df_doses['FECHA_PARSED'].dt.strftime('%d/%m/%Y')
df_doses['SR_APLICADAS'] = df_doses[['SR PRIMERA TOTAL', 'SR SEGUNDA TOTAL']].fillna(0).sum(axis=1)
df_doses['SRP_APLICADAS'] = df_doses[['SRP  PRIMERA TOTAL', 'SRP SEGUNDA TOTAL']].fillna(0).sum(axis=1)

# Map CSV jurisdiction names to standard format
csv_jur_map = {
    'DURANGO': 'JURISDICCIÓN No. 1',
    'SANTIAGO PAPASQUIARO': 'JURISDICCIÓN No. 2',
    'RODEO': 'JURISDICCIÓN No. 3',
}

def map_jur_csv(j):
    j = str(j).strip().upper()
    for k, v in csv_jur_map.items():
        if k in j:
            return v
    return j

df_doses['JUR_NORM'] = df_doses['JURISDICCION'].apply(map_jur_csv)

doses_agg = df_doses.groupby(['FECHA_STR', 'FECHA_PARSED', 'JUR_NORM']).agg(
    SR_APLICADAS=('SR_APLICADAS', 'sum'),
    SRP_APLICADAS=('SRP_APLICADAS', 'sum')
).reset_index()

#############################################
# 3. Merge stock + doses
#############################################
merged = pd.merge(
    df_jur_stock[['FECHA_STR', 'FECHA_PARSED', 'JUR_NORM', 'EXIST_SR_PUNTOS', 'EXIST_SRP_PUNTOS', 'CONSUMO_SR', 'CONSUMO_SRP']],
    doses_agg[['FECHA_STR', 'JUR_NORM', 'SR_APLICADAS', 'SRP_APLICADAS']],
    on=['FECHA_STR', 'JUR_NORM'],
    how='left'
).fillna(0)

merged['SR_APLICADAS'] = merged['SR_APLICADAS'].astype(int)
merged['SRP_APLICADAS'] = merged['SRP_APLICADAS'].astype(int)
merged['CONSUMO_SR'] = merged['CONSUMO_SR'].fillna(0).astype(int)
merged['CONSUMO_SRP'] = merged['CONSUMO_SRP'].fillna(0).astype(int)
merged['DIFERENCIA_SR'] = merged['CONSUMO_SR'] - merged['SR_APLICADAS']
merged['DIFERENCIA_SRP'] = merged['CONSUMO_SRP'] - merged['SRP_APLICADAS']
merged = merged.sort_values(['FECHA_PARSED', 'JUR_NORM'])

#############################################
# 4. Build Excel report
#############################################
wb = Workbook()

JURISDICTIONS = ['JURISDICCIÓN No. 1', 'JURISDICCIÓN No. 2', 'JURISDICCIÓN No. 3', 'JURISDICCIÓN No. 4']

header_dark  = PatternFill("solid", fgColor="1F3864")
header_blue  = PatternFill("solid", fgColor="2E75B6")
header_green = PatternFill("solid", fgColor="375623")
header_red   = PatternFill("solid", fgColor="C00000")
alt_fill     = PatternFill("solid", fgColor="ECF4FF")
white_fill   = PatternFill("solid", fgColor="FFFFFF")
yellow_fill  = PatternFill("solid", fgColor="FFEB9C")
red_fill     = PatternFill("solid", fgColor="FFC7CE")
green_fill   = PatternFill("solid", fgColor="C6EFCE")

def bold_white(size=10): return Font(bold=True, color="FFFFFF", size=size)
def bold_black(size=10): return Font(bold=True, size=size)

# --- SR correlation sheet ---
for jur_idx, jur_name in enumerate(JURISDICTIONS):
    sheet_name = f"SR - J{jur_idx+1}"
    ws = wb.active if jur_idx == 0 else wb.create_sheet(sheet_name)
    ws.title = sheet_name

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    tc = ws.cell(row=1, column=1, value=f"CORRELACIÓN SR - {jur_name} | CAMPAÑA SARAMPIÓN 2026")
    tc.font = Font(bold=True, color="FFFFFF", size=13)
    tc.fill = header_dark
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['FECHA', 'EXISTENCIA SR (PUNTOS)', 'CONSUMO TEÓRICO (Δ STOCK)', 'DOSIS APLICADAS', 'DIFERENCIA / DESFASE']
    fills_h = [header_dark, header_blue, header_blue, header_green, header_red]
    for ci, (h, f) in enumerate(zip(headers, fills_h)):
        c = ws.cell(row=2, column=ci+1, value=h)
        c.fill = f
        c.font = bold_white(11)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 35

    df_jur = merged[merged['JUR_NORM'] == jur_name].sort_values('FECHA_PARSED')

    for ri, (_, row) in enumerate(df_jur.iterrows()):
        r = ri + 3
        alt = alt_fill if ri % 2 == 0 else white_fill

        ws.cell(row=r, column=1, value=row['FECHA_STR']).fill = alt
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')

        ws.cell(row=r, column=2, value=int(row['EXIST_SR_PUNTOS'])).fill = alt
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')

        consumo = int(row['CONSUMO_SR'])
        c3 = ws.cell(row=r, column=3, value=consumo)
        c3.fill = alt
        c3.alignment = Alignment(horizontal='center')

        aplicadas = int(row['SR_APLICADAS'])
        ws.cell(row=r, column=4, value=aplicadas).fill = alt
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')

        diferencia = int(row['DIFERENCIA_SR'])
        c5 = ws.cell(row=r, column=5, value=diferencia)
        c5.fill = red_fill if abs(diferencia) > 100 else (green_fill if diferencia == 0 else yellow_fill)
        c5.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 22

# --- SRP correlation sheets ---
for jur_idx, jur_name in enumerate(JURISDICTIONS):
    sheet_name = f"SRP - J{jur_idx+1}"
    ws = wb.create_sheet(sheet_name)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    tc = ws.cell(row=1, column=1, value=f"CORRELACIÓN SRP - {jur_name} | CAMPAÑA SARAMPIÓN 2026")
    tc.font = Font(bold=True, color="FFFFFF", size=13)
    tc.fill = PatternFill("solid", fgColor="375623")
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['FECHA', 'EXISTENCIA SRP (PUNTOS)', 'CONSUMO TEÓRICO (Δ STOCK)', 'DOSIS APLICADAS', 'DIFERENCIA / DESFASE']
    for ci, h in enumerate(headers):
        c = ws.cell(row=2, column=ci+1, value=h)
        c.fill = PatternFill("solid", fgColor="375623")
        c.font = bold_white(11)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 35

    df_jur = merged[merged['JUR_NORM'] == jur_name].sort_values('FECHA_PARSED')

    for ri, (_, row) in enumerate(df_jur.iterrows()):
        r = ri + 3
        alt = alt_fill if ri % 2 == 0 else white_fill

        ws.cell(row=r, column=1, value=row['FECHA_STR']).fill = alt
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=int(row['EXIST_SRP_PUNTOS'])).fill = alt
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
        
        consumo = int(row['CONSUMO_SRP'])
        ws.cell(row=r, column=3, value=consumo).fill = alt
        ws.cell(row=r, column=3).alignment = Alignment(horizontal='center')
        
        aplicadas = int(row['SRP_APLICADAS'])
        ws.cell(row=r, column=4, value=aplicadas).fill = alt
        ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')

        diferencia = int(row['DIFERENCIA_SRP'])
        c5 = ws.cell(row=r, column=5, value=diferencia)
        c5.fill = red_fill if abs(diferencia) > 100 else (green_fill if diferencia == 0 else yellow_fill)
        c5.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 22

out_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\EXISTENCIAS SRP Y SR\CORRELACION_SR_SRP_POR_JURISDICCION.xlsx"
wb.save(out_path)
print(f"Saved to: {out_path}")
print(f"Rows in merged data: {len(merged)}")
print(f"Jurisdictions found: {merged['JUR_NORM'].unique()}")
