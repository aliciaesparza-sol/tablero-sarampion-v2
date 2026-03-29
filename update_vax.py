import pandas as pd
from openpyxl import load_workbook

# Paths
csv_path = r'c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\COBERTURA DE VACUNACIÓN\TABLERO SARAMPION\SRP-SR-2025_21-03-2026 04-30-33.csv'
excel_path = r'c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\Cobertura_SRP-SR_SSA_Durango_2026.xlsx'

# Load CSV
df = pd.read_csv(csv_path)
dur = df[df['ESTADO'] == 'DURANGO']

def get_sum(query):
    return dur.filter(like=query).fillna(0).sum().sum()

# Calculations
new_doses = {
    "6 a 11 meses": get_sum('6 A 11 MESES'),
    "1 año": get_sum('1 ANIO'),
    "18 meses": get_sum('18 MESES'),
    "Rezagados 2 a 12 años": get_sum('2 A 5 ANIOS') + get_sum('6 ANIOS') + get_sum('7 A 9 ANIOS') + get_sum('10 A 12 ANIOS'),
    "13 a 19 años": get_sum('13 A 19 ANIOS'),
    "20 a 39 años": get_sum('20 A 29 ANIOS') + get_sum('30 A 39 ANIOS'),
    "40 a 49 años": get_sum('40 A 49 ANIOS')
}

print("New calculated doses:")
for k, v in new_doses.items():
    print(f"  {k}: {int(v)}")

# Update Excel
wb = load_workbook(excel_path)
sheet = wb.active

# Row indices based on our inspection (CSV dump showed Row 3 was header, Row 4 was 6-11m)
# Categories mapping to rows (Row numbers are 1-based)
mapping = {
    "6 a 11 meses": 4,
    "1 año": 5,
    "18 meses": 6,
    "Rezagados 2 a 12 años": 7,
    "13 a 19 años": 8,
    "20 a 39 años": 9,
    "40 a 49 años": 10
}

# The column for 'Dosis Aplicadas (CSV)' is D (4th column)
d_col = 4

for category, row_idx in mapping.items():
    val = new_doses[category]
    sheet.cell(row=row_idx, column=d_col).value = int(val)

# Update the Source field in Row 2
source_val = f"Fuente dosis aplicadas: SRP-SR-2025_21-03-2026.csv (SRP-SR SSA)  |  Corte: 21 de marzo 2026"
sheet.cell(row=2, column=1).value = source_val

# Save changes
wb.save(excel_path)
print(f"\nExcel file updated successfully at: {excel_path}")
