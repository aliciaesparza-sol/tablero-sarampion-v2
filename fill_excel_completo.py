
"""
Script: fill_excel_completo.py
Propósito: Versión mejorada - usa CSV EFES + PDFs para rellenar campos vacíos
           en CASOS_NOTIFICADOS_2026_ACTUALIZADO.xlsx desde el caso 110 en adelante.
"""

import zipfile
import pdfplumber
import io
import os
import re
import openpyxl

# ── RUTAS ───────────────────────────────────────────────────────────────────
base = os.path.expanduser('~') + '/OneDrive/Escritorio/PVU'

excel_path = None
zip_path = None

for root, dirs, files in os.walk(base):
    for f in files:
        if f == 'CASOS_NOTIFICADOS_2026_ACTUALIZADO.xlsx' and 'EE casos' in root:
            excel_path = os.path.join(root, f)
        if f.endswith('.zip') and 'EE casos' in f:
            zip_path = os.path.join(root, f)

print(f"Excel : {excel_path}")
print(f"ZIP   : {zip_path}")

# ── LEER CSV DE EFES ─────────────────────────────────────────────────────────
print("\nLeyendo CSV EFES...")
efes_data = {}  # {efes_id: {campo: valor}}

with zipfile.ZipFile(zip_path, 'r') as z:
    csv_bytes = z.read('EE casos confirmados sarampion 2026/EFES.csv')
    csv_text = csv_bytes.decode('utf-8-sig', errors='replace')

csv_lines = csv_text.strip().split('\n')
headers = csv_lines[0].split('|')

# Índices de columnas relevantes
col = {h: i for i, h in enumerate(headers)}

# Mapa de CVE_TIP_VAC -> tipo de vacuna
tipo_vac_map = {
    '1': 'SR', '2': 'SRP', '3': 'SRP', '4': 'SR', '0': None
}

for line in csv_lines[1:]:
    if not line.strip():
        continue
    parts = line.split('|')

    def g(c, default=''):
        idx = col.get(c)
        if idx is None or idx >= len(parts):
            return default
        return parts[idx].strip()

    efes_id = g('EFES_ID')
    if not efes_id:
        continue

    # Tipo de localidad basado en MUNMAYORHAB y DES_LOC
    mayor_hab = g('MUNMAYORHAB', 'NO').upper()
    des_loc = g('DES_LOC', '').upper()
    tipo_loc = 'URBANA' if mayor_hab == 'SI' else 'RURAL'

    # Antecedente vacunal
    cta_vac = g('CTA_VAC_SAR', '0')
    if cta_vac == '1':
        ant_vac = 'Si'
    elif cta_vac == '2':
        ant_vac = 'No'
    else:
        ant_vac = None

    # Número de dosis (de la CURP y el número de dosis)
    num_dosis_val = g('DOSIS', '0')
    try:
        num_dosis_val = int(float(num_dosis_val)) if num_dosis_val and num_dosis_val != '0' else None
    except:
        num_dosis_val = None

    # Tipo de vacuna
    cveTipVac = g('CVE_TIP_VAC', '0')
    tipo_vac = tipo_vac_map.get(cveTipVac)

    # Clasificación final
    clas_final = g('CLAS_FINAL_SARAMPION', '')
    if 'CONFIRMADO' in clas_final.upper():
        resultado_lab = 'POSITIVO'
    elif 'DESCARTADO' in clas_final.upper():
        resultado_lab = 'NEGATIVO'
    else:
        resultado_lab = None

    # ES INDIGENA: 1=Si, 2=No, 3=Se desconoce
    es_ind_cve = g('ES_INDIGENA', '3')
    if es_ind_cve == '1':
        es_ind = 'Si'
    elif es_ind_cve == '2':
        es_ind = 'No'
    else:
        es_ind = 'Se desconoce'

    # Institución
    unidad_notif = g('UNIDAD_NOTIFICANTE', '').strip()
    clues = g('CLUES', '').strip()

    # Mapa institución de la CLUES
    institcion_notif = None
    if clues.startswith('DGIMS'):
        institcion_notif = 'IMSS'
    elif clues.startswith('DGSSA'):
        institcion_notif = 'SSA'
    elif clues.startswith('DGIST'):
        institcion_notif = 'ISSSTE'
    elif clues.startswith('DGSDN'):
        institcion_notif = 'SEDENA'
    elif clues.startswith('DGPEM'):
        institcion_notif = 'PEMEX'

    efes_data[efes_id] = {
        'rr_tipo_localidad': tipo_loc,
        'municipio': g('DES_MPO', '').strip().title(),
        'antecedente_vacunacion': ant_vac,
        'numero_dosis': num_dosis_val,
        'tipo_vacuna': tipo_vac,
        'resultado_laboratorio': resultado_lab,
        'es_indigena': es_ind,
        'rr_unidad_salud_clues': clues,
        'institucion_notificante': institcion_notif,
        'institucion_responsabilidad': institcion_notif,
    }

print(f"Registros en EFES CSV: {len(efes_data)}")

# ── FUNCIONES EXTRACCIÓN PDF ─────────────────────────────────────────────────
def clean(s):
    if not s:
        return ""
    return re.sub(r'\s+', ' ', s.strip())


def extract_bloqueo_from_pdf(pdf_bytes):
    """Extrae datos de bloqueo vacunal (sección VII) del PDF."""
    result = {
        'rr_tactica_tipo': None,
        'rr_tactica_fecha': None,
        'rr_manzanas_cubiertas': None,
        'rr_srp_dosis_aplicadas': None,
        'rr_sr_dosis_aplicadas': None,
        'rr_total_dosis_aplicadas': None,
        'rr_poblacion_objetivo_estimada': None,
        'rr_cobertura_alcanzada': None,
        'rr_ageb_domicilio': None,
        'observaciones': None,
    }

    full_text = ""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                full_text += t + "\n"
    except Exception as e:
        return result

    lines = full_text.split("\n")

    for i, line in enumerate(lines):
        line_c = clean(line)

        # BLOQUEO - sección VII
        if 'BLOQUEO' in line_c.upper():
            result['rr_tactica_tipo'] = 'Bloqueo vacunal'

        # Línea con INICIO...DOSIS:...COBERTURA:
        if 'INICIO' in line_c and ('DOSIS:' in line_c or 'TERMINACI' in line_c):
            # Fecha inicio del bloqueo
            m_inicio = re.search(r'INICIO\s+(\d{1,2}/\d{1,2}/\d{4})', line_c)
            if m_inicio:
                result['rr_tactica_fecha'] = m_inicio.group(1)

            # Dosis
            m_dosis = re.search(r'DOSIS:\s*(\d+)', line_c)
            if m_dosis:
                d = int(m_dosis.group(1))
                result['rr_total_dosis_aplicadas'] = d
                result['rr_srp_dosis_aplicadas'] = d
                result['rr_sr_dosis_aplicadas'] = 0

            # Cobertura
            m_cob = re.search(r'COBERTURA:\s*(\d+(?:\.\d+)?)\s*%', line_c)
            if m_cob:
                cob = float(m_cob.group(1))
                result['rr_cobertura_alcanzada'] = cob / 100 if cob > 1 else cob

        # AGEB
        m_ageb = re.search(r'AGEB[:\s]+([A-Z0-9\-]+)', line_c, re.IGNORECASE)
        if m_ageb and not result['rr_ageb_domicilio']:
            result['rr_ageb_domicilio'] = m_ageb.group(1)

        # Observaciones
        if 'Observaciones:' in line_c or 'OBSERVACIONES:' in line_c:
            obs_lines = []
            for j in range(i+1, min(i+6, len(lines))):
                nl = clean(lines[j])
                if nl and not nl.upper().startswith('XI') and 'MEDICO' not in nl.upper():
                    obs_lines.append(nl)
                else:
                    break
            if obs_lines:
                result['observaciones'] = ' '.join(obs_lines)

    return result


# ── INDEXAR PDFs del ZIP ─────────────────────────────────────────────────────
print("\nIndexando PDFs del ZIP...")
pdf_index    = {}   # {numero_caso: nombre_archivo}
efes_id_index = {} # {numero_caso: efes_id}

with zipfile.ZipFile(zip_path, 'r') as z:
    for fname in z.namelist():
        m = re.search(r'/(\d+)[. ]+(\d+)\s+', fname)
        if m:
            caso_num = int(m.group(1))
            efes_folio = m.group(2)
            pdf_index[caso_num] = fname
            efes_id_index[caso_num] = efes_folio

print(f"PDFs: {len(pdf_index)} | Rango: {min(pdf_index)}-{max(pdf_index)}")

# ── LEER EXCEL CON OPENPYXL ──────────────────────────────────────────────────
print("\nCargando Excel...")
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Mapa columna -> número
col_map = {}
for c in range(1, 40):
    val = ws.cell(1, c).value
    if val:
        col_map[str(val).strip()] = c

print("Columnas encontradas:", list(col_map.keys())[:36])

# ── PROCESAR FILAS ────────────────────────────────────────────────────────────
print("\nProcesando filas del caso 110 en adelante...\n")

col_ns = col_map.get('numero_seriado', 1)
updated = 0
no_pdf = []

with zipfile.ZipFile(zip_path, 'r') as z:
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row, col_ns).value
        if not cell_val:
            continue

        num_str = str(cell_val).strip()
        m = re.match(r'^(\d+)\.(\d+)\.', num_str)
        if not m:
            continue

        caso_num = int(m.group(1))
        efes_folio_excel = m.group(2)

        if caso_num < 110:
            continue

        print(f"  Caso {caso_num} (EFES: {efes_folio_excel})", end=" -> ")

        # Obtener datos del CSV EFES usando el folio del numero_seriado
        efes_row = efes_data.get(efes_folio_excel, {})

        # Obtener datos del PDF
        bloqueo = {}
        if caso_num in pdf_index:
            try:
                pdf_bytes = z.read(pdf_index[caso_num])
                bloqueo = extract_bloqueo_from_pdf(pdf_bytes)
            except Exception as e:
                print(f"\n    ERROR PDF: {e}", end="")
        else:
            no_pdf.append(caso_num)

        # Combinar datos (CSV EFES tiene prioridad para datos del paciente;
        # PDF tiene prioridad para datos de bloqueo)
        all_data = {**efes_row, **bloqueo}

        # Función helper para llenar solo si está vacío
        def fill(col_name, value):
            if col_name not in col_map:
                return
            c = col_map[col_name]
            cur = ws.cell(row, c).value
            # Considera "vacío" si es None, '', 0 (para dosis), o 0.0
            is_empty = cur is None or str(cur).strip() in ('', 'None', 'NaN', 'nan')
            if is_empty and value is not None and str(value).strip() not in ('', 'None', 'nan'):
                ws.cell(row, c).value = value

        fill('rr_unidad_salud_clues', all_data.get('rr_unidad_salud_clues'))
        fill('rr_ageb_domicilio', all_data.get('rr_ageb_domicilio'))
        fill('rr_tipo_localidad', all_data.get('rr_tipo_localidad'))
        fill('rr_tactica_tipo', all_data.get('rr_tactica_tipo'))
        fill('rr_tactica_fecha', all_data.get('rr_tactica_fecha'))
        fill('rr_manzanas_cubiertas', all_data.get('rr_manzanas_cubiertas'))

        # Dosis: solo si > 0
        srp = all_data.get('rr_srp_dosis_aplicadas')
        sr  = all_data.get('rr_sr_dosis_aplicadas')
        tot = all_data.get('rr_total_dosis_aplicadas')
        if tot and tot > 0:
            fill('rr_srp_dosis_aplicadas', srp)
            fill('rr_sr_dosis_aplicadas', sr)
            fill('rr_total_dosis_aplicadas', tot)

        # Cobertura: solo si > 0
        cob = all_data.get('rr_cobertura_alcanzada')
        if cob and cob > 0:
            fill('rr_cobertura_alcanzada', cob)

        fill('rr_poblacion_objetivo_estimada', all_data.get('rr_poblacion_objetivo_estimada'))
        fill('antecedente_vacunacion', all_data.get('antecedente_vacunacion'))
        fill('numero_dosis', all_data.get('numero_dosis'))
        fill('tipo_vacuna', all_data.get('tipo_vacuna'))
        fill('es_indigena', all_data.get('es_indigena'))
        fill('resultado_laboratorio', all_data.get('resultado_laboratorio'))
        fill('municipio', all_data.get('municipio'))
        fill('institucion_notificante', all_data.get('institucion_notificante'))
        fill('institucion_responsabilidad', all_data.get('institucion_responsabilidad'))
        fill('Observaciones', all_data.get('observaciones'))

        print(f"CLUES={all_data.get('rr_unidad_salud_clues','-')}, "
              f"Loc={all_data.get('rr_tipo_localidad','-')}, "
              f"Táctica={all_data.get('rr_tactica_tipo','-')}, "
              f"Fecha={all_data.get('rr_tactica_fecha','-')}, "
              f"Clas={all_data.get('resultado_laboratorio','-')}")

        updated += 1

# ── GUARDAR ──────────────────────────────────────────────────────────────────
output_path = excel_path.replace('.xlsx', '_COMPLETO_V2.xlsx')
print(f"\nGuardando en: {output_path}")
wb.save(output_path)

print(f"\n{'='*65}")
print(f"RESUMEN FINAL:")
print(f"  Filas procesadas (>=110) : {updated}")
print(f"  PDFs no encontrados      : {len(no_pdf)} -> {no_pdf}")
print(f"  Registros en EFES CSV    : {len(efes_data)}")
print(f"  Archivo generado         : {output_path}")
print('='*65)
