import os
import re
import shutil
import openpyxl
from copy import copy
from datetime import datetime

# ─── Rutas ──────────────────────────────────────────────────────────────────
base = r'c:\Users\aicil\OneDrive\Escritorio\PVU'
sarampion_dir = [d for d in os.listdir(base)
                 if 'SARAMP' in d.upper() and os.path.isdir(os.path.join(base, d))][0]
folder = os.path.join(base, sarampion_dir, 'CASOS NOTIFICADOS', 'CASOS NOTIFICADOS 2026')
xlsx_name = [f for f in os.listdir(folder)
             if 'positivos' in f.lower() and f.endswith('.xlsx') and not f.startswith('~$')][0]
filepath = os.path.join(folder, xlsx_name)

# Crear backup con timestamp
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
backup_path = os.path.join(folder, f'BACKUP_{ts}_{xlsx_name}')
shutil.copy2(filepath, backup_path)
print(f'Backup creado: {backup_path}')

# ─── Cargar workbook ─────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(filepath)
ws = wb.active

# ─── Leer todas las filas (incluyendo encabezado) ────────────────────────────
# Guardamos el encabezado separado
header_row = [cell.value for cell in ws[1]]

# Leer filas de datos (fila 2 en adelante) como listas de valores
data_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    # Solo incluir filas que no estén completamente vacías
    if any(v is not None for v in row):
        data_rows.append(list(row))

print(f'Total filas de datos leídas: {len(data_rows)}')
print('Primeros 5 números seriados:')
for r in data_rows[:5]:
    print(' ', r[0])

# ─── Función para extraer la clave de ordenamiento ──────────────────────────
def sort_key(row):
    """
    El número seriado tiene el formato: N.NNNNN.N.XXXX
    Ordenamos por el PRIMER segmento numérico (número consecutivo).
    Ejemplo: '1.48157.0.GRRH' → clave = 1
             '2.49391.MRAA'   → clave = 2
    Si el valor no existe o no es parseable, lo enviamos al final.
    """
    seriado = row[0]
    if seriado is None:
        return float('inf')
    seriado_str = str(seriado).strip()
    # Tomar el primer segmento antes del primer punto
    parts = seriado_str.split('.')
    try:
        return int(parts[0])
    except (ValueError, IndexError):
        return float('inf')

# ─── Ordenar ─────────────────────────────────────────────────────────────────
data_rows_sorted = sorted(data_rows, key=sort_key)

print('\nOrden después de ordenar (primeros 10):')
for r in data_rows_sorted[:10]:
    print(f'  {r[0]}')

# ─── Escribir de vuelta al workbook ──────────────────────────────────────────
# Limpiar filas de datos existentes (dejar solo el encabezado en fila 1)
# Borramos desde fila 2 hacia adelante y re-escribimos
for i, sorted_row in enumerate(data_rows_sorted, start=2):
    for j, value in enumerate(sorted_row, start=1):
        ws.cell(row=i, column=j, value=value)

# Si el archivo original tenía más filas que las ordenadas, limpiar el exceso
max_row_original = ws.max_row
for extra_row in range(len(data_rows_sorted) + 2, max_row_original + 1):
    for col in range(1, ws.max_column + 1):
        ws.cell(row=extra_row, column=col, value=None)

# ─── Guardar ─────────────────────────────────────────────────────────────────
wb.save(filepath)
print(f'\n✅ Archivo guardado y ordenado correctamente: {filepath}')
print(f'   Total registros ordenados: {len(data_rows_sorted)}')
