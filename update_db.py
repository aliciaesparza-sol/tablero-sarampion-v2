import pandas as pd
import sys
import re
import os
from openpyxl import load_workbook

def parse_date_from_filename(filename):
    basename = os.path.basename(filename)
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', basename)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    return None

def process_daily_file(filepath):
    xl = pd.ExcelFile(filepath)
    df = xl.parse("Existencias en puntos")
    
    sr_row = df[df['Tipo de biológico'].astype(str).str.contains(r'SR\s*\(Doble viral', na=False, regex=True)]
    srp_row = df[df['Tipo de biológico'].astype(str).str.contains(r'SRP\s*\(Triple viral', na=False, regex=True)]
    
    sr_almacen = int(sr_row['Existencias de vacunas en almacenes'].iloc[0]) if not sr_row.empty else 0
    sr_puntos = int(sr_row['Existencias de vacunas en puntos de vacunación'].iloc[0]) if not sr_row.empty else 0
    
    srp_almacen = int(srp_row['Existencias de vacunas en almacenes'].iloc[0]) if not srp_row.empty else 0
    srp_puntos = int(srp_row['Existencias de vacunas en puntos de vacunación'].iloc[0]) if not srp_row.empty else 0
    
    total_sr = sr_almacen + sr_puntos
    total_srp = srp_almacen + srp_puntos
    
    return {
        'SR ALMACÉN': sr_almacen, 'SR PUNTOS': sr_puntos, 'TOTAL SR': total_sr,
        'SRP ALMACÉN': srp_almacen, 'SRP PUNTOS': srp_puntos, 'TOTAL SRP': total_srp,
        'GRAN TOTAL': total_sr + total_srp
    }

def update_db(db_path, daily_files):
    wb = load_workbook(db_path)
    
    if 'Resumen ' in wb.sheetnames:
        ws_resumen = wb['Resumen ']
    else:
        print("Sheet 'Resumen ' not found!")
        return
        
    if 'Detalle por Jurisdicción' in wb.sheetnames:
        ws_detalle = wb['Detalle por Jurisdicción']
    else:
        print("Sheet 'Detalle por Jurisdicción' not found!")
        return

    df_res = pd.read_excel(db_path, sheet_name='Resumen ')
    if 'FECHA' in df_res.columns:
        # handle datetime or strings
        def fmt_date(d):
            if pd.isna(d): return ""
            if isinstance(d, str):
                m = re.search(r'(\d{2})[/-](\d{2})[/-](\d{4})', d)
                if m: return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
                elif "00:00:00" in d: # sometimes "2026-02-17 00:00:00"
                   dt = pd.to_datetime(d, errors='coerce')
                   if not pd.isna(dt): return dt.strftime('%d/%m/%Y')
                return d.strip()
            else:
                return pd.to_datetime(d).strftime('%d/%m/%Y')
        existing_dates_res = set(df_res['FECHA'].apply(fmt_date))
    else:
        existing_dates_res = set()

    df_det = pd.read_excel(db_path, sheet_name='Detalle por Jurisdicción')
    if 'FECHA' in df_det.columns:
        existing_dates_det = set(df_det['FECHA'].apply(fmt_date))
    else:
        existing_dates_det = set()

    for fp in daily_files:
        date_str = parse_date_from_filename(fp)
        if not date_str:
            print(f"Could not parse date from {fp}")
            continue
            
        print(f"Processing date: {date_str} from {fp}")
        try:
            data = process_daily_file(fp)
            print("  Data extracted:", data)
        except Exception as e:
            print(f"  Error processing {fp}: {e}")
            continue

        if date_str in existing_dates_res:
             print(f"  Date {date_str} already in Resumen. Skipping.")
        else:
             ws_resumen.append([date_str, data['SR ALMACÉN'], data['SR PUNTOS'], data['TOTAL SR'], data['SRP ALMACÉN'], data['SRP PUNTOS'], data['TOTAL SRP'], data['GRAN TOTAL']])
             print(f"  Added {date_str} to Resumen.")
             existing_dates_res.add(date_str)

        if date_str in existing_dates_det:
             print(f"  Date {date_str} already in Detalle por Jurisdicción. Skipping.")
        else:
             ws_detalle.append([date_str, 'TOTAL GENERAL', data['SR ALMACÉN'], data['SR PUNTOS'], data['SRP ALMACÉN'], data['SRP PUNTOS'], data['TOTAL SR'], data['TOTAL SRP']])
             print(f"  Added {date_str} to Detalle por Jurisdicción.")
             existing_dates_det.add(date_str)
            
    wb.save(db_path)
    print("Database saved.")

if __name__ == "__main__":
    db = sys.argv[1]
    files = sys.argv[2:]
    update_db(db, files)
