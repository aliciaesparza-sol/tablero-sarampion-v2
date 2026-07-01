
# -*- coding: utf-8 -*-
"""
Actualiza SOLO la tabla de RESUMEN MUNICIPIOS en el tablero de Sarampión
Corrige la columna Nom. 2026 y recalcula Total, Pendientes, Coberturas
Corte: 21 de junio 2026
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys

sys.stdout.reconfigure(encoding='utf-8')

CSV_PATH = r'C:\SRP\SRP-SR-2025_21-06-2026 08-32-01.csv'
XLSX_IN  = r'C:\Users\aicil\.gemini\antigravity-ide\scratch\COBERTURA_SARAMPION_21JUN2026_ACTUALIZADO.xlsx'
XLSX_OUT = r'C:\Users\aicil\.gemini\antigravity-ide\scratch\COBERTURA_SARAMPION_21JUN2026_ACTUALIZADO.xlsx'

# ─── Mapa Excel → CSV ────────────────────────────────────────────────────────
MAPA = {
    'Peñón Blanco':         'PEÃ\x91ON BLANCO',   # encoding latin-1 exacto
    'Santa Clara':          'SANTA CLARA',
    'Guanaceví':            'GUANACEVI',
    'San Bernardo':         'SAN BERNARDO',
    'San Pedro del Gallo':  'SAN PEDRO DEL GALLO',
    'Topia':                'TOPIA',
    'Tepehuanes':           'TEPEHUANES',
    'San Juan de Guadalupe':'SAN JUAN DE GUADALUPE',
    'Canelas':              'CANELAS',
    'Súchil':               'SUCHIL',
    'Otáez':                'OTAEZ',
    'Poanas':               'POANAS',
    'Cuencamé':             'CUENCAME',
    'Nombre de Dios':       'NOMBRE DE DIOS',
    'Indé':                 'INDE',
    'San Dimas':            'SAN DIMAS',
    'Mezquital':            'MEZQUITAL',
    'Pueblo Nuevo':         'PUEBLO NUEVO DGO',
    'General Simón Bolívar':'GENERAL SIMON BOLIVAR',
    'Hidalgo':              'HIDALGO DGO',
    'Mapimí':               'MAPIMI',
    'Coneto de Comonfort':  'CONETO DE COMONFORT',
    'Nazas':                'NAZAS',
    'Canatlán':             'CANATLAN',
    'Ocampo':               'OCAMPO DGO',
    'El Oro':               'ORO EL',
    'Rodeo':                'RODEO',
    'Durango':              'DURANGO',
    'Lerdo':                'LERDO',
    'Pánuco de Coronado':   'PANUCO DE CORONADO',
    'Gómez Palacio':        'GOMEZ PALACIO',
    'Nuevo Ideal':          'NUEVO IDEAL',
    'Tamazula':             'TAMAZULA',
    'Tlahualilo':           'TLAHUALILO',
    'Guadalupe Victoria':   'GUADALUPE VICTORIA DGO',
    'San Luis del Cordero': 'SAN LUIS DEL CORDERO',
    'San Juan del Río':     'SAN JUAN DEL RIO DGO',
    'Vicente Guerrero':     'VICENTE GUERRERO',
    'Santiago Papasquiaro': 'SANTIAGO PAPASQUIARO',
}

# ─── Columnas del tablero por grupo etario ───────────────────────────────────
TODAS_COLS = [
    'SRP 6 A 11 MESES PRIMERA', 'SR 6 A 11 MESES PRIMERA',
    'SRP 1 ANIO  PRIMERA', 'SR 1 ANIO PRIMERA',
    'SRP 18 MESES SEGUNDA', 'SR 18 MESES SEGUNDA',
    'SRP 6 ANIOS PRIMERA', 'SRP 6 ANIOS SEGUNDA', 'SR 6 ANIOS PRIMERA', 'SR 6 ANIOS SEGUNDA',
    'SRP 2 A 5 ANIOS PRIMERA', 'SRP 2 A 5 ANIOS SEGUNDA',
    'SRP 7 A 9 ANIOS PRIMERA', 'SRP 7 A 9 ANIOS SEGUNDA',
    'SRP 10 A 12 ANIOS PRIMERA', 'SRP 10 A 12 ANIOS SEGUNDA',
    'SR 2 A 5 ANIOS PRIMERA', 'SR 2 A 5 ANIOS SEGUNDA',
    'SR 7 A 9 ANIOS PRIMERA', 'SR 7 A 9 ANIOS SEGUNDA',
    'SR 10 A 12 ANIOS PRIMERA', 'SR 10 A 12 ANIOS SEGUNDA',
    'SRP 13 A 19 ANIOS PRIMERA', 'SRP 13 A 19 ANIOS SEGUNDA',
    'SR 13 A 19 ANIOS PRIMERA', 'SR 13 A 19 ANIOS SEGUNDA',
    'SRP 20 A 29 ANIOS PRIMERA', 'SRP 20 A 29 ANIOS SEGUNDA',
    'SRP 30 A 39 ANIOS PRIMERA', 'SRP 30 A 39 ANIOS SEGUNDA',
    'SR 20 A 29 ANIOS PRIMERA', 'SR 20 A 29 ANIOS SEGUNDA',
    'SR 30 A 39 ANIOS PRIMERA', 'SR 30 A 39 ANIOS SEGUNDA',
    'SRP 40 A 49 ANIOS PRIMERA', 'SRP 40 A 49 ANIOS SEGUNDA',
    'SR 40 A 49 ANIOS PRIMERA', 'SR 40 A 49 ANIOS SEGUNDA',
]

def semaforo(cob_pct):
    if cob_pct >= 85:
        return '✅ META CUMPLIDA'
    elif cob_pct >= 60:
        return '🟡 EN PROCESO'
    else:
        return '🔴 CRÍTICO'

def color_sem(texto):
    if '✅' in texto: return '00AA44'
    if '🟡' in texto: return 'FFC000'
    return 'FF0000'

def parse_num(val, default=0):
    if val is None: return default
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip().replace(',','').replace('%','').replace(' ','')
    try: return float(s)
    except: return default

# ─── Leer CSV ────────────────────────────────────────────────────────────────
print('Leyendo CSV...')
df = pd.read_csv(CSV_PATH, encoding='latin-1', low_memory=False)
dur = df[df['ESTADO'].str.upper().str.contains('DURANGO', na=False)].copy()

# Solo temporada 2026
dur26 = dur[dur['Temporada'] == 2026].copy()
print(f'  Registros Durango 2026: {len(dur26)}')

# Normalizar y procesar columnas
dur26['MUN_NORM'] = dur26['MUNICIPIO'].str.upper().str.strip()
cols_ok = [c for c in TODAS_COLS if c in dur26.columns]
for c in cols_ok:
    dur26[c] = pd.to_numeric(dur26[c], errors='coerce').fillna(0)

# Agregar dosis 2026 por municipio
agg26 = dur26.groupby('MUN_NORM')[cols_ok].sum().sum(axis=1)

print()
print('Dosis 2026 por municipio (columnas tablero):')
for mun, v in agg26.sort_index().items():
    print(f'  {mun}: {int(v):,}')

# ─── Abrir Excel ─────────────────────────────────────────────────────────────
print('\nActualizando RESUMEN MUNICIPIOS...')
wb = openpyxl.load_workbook(XLSX_IN)
ws = wb['RESUMEN MUNICIPIOS']

# Estructura de la tabla municipios:
# Col B=2: Municipio | C=3: Universo | D=4: Meta | E=5: Cubos | F=6: Nom2025 | G=7: Nom2026 | H=8: Total | I=9: Pendientes | J=10: Cob Meta | K=11: Cob Univ

# Actualizar fila 2 (corte)
cell2 = ws.cell(row=2, column=2)
if cell2.value:
    nuevo = str(cell2.value)
    # Reemplazar fecha de corte
    import re
    nuevo = re.sub(r'Corte[:\s]*\d+/\d+/\d+', 'Corte: 21/06/2026', nuevo)
    nuevo = re.sub(r'\d+\s+de\s+\w+\s+\d{4}', '21 de junio 2026', nuevo)
    ws.cell(row=2, column=2).value = nuevo

# Fila 1 título
cell1 = ws.cell(row=1, column=2)
if cell1.value and isinstance(cell1.value, str):
    ws.cell(row=1, column=2).value = cell1.value

# Acumuladores para fila TOTAL
tot_cubos = 0
tot_nom25 = 0
tot_nom26 = 0
tot_total = 0
tot_meta  = 0
tot_univ  = 0

# Procesar cada municipio (filas 9 a 47)
for row in range(9, 49):
    mun_val = ws.cell(row=row, column=2).value
    if mun_val is None:
        continue
    mun_str = str(mun_val).strip()

    # Fila TOTAL
    if 'TOTAL' in mun_str.upper():
        # Actualizar al final
        continue

    # Obtener datos actuales
    universo = parse_num(ws.cell(row=row, column=3).value)
    meta     = parse_num(ws.cell(row=row, column=4).value)
    cubos    = parse_num(ws.cell(row=row, column=5).value)
    nom25    = parse_num(ws.cell(row=row, column=6).value)

    # Buscar dosis 2026 del CSV
    csv_key = MAPA.get(mun_str)
    if csv_key is None:
        # Intentar con el mismo nombre en mayúsculas
        csv_key = mun_str.upper()
        print(f'  ⚠ No mapeado: "{mun_str}" - intentando con: {csv_key}')

    nom26_nuevo = int(agg26.get(csv_key, 0))

    # Calcular total y cobertura
    total_dosis = cubos + nom25 + nom26_nuevo
    pend = meta - total_dosis

    if pend <= 0:
        pend_str = f'SUPER +{int(abs(pend)):,}'
    else:
        pend_str = int(pend)

    cob_meta = (total_dosis / meta * 100) if meta > 0 else 0
    cob_univ = (total_dosis / universo * 100) if universo > 0 else 0

    cob_meta_str = f'{cob_meta:.1f}%'
    cob_univ_str = f'{cob_univ:.1f}%'

    # Escribir
    ws.cell(row=row, column=7).value  = nom26_nuevo     # Nom 2026
    ws.cell(row=row, column=8).value  = total_dosis     # Total
    ws.cell(row=row, column=9).value  = pend_str        # Pendientes
    ws.cell(row=row, column=10).value = cob_meta_str    # % Cob Meta
    ws.cell(row=row, column=11).value = cob_univ_str    # % Cob Univ

    # Color cobertura
    for col_cob in [10, 11]:
        c = ws.cell(row=row, column=col_cob)
        if cob_meta >= 85:
            c.fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
            c.font = Font(bold=True, color='276221')
        elif cob_meta >= 60:
            c.fill = PatternFill(fill_type='solid', fgColor='FFEB9C')
            c.font = Font(bold=True, color='9C6500')
        else:
            c.fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
            c.font = Font(bold=True, color='9C0006')

    # Acumular
    tot_cubos += cubos
    tot_nom25 += nom25
    tot_nom26 += nom26_nuevo
    tot_total += total_dosis
    tot_meta  += meta
    tot_univ  += universo

    print(f'  {mun_str}: Nom26={nom26_nuevo:,}, Total={total_dosis:,}, Cob={cob_meta:.1f}%')

# ─── Actualizar fila TOTAL DURANGO ───────────────────────────────────────────
for row in range(9, 50):
    mun_val = ws.cell(row=row, column=2).value
    if mun_val and 'TOTAL' in str(mun_val).upper():
        pend_tot = tot_meta - tot_total
        pend_str = f'SUPER +{int(abs(pend_tot)):,}' if pend_tot <= 0 else int(pend_tot)
        cob_meta_tot = (tot_total / tot_meta * 100) if tot_meta > 0 else 0
        cob_univ_tot = (tot_total / tot_univ * 100) if tot_univ > 0 else 0

        ws.cell(row=row, column=7).value  = int(tot_nom26)
        ws.cell(row=row, column=8).value  = int(tot_total)
        ws.cell(row=row, column=9).value  = pend_str
        ws.cell(row=row, column=10).value = f'{cob_meta_tot:.1f}%'
        ws.cell(row=row, column=11).value = f'{cob_univ_tot:.1f}%'

        print()
        print(f'TOTAL DURANGO: Nom25={int(tot_nom25):,}, Nom26={int(tot_nom26):,}, Total={int(tot_total):,}, Cob={cob_meta_tot:.1f}%')
        break

# ─── Actualizar fila resumen global (fila 5) ─────────────────────────────────
ws.cell(row=5, column=5).value = f'{int(tot_cubos):,}'
ws.cell(row=5, column=6).value = f'{int(tot_nom26):,}'
ws.cell(row=5, column=7).value = int(tot_total)
cob_global = (tot_total / tot_meta * 100) if tot_meta > 0 else 0
ws.cell(row=5, column=8).value = f'{cob_global:.1f}%'  # Cob global

# Actualizar fila 2 con resumen
ws.cell(row=2, column=2).value = (
    f'Universo CONAPO 2026 | Cubos ene-may 2025: {int(tot_cubos):,} dosis | '
    f'Nominal jun 2025-junio 2026 | Corte: 21/06/2026'
)

# ─── Guardar ─────────────────────────────────────────────────────────────────
print(f'\nGuardando: {XLSX_OUT}')
wb.save(XLSX_OUT)
print('Listo!')
