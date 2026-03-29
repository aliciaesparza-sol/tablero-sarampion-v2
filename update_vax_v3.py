import pandas as pd
from openpyxl import load_workbook
import copy

# Paths
csv_path = r'SRP-SR-2028.csv'
excel_path = r'c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\Cobertura_SRP-SR_SSA_Durango_2026_PoblaciónSusceptible_ALCORTE_21_03_2026.xlsx'
output_path = r'c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\Cobertura_SRP-SR_SSA_Durango_2026_PoblaciónSusceptible_ALCORTE_28_03_2026.xlsx'

# Load CSV
df = pd.read_csv(csv_path)
dur = df[df['ESTADO'] == 'DURANGO']

def get_sum(query):
    return int(dur.filter(like=query).fillna(0).sum().sum())

# Calculations for March 21
new_doses = {
    "6 a 11 meses": get_sum('6 A 11 MESES'),
    "1 año": get_sum('1 ANIO'),
    "18 meses": get_sum('18 MESES'),
    "Rezagados 2 a 12 años": get_sum('2 A 5 ANIOS') + get_sum('6 ANIOS') + get_sum('7 A 9 ANIOS') + get_sum('10 A 12 ANIOS'),
    "13 a 19 años": get_sum('13 A 19 ANIOS'),
    "20 a 39 años": get_sum('20 A 29 ANIOS') + get_sum('30 A 39 ANIOS'),
    "40 a 49 años": get_sum('40 A 49 ANIOS')
}

# Load Excel
wb = load_workbook(excel_path)
source_sheet = wb.active
new_sheet_name = "28-03-2026"

# Create new sheet if it doesn't exist, copying from source
if new_sheet_name in wb.sheetnames:
    del wb[new_sheet_name]
new_sheet = wb.copy_worksheet(source_sheet)
new_sheet.title = new_sheet_name

# Row mapping (matches the structure we saw in CSV dump)
mapping = {
    "6 a 11 meses": 5,
    "1 año": 6,
    "18 meses": 7,
    "Rezagados 2 a 12 años": 8,
    "13 a 19 años": 9,
    "20 a 39 años": 10,
    "40 a 49 años": 11
}

# Columns: C=Meta (3), D=Dosis (4), E=Susceptible (5), F=Cobertura (6)
meta_col = 3
dosis_col = 4
susc_col = 5
cob_col = 6
avance_col = 7

total_meta = 0
total_dosis = 0

print("Category | New Doses | %")
for category, row_idx in mapping.items():
    meta_val = new_sheet.cell(row=row_idx, column=meta_col).value or 0
    doses_val = new_doses[category]
    
    # Update row
    new_sheet.cell(row=row_idx, column=dosis_col).value = doses_val
    
    # Recalculate Susceptible: Meta - Doses (min 0)
    susc_val = max(0, meta_val - doses_val)
    new_sheet.cell(row=row_idx, column=susc_col).value = susc_val
    
    # Recalculate Cobertura (%): Doses / Meta * 100
    cob_val = (doses_val / meta_val * 100) if meta_val > 0 else 0
    new_sheet.cell(row=row_idx, column=cob_col).value = round(cob_val, 2)
    
    # Avance Cobertura (usually same as % but decimal 0-1)
    new_sheet.cell(row=row_idx, column=avance_col).value = round(cob_val / 100, 4)
    
    total_meta += meta_val
    total_dosis += doses_val
    print(f"{category:25} | {doses_val:8} | {cob_val:6.2f}%")

# Update TOTAL DURANGO row (Row 12)
new_sheet.cell(row=12, column=meta_col).value = total_meta
new_sheet.cell(row=12, column=dosis_col).value = total_dosis
new_sheet.cell(row=12, column=susc_col).value = max(0, total_meta - total_dosis)
new_sheet.cell(row=12, column=cob_col).value = round((total_dosis / total_meta * 100), 2) if total_meta > 0 else 0
new_sheet.cell(row=12, column=avance_col).value = round((total_dosis / total_meta), 4) if total_meta > 0 else 0

# Update Source in Row 2
source_val = f"Fuente: SRP-SR-2025_28-03-2026.csv | Corte: 28 de marzo 2026 | Actualizado: 29/03/2026"
new_sheet.cell(row=2, column=1).value = source_val

wb.save(output_path)
print(f"\nExcel file updated with new sheet '{new_sheet_name}'.")
print(f"Saved as: {output_path}")
