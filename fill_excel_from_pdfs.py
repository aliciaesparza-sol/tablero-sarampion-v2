
"""
Script: fill_excel_from_pdfs.py
Propósito: Extraer información de estudios epidemiológicos (PDF) del ZIP y
           rellenar los campos vacíos del Excel CASOS_NOTIFICADOS_2026_ACTUALIZADO.xlsx
           a partir del caso 110 en adelante.
"""

import zipfile
import pdfplumber
import io
import os
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

# ── RUTAS ───────────────────────────────────────────────────────────────────
base = os.path.expanduser('~') + '/OneDrive/Escritorio/PVU'

excel_path = None
zip_path = None

for root, dirs, files in os.walk(base):
    for f in files:
        if f == 'CASOS_NOTIFICADOS_2026_ACTUALIZADO.xlsx' and 'EE casos' in root:
            excel_path = os.path.join(root, f)
        if f.endswith('.zip') and 'EE casos' in f:
            zip_path = os.path.join(root, f)

print(f"Excel : {excel_path}")
print(f"ZIP   : {zip_path}")

# ── FUNCIONES DE EXTRACCIÓN ──────────────────────────────────────────────────

def clean(s):
    """Normaliza texto quitando espacios extra."""
    if not s:
        return ""
    return re.sub(r'\s+', ' ', s.strip())


def extract_pdf_data(pdf_bytes):
    """
    Lee un PDF de EFE SINAVE y extrae los campos relevantes para el Excel.
    Retorna un diccionario con los valores encontrados.
    """
    data = {
        'rr_unidad_salud_clues': None,
        'rr_ageb_domicilio': None,
        'rr_tipo_localidad': None,
        'rr_tactica_tipo': None,
        'rr_tactica_fecha': None,
        'rr_manzanas_cubiertas': None,
        'rr_srp_dosis_aplicadas': None,
        'rr_sr_dosis_aplicadas': None,
        'rr_total_dosis_aplicadas': None,
        'rr_poblacion_objetivo_estimada': None,
        'rr_cobertura_alcanzada': None,
        # Campos adicionales que podrían estar vacíos en el Excel
        'antecedente_vacunacion': None,
        'numero_dosis': None,
        'es_indigena': None,
        'es_personal_medico': None,
        'viaje_reciente': None,
        'resultado_laboratorio': None,
        'municipio': None,
        'institucion_notificante': None,
        'rr_clues_unidad': None,   # CLUES de la unidad notificante
        'observaciones': None,
    }

    full_text = ""
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                t = page.extract_text() or ""
                full_text += t + "\n"
    except Exception as e:
        print(f"  ERROR leyendo PDF: {e}")
        return data

    lines = full_text.split("\n")

    # ─── VII Acciones de control ─────────────────────────────────────────────
    # COBERTURA EN EL MUNICIPIO PREVIO AL BLOQUEO: 80 % BLOQUEO Sí No ?
    # INICIO 28/02/2026 TERMINACIÓN01/03/2026 DOSIS: 10 COBERTURA:90 %

    for i, line in enumerate(lines):
        line_c = clean(line)

        # BLOQUEO VACUNAL - buscar sección VII
        if 'INICIO' in line_c and 'TERMINACI' in line_c and 'DOSIS:' in line_c:
            # Extraer fecha de inicio del bloqueo
            m_inicio = re.search(r'INICIO\s+(\d{1,2}/\d{1,2}/\d{4})', line_c)
            if m_inicio:
                data['rr_tactica_fecha'] = m_inicio.group(1)
            # Extraer dosis
            m_dosis = re.search(r'DOSIS:\s*(\d+)', line_c)
            if m_dosis:
                data['rr_total_dosis_aplicadas'] = int(m_dosis.group(1))
                data['rr_srp_dosis_aplicadas'] = int(m_dosis.group(1))
                data['rr_sr_dosis_aplicadas'] = 0
            # Extraer cobertura
            m_cob = re.search(r'COBERTURA:\s*(\d+(?:\.\d+)?)\s*%', line_c)
            if m_cob:
                data['rr_cobertura_alcanzada'] = float(m_cob.group(1)) / 100

        # BLOQUEO Sí/No
        if 'BLOQUEO' in line_c and ('S\u00cd' in line_c.upper() or 'SI' in line_c.upper() or 'S\xcd' in line_c.upper()):
            data['rr_tactica_tipo'] = 'Bloqueo vacunal'

        # COBERTURA EN MUNICIPIO PREVIO
        m_prev = re.search(r'COBERTURA EN EL MUNICIPIO PREVIO.*?:\s*(\d+(?:\.\d+)?)\s*%', line_c, re.IGNORECASE)
        if m_prev:
            data['rr_poblacion_objetivo_estimada'] = None  # la poblacion viene de otro campo

        # CLUES de la unidad notificante (sección II y III)
        m_clues = re.search(r'CLUES:\s*([A-Z0-9]+)', line_c)
        if m_clues:
            data['rr_unidad_salud_clues'] = m_clues.group(1)

        # Tipo de localidad - buscar URBANA / RURAL
        if re.search(r'\bURBANA\b', line_c, re.IGNORECASE):
            if data['rr_tipo_localidad'] is None:
                data['rr_tipo_localidad'] = 'URBANA'
        if re.search(r'\bRURAL\b', line_c, re.IGNORECASE):
            if data['rr_tipo_localidad'] is None:
                data['rr_tipo_localidad'] = 'RURAL'

        # Municipio - sección II
        m_mun = re.search(r'Municipio:\s*([A-ZÁÉÍÓÚÑ\s]+?)(?:\s+Localidad:|\s+Instituci)', line_c, re.IGNORECASE)
        if m_mun and not data['municipio']:
            data['municipio'] = clean(m_mun.group(1))

        # Institución notificante
        m_inst = re.search(r'Instituci[oó]n:\s*(SSA|IMSS|ISSSTE|SEDENA|IMSS-B|PEMEX|DIF|PRIVADA)', line_c, re.IGNORECASE)
        if m_inst and not data['institucion_notificante']:
            data['institucion_notificante'] = m_inst.group(1).upper()

        # Vacunación antisarampión
        if 'VACUNACI' in line_c.upper() and 'ANTISARAMPI' in line_c.upper():
            if re.search(r'S[IÍ]\b', line_c):
                data['antecedente_vacunacion'] = 'Sí'
            elif re.search(r'\bNO\b', line_c):
                data['antecedente_vacunacion'] = 'No'

        # Número de dosis
        m_ndosis = re.search(r'N[ÚU]MERO DE DOSIS:\s*(\d+)', line_c, re.IGNORECASE)
        if m_ndosis and not data['numero_dosis']:
            data['numero_dosis'] = int(m_ndosis.group(1))

        # ¿Es indígena?
        if 'RECONOCE COMO IND' in line_c.upper():
            if re.search(r'S[IÍ]\b', line_c):
                data['es_indigena'] = 'Sí'
            elif re.search(r'\bNO\b', line_c):
                data['es_indigena'] = 'No'

        # Resultado laboratorio
        if 'POSITIVO' in line_c and ('SARAMPION' in line_c.upper() or 'SARAMPI' in line_c.upper() or 'SUERO' in line_c.upper() or 'IGM' in line_c.upper()):
            data['resultado_laboratorio'] = 'POSITIVO'
        
        if 'CLASIFICACI' in line_c.upper() and 'FINAL' in line_c.upper() and 'SARAMPI' in line_c.upper():
            if 'CONFIRMADO' in line_c.upper():
                data['resultado_laboratorio'] = 'POSITIVO'
            elif 'EN ESTUDIO' in line_c.upper():
                if not data['resultado_laboratorio']:
                    data['resultado_laboratorio'] = 'EN ESTUDIO'

        # Observaciones (sección X)
        if 'Observaciones:' in line_c:
            # Recoger las siguientes líneas
            obs_lines = []
            for j in range(i+1, min(i+5, len(lines))):
                nl = clean(lines[j])
                if nl and not nl.startswith('XI'):
                    obs_lines.append(nl)
            if obs_lines:
                data['observaciones'] = ' '.join(obs_lines)

    # Asegurar tipo de táctica si hay bloqueo pero no había bandera explícita
    if data['rr_tactica_fecha'] and not data['rr_tactica_tipo']:
        data['rr_tactica_tipo'] = 'Bloqueo vacunal'

    return data


# ── LEER EL EXCEL (OPENPYXL para preservar formato) ─────────────────────────
print("\nCargando Excel (puede tardar unos segundos)...")
wb = load_workbook(excel_path)
ws = wb.active

# Identificar columnas por encabezado
header_row = 1
col_map = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(header_row, col).value
    if val:
        col_map[str(val).strip()] = col

print(f"Columnas de interés encontradas:")
cols_needed = [
    'numero_seriado', 'nombre', 'sexo', 'edad',
    'antecedente_vacunacion', 'numero_dosis',
    'fecha_dosis_1', 'fecha_dosis_2', 'fecha_dosis_3', 'fecha_ultima_dosis',
    'tipo_vacuna', 'es_indigena', 'es_personal_medico', 'viaje_reciente',
    'fecha_inicio_fiebre', 'fecha_inicio_exantema', 'resultado_laboratorio',
    'direccion', 'colonia', 'codigo_postal', 'municipio',
    'institucion_notificante', 'institucion_responsabilidad',
    'rr_unidad_salud_clues', 'rr_ageb_domicilio', 'rr_tipo_localidad',
    'rr_tactica_tipo', 'rr_tactica_fecha',
    'rr_manzanas_cubiertas', 'rr_srp_dosis_aplicadas', 'rr_sr_dosis_aplicadas',
    'rr_total_dosis_aplicadas', 'rr_poblacion_objetivo_estimada',
    'rr_cobertura_alcanzada', 'Observaciones'
]

for c in cols_needed:
    if c in col_map:
        print(f"  OK '{c}' -> col {col_map[c]}")
    else:
        print(f"  XX '{c}' NO ENCONTRADA")

# ── OBTENER ARCHIVOS PDF DEL ZIP POR NÚMERO DE CASO ─────────────────────────
print("\nIndexando PDFs del ZIP...")
pdf_index = {}  # {numero_caso: nombre_archivo_en_zip}

with zipfile.ZipFile(zip_path, 'r') as z:
    for fname in z.namelist():
        m = re.search(r'/(\d+)\.\s+', fname)
        if m:
            num = int(m.group(1))
            pdf_index[num] = fname

print(f"PDFs encontrados: {len(pdf_index)}")
print(f"Rango: {min(pdf_index)} - {max(pdf_index)}")

# ── PROCESAR FILAS DEL EXCEL DESDE CASO 110 ──────────────────────────────────
print("\nProcesando filas (caso 110 en adelante)...")

col_num_ser = col_map.get('numero_seriado', 1)

updated_rows = 0
skipped_rows = 0
not_found_pdfs = []

with zipfile.ZipFile(zip_path, 'r') as z:
    for row in range(2, ws.max_row + 1):  # fila 2 en adelante (fila 1 = encabezados)
        cell_num = ws.cell(row, col_num_ser).value
        if not cell_num:
            continue

        # Extraer número del caso del valor en la columna numero_seriado
        # Formato típico: "111.62061.AMDA" o "110.60438.NSN"
        num_str = str(cell_num).strip()
        m = re.match(r'^(\d+)\.', num_str)
        if not m:
            continue

        caso_num = int(m.group(1))

        # Solo procesar desde caso 110 en adelante
        if caso_num < 110:
            skipped_rows += 1
            continue

        # Buscar PDF correspondiente
        if caso_num not in pdf_index:
            print(f"  ⚠ Caso {caso_num}: PDF no encontrado en ZIP")
            not_found_pdfs.append(caso_num)
            continue

        pdf_name = pdf_index[caso_num]
        print(f"  → Procesando caso {caso_num} ({pdf_name.split('/')[-1]})...")

        # Leer PDF
        try:
            pdf_bytes = z.read(pdf_name)
        except Exception as e:
            print(f"    ERROR leyendo ZIP entry: {e}")
            continue

        # Extraer datos
        data = extract_pdf_data(pdf_bytes)

        # ── RELLENAR CELDAS VACÍAS ────────────────────────────────────────────
        def fill_if_empty(col_name, value):
            if col_name not in col_map:
                return
            c = col_map[col_name]
            current = ws.cell(row, c).value
            if (current is None or str(current).strip() == '' or str(current).strip() in ['NaN', 'nan', '0.0'] or current == 0) and value is not None:
                ws.cell(row, c).value = value

        fill_if_empty('rr_unidad_salud_clues', data['rr_unidad_salud_clues'])
        fill_if_empty('rr_ageb_domicilio', data['rr_ageb_domicilio'])
        fill_if_empty('rr_tipo_localidad', data['rr_tipo_localidad'])
        fill_if_empty('rr_tactica_tipo', data['rr_tactica_tipo'])
        fill_if_empty('rr_tactica_fecha', data['rr_tactica_fecha'])
        fill_if_empty('rr_manzanas_cubiertas', data['rr_manzanas_cubiertas'])
        fill_if_empty('rr_srp_dosis_aplicadas', data['rr_srp_dosis_aplicadas'])
        fill_if_empty('rr_sr_dosis_aplicadas', data['rr_sr_dosis_aplicadas'])
        fill_if_empty('rr_total_dosis_aplicadas', data['rr_total_dosis_aplicadas'])
        fill_if_empty('rr_poblacion_objetivo_estimada', data['rr_poblacion_objetivo_estimada'])
        fill_if_empty('rr_cobertura_alcanzada', data['rr_cobertura_alcanzada'])
        fill_if_empty('antecedente_vacunacion', data['antecedente_vacunacion'])
        fill_if_empty('numero_dosis', data['numero_dosis'])
        fill_if_empty('es_indigena', data['es_indigena'])
        fill_if_empty('resultado_laboratorio', data['resultado_laboratorio'])
        fill_if_empty('Observaciones', data['observaciones'])

        updated_rows += 1
        print(f"    Datos: CLUES={data['rr_unidad_salud_clues']}, Localidad={data['rr_tipo_localidad']}, "
              f"Táctica={data['rr_tactica_tipo']}, Fecha={data['rr_tactica_fecha']}, "
              f"Dosis={data['rr_total_dosis_aplicadas']}, Cob={data['rr_cobertura_alcanzada']}")

# ── GUARDAR EXCEL ─────────────────────────────────────────────────────────────
output_path = excel_path.replace('.xlsx', '_COMPLETO.xlsx')
print(f"\nGuardando resultado en:\n  {output_path}")
wb.save(output_path)

print(f"\n{'='*60}")
print(f"RESUMEN:")
print(f"  Filas actualizadas  : {updated_rows}")
print(f"  Filas omitidas (<110): {skipped_rows}")
print(f"  PDFs no encontrados : {len(not_found_pdfs)} -> {not_found_pdfs}")
print(f"  Archivo guardado    : {output_path}")
print('='*60)
