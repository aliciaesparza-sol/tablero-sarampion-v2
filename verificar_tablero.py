
# -*- coding: utf-8 -*-
"""
Verificación: compara lo que quedó en el Excel actualizado
contra lo que el CSV realmente dice, para cada pestaña y municipio.
"""
import pandas as pd
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

CSV_PATH  = r'C:\SRP\SRP-SR-2025_21-06-2026 08-32-01.csv'
XLSX_PATH = r'C:\Users\aicil\.gemini\antigravity-ide\scratch\COBERTURA_SARAMPION_21JUN2026_ACTUALIZADO.xlsx'

# ── Mapa Excel → CSV (municipio) ──────────────────────────────────────────────
MAPA = {
    'Peñón Blanco':         'PEÃ\x91ON BLANCO',
    'Santa Clara':          'SANTA CLARA',
    'Guanaceví':            'GUANACEVI',
    'San Bernardo':         'SAN BERNARDO',
    'San Pedro del Gallo':  'SAN PEDRO DEL GALLO',
    'Topia':                'TOPIA',
    'Tepehuanes':           'TEPEHUANES',
    'San Juan de Guadalupe':'SAN JUAN DE GUADALUPE',
    'Canelas':              'CANELAS',
    'Súchil':               'SUCHIL',
    'Otáez':                'OTAEZ',
    'Poanas':               'POANAS',
    'Cuencamé':             'CUENCAME',
    'Nombre de Dios':       'NOMBRE DE DIOS',
    'Indé':                 'INDE',
    'San Dimas':            'SAN DIMAS',
    'Mezquital':            'MEZQUITAL',
    'Pueblo Nuevo':         'PUEBLO NUEVO DGO',
    'General Simón Bolívar':'GENERAL SIMON BOLIVAR',
    'Hidalgo':              'HIDALGO DGO',
    'Mapimí':               'MAPIMI',
    'Coneto de Comonfort':  'CONETO DE COMONFORT',
    'Nazas':                'NAZAS',
    'Canatlán':             'CANATLAN',
    'Ocampo':               'OCAMPO DGO',
    'El Oro':               'ORO EL',
    'Rodeo':                'RODEO',
    'Durango':              'DURANGO',
    'Lerdo':                'LERDO',
    'Pánuco de Coronado':   'PANUCO DE CORONADO',
    'Gómez Palacio':        'GOMEZ PALACIO',
    'Nuevo Ideal':          'NUEVO IDEAL',
    'Tamazula':             'TAMAZULA',
    'Tlahualilo':           'TLAHUALILO',
    'Guadalupe Victoria':   'GUADALUPE VICTORIA DGO',
    'San Luis del Cordero': 'SAN LUIS DEL CORDERO',
    'San Juan del Río':     'SAN JUAN DEL RIO DGO',
    'Vicente Guerrero':     'VICENTE GUERRERO',
    'Santiago Papasquiaro': 'SANTIAGO PAPASQUIARO',
}

# ── Columnas CSV por grupo etario ─────────────────────────────────────────────
GRUPOS_CSV = {
    '6-11 Meses':      ['SRP 6 A 11 MESES PRIMERA', 'SR 6 A 11 MESES PRIMERA'],
    '1 Año':           ['SRP 1 ANIO  PRIMERA', 'SR 1 ANIO PRIMERA'],
    '18 Meses':        ['SRP 18 MESES SEGUNDA', 'SR 18 MESES SEGUNDA'],
    '6 Años':          ['SRP 6 ANIOS PRIMERA', 'SRP 6 ANIOS SEGUNDA',
                        'SR 6 ANIOS PRIMERA',  'SR 6 ANIOS SEGUNDA'],
    'Rezag 2-12 Años': ['SRP 2 A 5 ANIOS PRIMERA', 'SRP 2 A 5 ANIOS SEGUNDA',
                        'SRP 7 A 9 ANIOS PRIMERA',  'SRP 7 A 9 ANIOS SEGUNDA',
                        'SRP 10 A 12 ANIOS PRIMERA','SRP 10 A 12 ANIOS SEGUNDA',
                        'SR 2 A 5 ANIOS PRIMERA',   'SR 2 A 5 ANIOS SEGUNDA',
                        'SR 7 A 9 ANIOS PRIMERA',   'SR 7 A 9 ANIOS SEGUNDA',
                        'SR 10 A 12 ANIOS PRIMERA',  'SR 10 A 12 ANIOS SEGUNDA'],
    '13-19 Años':      ['SRP 13 A 19 ANIOS PRIMERA', 'SRP 13 A 19 ANIOS SEGUNDA',
                        'SR 13 A 19 ANIOS PRIMERA',   'SR 13 A 19 ANIOS SEGUNDA'],
    '20-39 Años':      ['SRP 20 A 29 ANIOS PRIMERA', 'SRP 20 A 29 ANIOS SEGUNDA',
                        'SRP 30 A 39 ANIOS PRIMERA', 'SRP 30 A 39 ANIOS SEGUNDA',
                        'SR 20 A 29 ANIOS PRIMERA',  'SR 20 A 29 ANIOS SEGUNDA',
                        'SR 30 A 39 ANIOS PRIMERA',  'SR 30 A 39 ANIOS SEGUNDA'],
    '40-49 Años':      ['SRP 40 A 49 ANIOS PRIMERA', 'SRP 40 A 49 ANIOS SEGUNDA',
                        'SR 40 A 49 ANIOS PRIMERA',  'SR 40 A 49 ANIOS SEGUNDA'],
}

# ── Estructura de columnas por hoja ──────────────────────────────────────────
HOJA_CFG = {
    '6-11 Meses':      {'col_mun':2, 'col_nominal':7},
    '1 Año':           {'col_mun':2, 'col_nominal':7},
    '18 Meses':        {'col_mun':2, 'col_nominal':7},
    '6 Años':          {'col_mun':1, 'col_nominal':6},
    'Rezag 2-12 Años': {'col_mun':2, 'col_nominal':7},
    '13-19 Años':      {'col_mun':2, 'col_nominal':7},
    '20-39 Años':      {'col_mun':2, 'col_nominal':7},
    '40-49 Años':      {'col_mun':2, 'col_nominal':7},
}

# ── Leer CSV ──────────────────────────────────────────────────────────────────
print('Leyendo CSV...')
df = pd.read_csv(CSV_PATH, encoding='latin-1', low_memory=False)
dur = df[df['ESTADO'].str.upper().str.contains('DURANGO', na=False)].copy()
dur['MUN_NORM'] = dur['MUNICIPIO'].str.strip()

# Pre-calcular dosis del CSV por grupo y municipio (SIN filtrar por temporada:
# el CSV ya contiene datos Jun2025-Jun2026, es la fuente única de nominal)
dosis_csv = {}
for grp, cols in GRUPOS_CSV.items():
    ok_cols = [c for c in cols if c in dur.columns]
    for c in ok_cols:
        dur[c] = pd.to_numeric(dur[c], errors='coerce').fillna(0)
    agg = dur.groupby('MUN_NORM')[ok_cols].sum().sum(axis=1)
    dosis_csv[grp] = agg

# ── Leer Excel actualizado ────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

print()
print('=' * 80)
print('VERIFICACIÓN: CSV vs EXCEL ACTUALIZADO')
print('=' * 80)

errores = []

for sname, cfg in HOJA_CFG.items():
    if sname not in wb.sheetnames:
        print(f'\n⚠ Hoja "{sname}" no existe en el Excel')
        continue

    ws = wb[sname]
    cm  = cfg['col_mun']
    cn  = cfg['col_nominal']
    agg = dosis_csv.get(sname, {})

    print(f'\n{"─"*70}')
    print(f'HOJA: {sname}')
    print(f'{"─"*70}')
    print(f'{"Municipio":<28} {"CSV":>8} {"Excel":>8} {"Diferencia":>10} {"Estado"}')
    print(f'{"─"*28} {"─"*8} {"─"*8} {"─"*10} {"─"*15}')

    total_csv   = 0
    total_excel = 0

    for row in range(5, ws.max_row + 1):
        mun_val = ws.cell(row=row, column=cm).value
        if mun_val is None:
            continue
        mun_str = str(mun_val).strip()
        if 'TOTAL' in mun_str.upper():
            continue

        # Valor en Excel
        excel_val = ws.cell(row=row, column=cn).value
        try:
            excel_num = int(float(str(excel_val).replace(',',''))) if excel_val is not None else 0
        except:
            excel_num = 0

        # Valor del CSV
        csv_key = MAPA.get(mun_str)
        if csv_key:
            csv_num = int(agg.get(csv_key, 0))
        else:
            csv_num = -999   # no mapeado

        diff = excel_num - csv_num
        total_csv   += max(csv_num, 0)
        total_excel += excel_num

        if csv_num == -999:
            estado = '⚠ SIN MAPEO'
        elif diff == 0:
            estado = '✅ OK'
        else:
            estado = f'❌ DIFF={diff:+,}'
            errores.append(f'{sname} | {mun_str}: CSV={csv_num:,} Excel={excel_num:,} Diff={diff:+,}')

        csv_disp = f'{csv_num:,}' if csv_num >= 0 else 'N/A'
        print(f'{mun_str:<28} {csv_disp:>8} {excel_num:>8,} {diff if csv_num>=0 else 0:>+10,} {estado}')

    print(f'{"TOTAL":<28} {total_csv:>8,} {total_excel:>8,} {total_excel-total_csv:>+10,}')

print()
print('=' * 80)
if errores:
    print(f'DISCREPANCIAS ENCONTRADAS: {len(errores)}')
    for e in errores:
        print(f'  {e}')
else:
    print('✅ TODOS LOS VALORES COINCIDEN EXACTAMENTE CON EL CSV')
print('=' * 80)
