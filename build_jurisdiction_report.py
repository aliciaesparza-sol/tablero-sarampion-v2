import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# Load raw CSV
csv_path = r'c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\REPORTE_SRP-SR-CENSIA\SRP-SR-2025_14-03-2026 06-41-16.csv'
df = pd.read_csv(csv_path, usecols=[
    'JURISDICCION', 'Fecha de registro',
    'SR PRIMERA TOTAL', 'SR SEGUNDA TOTAL',
    'SRP  PRIMERA TOTAL', 'SRP SEGUNDA TOTAL'
])

# Parse dates
df['Fecha'] = pd.to_datetime(df['Fecha de registro'], errors='coerce')
df['Fecha_str'] = df['Fecha'].dt.strftime('%d/%m/%Y')

# Only keep rows from 2026 (current campaign)
df = df[df['Fecha'].dt.year >= 2026]

# Compute total SR and SRP
df['SR TOTAL'] = df[['SR PRIMERA TOTAL', 'SR SEGUNDA TOTAL']].fillna(0).sum(axis=1)
df['SRP TOTAL'] = df[['SRP  PRIMERA TOTAL', 'SRP SEGUNDA TOTAL']].fillna(0).sum(axis=1)

# Normalize jurisdiction names
df['JURISDICCION'] = df['JURISDICCION'].astype(str).str.strip().str.upper()

# Group by date + jurisdiction
summary = df.groupby(['Fecha_str', 'Fecha', 'JURISDICCION']).agg(
    SR=('SR TOTAL', 'sum'),
    SRP=('SRP TOTAL', 'sum')
).reset_index()
summary['TOTAL'] = summary['SR'] + summary['SRP']
summary = summary.sort_values(['Fecha', 'JURISDICCION'])

# Build pivot: dates as rows, jurisdictions as columns
jurisdictions = sorted(summary['JURISDICCION'].unique())
dates = summary.drop_duplicates('Fecha_str')[['Fecha', 'Fecha_str']].sort_values('Fecha')

# Build an Excel workbook
wb = Workbook()

# ---- Sheet 1: SR por Jurisdicción por Día ----
ws1 = wb.active
ws1.title = "SR por Jurisdicción"

header_fill = PatternFill("solid", fgColor="1F3864")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill = PatternFill("solid", fgColor="DCE6F1")
total_fill = PatternFill("solid", fgColor="BDD7EE")
total_font = Font(bold=True, size=11)

# Row 1: Title
ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(jurisdictions)+3)
title_cell = ws1.cell(row=1, column=1, value="DOSIS SR APLICADAS POR JURISDICCIÓN Y DÍA - CAMPAÑA SARAMPIÓN 2026")
title_cell.font = Font(bold=True, color="FFFFFF", size=13)
title_cell.fill = PatternFill("solid", fgColor="1F3864")
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 25

# Row 2: column headers
ws1.cell(row=2, column=1, value="FECHA").fill = header_fill
ws1.cell(row=2, column=1).font = header_font
ws1.cell(row=2, column=1).alignment = Alignment(horizontal='center')

for i, jur in enumerate(jurisdictions):
    c = ws1.cell(row=2, column=i+2, value=jur)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

total_col = len(jurisdictions) + 2
ws1.cell(row=2, column=total_col, value="TOTAL").fill = header_fill
ws1.cell(row=2, column=total_col).font = header_font
ws1.cell(row=2, column=total_col).alignment = Alignment(horizontal='center')
ws1.row_dimensions[2].height = 30

# Data rows
sr_pivot = summary.pivot_table(index='Fecha_str', columns='JURISDICCION', values='SR', aggfunc='sum', fill_value=0)
srp_pivot = summary.pivot_table(index='Fecha_str', columns='JURISDICCION', values='SRP', aggfunc='sum', fill_value=0)
total_pivot = summary.pivot_table(index='Fecha_str', columns='JURISDICCION', values='TOTAL', aggfunc='sum', fill_value=0)

date_order = dates['Fecha_str'].tolist()

for row_idx, date_str in enumerate(date_order):
    r = row_idx + 3
    fill = alt_fill if row_idx % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")
    ws1.cell(row=r, column=1, value=date_str).fill = fill
    ws1.cell(row=r, column=1).font = Font(bold=False)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    
    row_total = 0
    for i, jur in enumerate(jurisdictions):
        val = int(sr_pivot.loc[date_str, jur]) if date_str in sr_pivot.index and jur in sr_pivot.columns else 0
        c = ws1.cell(row=r, column=i+2, value=val)
        c.fill = fill
        c.alignment = Alignment(horizontal='center')
        row_total += val
    
    tot_cell = ws1.cell(row=r, column=total_col, value=row_total)
    tot_cell.fill = total_fill
    tot_cell.font = total_font
    tot_cell.alignment = Alignment(horizontal='center')

# Grand total row
last_r = len(date_order) + 3
ws1.cell(row=last_r, column=1, value="TOTAL GENERAL").fill = total_fill
ws1.cell(row=last_r, column=1).font = total_font
ws1.cell(row=last_r, column=1).alignment = Alignment(horizontal='center')
grand_total = 0
for i, jur in enumerate(jurisdictions):
    col_total = int(sr_pivot[jur].sum()) if jur in sr_pivot.columns else 0
    c = ws1.cell(row=last_r, column=i+2, value=col_total)
    c.fill = total_fill
    c.font = total_font
    c.alignment = Alignment(horizontal='center')
    grand_total += col_total
ws1.cell(row=last_r, column=total_col, value=grand_total).fill = PatternFill("solid", fgColor="1F3864")
ws1.cell(row=last_r, column=total_col).font = Font(bold=True, color="FFFFFF", size=11)
ws1.cell(row=last_r, column=total_col).alignment = Alignment(horizontal='center')

# Column widths
ws1.column_dimensions['A'].width = 14
for i in range(1, len(jurisdictions)+2):
    ws1.column_dimensions[get_column_letter(i+1)].width = 20

# ---- Sheet 2: SRP por Jurisdicción por Día ----
ws2 = wb.create_sheet("SRP por Jurisdicción")

ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(jurisdictions)+3)
title_cell2 = ws2.cell(row=1, column=1, value="DOSIS SRP APLICADAS POR JURISDICCIÓN Y DÍA - CAMPAÑA SARAMPIÓN 2026")
title_cell2.font = Font(bold=True, color="FFFFFF", size=13)
title_cell2.fill = PatternFill("solid", fgColor="375623")
title_cell2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 25

header_fill2 = PatternFill("solid", fgColor="375623")
total_fill2 = PatternFill("solid", fgColor="E2EFDA")

ws2.cell(row=2, column=1, value="FECHA").fill = header_fill2
ws2.cell(row=2, column=1).font = header_font
ws2.cell(row=2, column=1).alignment = Alignment(horizontal='center')

for i, jur in enumerate(jurisdictions):
    c = ws2.cell(row=2, column=i+2, value=jur)
    c.fill = header_fill2
    c.font = header_font
    c.alignment = Alignment(horizontal='center', wrap_text=True)

ws2.cell(row=2, column=total_col, value="TOTAL").fill = header_fill2
ws2.cell(row=2, column=total_col).font = header_font
ws2.cell(row=2, column=total_col).alignment = Alignment(horizontal='center')
ws2.row_dimensions[2].height = 30

for row_idx, date_str in enumerate(date_order):
    r = row_idx + 3
    fill = alt_fill if row_idx % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")
    ws2.cell(row=r, column=1, value=date_str).fill = fill
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal='center')
    
    row_total = 0
    for i, jur in enumerate(jurisdictions):
        val = int(srp_pivot.loc[date_str, jur]) if date_str in srp_pivot.index and jur in srp_pivot.columns else 0
        c = ws2.cell(row=r, column=i+2, value=val)
        c.fill = fill
        c.alignment = Alignment(horizontal='center')
        row_total += val

    tot_cell = ws2.cell(row=r, column=total_col, value=row_total)
    tot_cell.fill = total_fill2
    tot_cell.font = total_font
    tot_cell.alignment = Alignment(horizontal='center')

last_r2 = len(date_order) + 3
ws2.cell(row=last_r2, column=1, value="TOTAL GENERAL").fill = total_fill2
ws2.cell(row=last_r2, column=1).font = total_font
ws2.cell(row=last_r2, column=1).alignment = Alignment(horizontal='center')
grand_total2 = 0
for i, jur in enumerate(jurisdictions):
    col_total = int(srp_pivot[jur].sum()) if jur in srp_pivot.columns else 0
    c = ws2.cell(row=last_r2, column=i+2, value=col_total)
    c.fill = total_fill2
    c.font = total_font
    c.alignment = Alignment(horizontal='center')
    grand_total2 += col_total
ws2.cell(row=last_r2, column=total_col, value=grand_total2).fill = PatternFill("solid", fgColor="375623")
ws2.cell(row=last_r2, column=total_col).font = Font(bold=True, color="FFFFFF", size=11)
ws2.cell(row=last_r2, column=total_col).alignment = Alignment(horizontal='center')

ws2.column_dimensions['A'].width = 14
for i in range(1, len(jurisdictions)+2):
    ws2.column_dimensions[get_column_letter(i+1)].width = 20

# Save
out_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\EXISTENCIAS SRP Y SR\DOSIS_APLICADAS_POR_JURISDICCION.xlsx"
wb.save(out_path)
print(f"Report saved to: {out_path}")
print(f"Dates in report: {len(date_order)} days")
print(f"Jurisdictions: {jurisdictions}")
