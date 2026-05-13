import openpyxl
import pandas as pd

excel_file = r'c:\Users\aicil\OneDrive\Escritorio\PVU\VPH\CAMPAÑA VPH 2025\TABLERO VPH 2025\TABLERO_VPH_10-05-2026.xlsx'
csv_file = r'c:\Descargas_VPH\VPH 10-05-2026 05-36-43.csv'

# Calculate totals from CSV
df = pd.read_csv(csv_file)
total_jur = df.groupby('JURISDICCION')['CANT APLICACIONES'].sum()
total_issste = df[df['INSTITUCION'] == 'ISSSTE']['CANT APLICACIONES'].sum()

# Load Excel
wb = openpyxl.load_workbook(excel_file)
ws_jur = wb['JURISDICCIONES']

# Find Table 2
start_row_t2 = 0
for r in range(8, 20):
    if ws_jur.cell(row=r, column=1).value == 'Jurisdicción' and ws_jur.cell(row=r, column=2).value == 'Exist. Inicial':
        start_row_t2 = r
        break

if start_row_t2 > 0:
    # Row mappings:
    # start_row_t2 + 1 -> N 1 (Durango)
    # start_row_t2 + 2 -> N 2 (Gomez Palacio)
    # start_row_t2 + 3 -> N 3 (Santiago Papasquiaro)
    # start_row_t2 + 4 -> N 4 (Rodeo)
    # start_row_t2 + 5 -> ISSSTE
    # start_row_t2 + 6 -> Total
    
    apps_n1 = int(total_jur.get('DURANGO', 0))
    apps_n2 = int(total_jur.get('GOMEZ PALACIO', 0))
    apps_n3 = int(total_jur.get('SANTIAGO PAPASQUIARO', 0))
    apps_n4 = int(total_jur.get('RODEO', 0))
    apps_issste = int(total_issste)
    
    # Update Applications (Col 3)
    ws_jur.cell(row=start_row_t2+1, column=3).value = apps_n1
    ws_jur.cell(row=start_row_t2+2, column=3).value = apps_n2
    ws_jur.cell(row=start_row_t2+3, column=3).value = apps_n3
    ws_jur.cell(row=start_row_t2+4, column=3).value = apps_n4
    ws_jur.cell(row=start_row_t2+5, column=3).value = apps_issste
    ws_jur.cell(row=start_row_t2+6, column=3).value = apps_n1 + apps_n2 + apps_n3 + apps_n4 + apps_issste
    
    # Format numbers
    for r in range(start_row_t2+1, start_row_t2+7):
        ws_jur.cell(row=r, column=3).number_format = '#,##0'
        
    # Recalculate Exist. Final (Col 4) = Exist. Inicial (Col 2) - Aplicaciones (Col 3)
    for r in range(start_row_t2+1, start_row_t2+7):
        exist_ini_cell = ws_jur.cell(row=r, column=2)
        exist_ini_val = exist_ini_cell.value
        
        if isinstance(exist_ini_val, str):
            exist_ini_val = int(exist_ini_val.replace(',', '').replace(' ', ''))
            
        apps = ws_jur.cell(row=r, column=3).value
        exist_final = exist_ini_val - apps
        
        ws_jur.cell(row=r, column=4).value = exist_final
        ws_jur.cell(row=r, column=4).number_format = '#,##0'

wb.save(excel_file)
print("Updated table successfully.")
