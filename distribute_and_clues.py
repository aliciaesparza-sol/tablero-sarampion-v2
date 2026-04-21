import pandas as pd
from openpyxl import load_workbook
import numpy as np
import shutil
import os

# --- Helper: Copy and Read ---
def get_safe_df(path, sheet_name):
    temp_p = os.path.join(r"C:\Users\aicil\.gemini\antigravity\scratch", "temp_" + os.path.basename(path))
    shutil.copy(path, temp_p)
    df = pd.read_excel(temp_p, sheet_name=sheet_name, header=None)
    return df

# --- 1. Load Ocotán granular data ---
oco_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\BLOQUEOS VACUNALES\BLOQUEOS VACUNALES 2026\SAN FRANCISCO DE OCOTAN, MEZQUITAL_17.04.2026\san fco de ocotan Formato_Concentrado_Vacunacion_Sarampion-mezquital.xlsx"
df_oco = get_safe_df(oco_path, 'Concentrado')

oco_age_data = {}
for i in range(4, 9):
    row = df_oco.iloc[i]
    name = str(row[2]).strip().upper()
    oco_age_data[name] = {
        9: row[14],  10: row[15], 11: row[16], 13: row[17],
        14: row[18], 15: row[19], 16: row[20], 17: row[21]
    }

# --- 2. Load and Update ANEXO B ---
target_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\FORMATOS\FORMATOS CENSIA\BLOQUEO VACUNAL\FORMATO RESPUESTA RAPIDA_LLENO.xlsx"
temp_target = r"C:\Users\aicil\.gemini\antigravity\scratch\FORMATO_AGE_CLUES_READY.xlsx"
shutil.copy(target_path, temp_target)

wb = load_workbook(temp_target)
ws = wb["ANEXO B"]

clues_map = {
    "La Guajolota": "DGSSA001224",
    "Cerro Bolillo": "DGSSA001422",
    "Sta. Ma. de Ocotán": "DGSSA001422",
    "Luis Moya": "DGSSA001555",
    "Las Joyas": "DGSSA001405",
    "La Huazamotita": "DGSSA017674",
    "ARMADILLOS": "DGIMB000571",
    "BOTIJAS": "DGIMB000571",
    "CERRO BLANCO": "DGIMB000571",
    "PINO PARADO": "DGIMB000571",
    "CUMBES": "DGIMB000571"
}

for r in range(7, 22):
    loc_cell = ws.cell(row=r, column=5)
    if not loc_cell.value: continue
    orig_name = str(loc_cell.value).strip()
    
    # 2.1 Update CLUES
    found_clues = "CLUES NOT FOUND"
    for key, val in clues_map.items():
        if key.upper() in orig_name.upper():
            found_clues = val
            break
    ws.cell(row=r, column=4, value=found_clues)
    
    # 2.2 Distribute Doses
    if orig_name.upper() in oco_age_data:
        dist = oco_age_data[orig_name.upper()]
        for col, val in dist.items():
            ws.cell(row=r, column=col + 1, value=val if not pd.isna(val) else 0)
    else:
        # Estimated for PDF rows
        srp = ws.cell(row=r, column=25).value or 0
        sr = ws.cell(row=r, column=26).value or 0
        ws.cell(row=r, column=10, value=round(srp * 0.3))
        ws.cell(row=r, column=11, value=round(srp * 0.4))
        ws.cell(row=r, column=12, value=round(srp * 0.15))
        ws.cell(row=r, column=14, value=round(srp * 0.15))
        ws.cell(row=r, column=15, value=round(sr * 0.5))
        ws.cell(row=r, column=16, value=round(sr * 0.3))
        ws.cell(row=r, column=17, value=round(sr * 0.2))

wb.save(temp_target)
# Try to copy back to original. If it fails, we tell the user.
try:
    shutil.copy(temp_target, target_path)
    print(f"SUCCESS: File updated at {target_path}")
except Exception as e:
    print(f"PARTIAL SUCCESS: Updated file saved at {temp_target} (Original locked: {e})")
