import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import unicodedata
from pathlib import Path
import numpy as np

def normalize_name(name):
    if pd.isna(name): return ""
    name = str(name).strip().upper()
    name = "".join(c for c in unicodedata.normalize('NFD', name) if unicodedata.category(c) != 'Mn')
    name = name.replace('Ñ', 'N').replace('DDEL', 'DEL')
    return name

# Paths
original_file = Path(r"C:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\COBERTURA POR MUNICIPIO Y SEMANA EPIDEMIOLÒGICA\COBERTURA_SARAMPION_POR_MUNICIPIO_2026_13mayo2026.xlsx")
temp_file = Path("temp.xlsx")
pop_path = Path(r"C:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\COBERTURA DE VACUNACIÓN\TABLERO 3\Poblacion_municipio_edad_simple_y_sexo_Mexico_2026_CENJSIA_EGM.xlsx")
doses_6a_path = Path(r"C:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\COBERTURA POR MUNICIPIO Y SEMANA EPIDEMIOLÒGICA\COBERTURAS POR MUNICIPIO SRP Y SR 2025 12M,18M Y 6A POR SEMANA EPIDEMIOLOGICA.xlsx")

# 1. Get Municipality Order from '1 Año'
df_order = pd.read_excel(temp_file, sheet_name='1 Año', header=3)
muni_order_raw = df_order['Municipio'].tolist()
muni_order_norm = [normalize_name(m) for m in muni_order_raw]

# 2. Get Pop 6y
df_pop_raw = pd.read_excel(pop_path, sheet_name='Durango', header=3)
pop_muni_names = df_pop_raw.iloc[0, 1:40].tolist()
pop_values = df_pop_raw.iloc[8, 1:40].tolist() # Row 8 is Age 6
pop_df = pd.DataFrame({'Municipio_Raw': pop_muni_names, 'Universo 2026': pop_values})
pop_df['Muni_Norm'] = pop_df['Municipio_Raw'].apply(normalize_name)

# 3. Get Doses 6y
df_doses_raw = pd.read_excel(doses_6a_path, sheet_name='SE 53', header=5)
doses_df = df_doses_raw.iloc[:, [0, 8]].copy()
doses_df.columns = ['Municipio_Raw', 'Cubos Ene-May 25']
doses_df['Muni_Norm'] = doses_df['Municipio_Raw'].apply(normalize_name)

# 4. Merge
merged = pd.DataFrame({'Muni_Norm': muni_order_norm, 'Municipio_Orig': muni_order_raw})
merged = pd.merge(merged, pop_df[['Muni_Norm', 'Universo 2026']], on='Muni_Norm', how='left')
merged = pd.merge(merged, doses_df[['Muni_Norm', 'Cubos Ene-May 25']], on='Muni_Norm', how='left')

# Fill TOTAL DURANGO
merged.loc[merged['Municipio_Orig'] == 'TOTAL DURANGO', 'Universo 2026'] = merged['Universo 2026'].sum()
merged.loc[merged['Municipio_Orig'] == 'TOTAL DURANGO', 'Cubos Ene-May 25'] = merged['Cubos Ene-May 25'].sum()

# Calculations
merged['% Meta'] = 1.0 # 100%
merged['Meta Sect.'] = merged['Universo 2026']
merged['Nominal Jun25-May26*'] = 0
merged['Total Dosis'] = merged['Cubos Ene-May 25'] + merged['Nominal Jun25-May26*']
merged['Pendientes'] = merged['Meta Sect.'] - merged['Total Dosis']
merged['Cob. vs Meta (%)'] = np.where(merged['Meta Sect.'] > 0, merged['Total Dosis'] / merged['Meta Sect.'], 0)

def get_semaforo_text(row):
    val = row['Cob. vs Meta (%)'] * 100
    if val < 80: return "🔴 CRÍTICO"
    elif val < 95: return "🟡 EN RIESGO"
    else: return "🟢 ALCANZADA"

merged['Semáforo'] = merged.apply(get_semaforo_text, axis=1)

# 5. Add to Excel using openpyxl
wb = openpyxl.load_workbook(temp_file)
if '6 Años' in wb.sheetnames:
    del wb['6 Años']
ws = wb.create_sheet('6 Años')

# Header at row 4
headers = ['Municipio', 'Universo 2026', '% Meta', 'Meta Sect.', 'Cubos Ene-May 25', 'Nominal Jun25-May26*', 'Total Dosis', 'Pendientes', 'Cob. vs Meta (%)', 'Semáforo']
for i in range(3): ws.append([]) # Add 3 empty rows
ws.append(headers)

# Styles
purple_fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
white_font = Font(color='FFFFFF', bold=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Semáforo Fills
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
red_font = Font(color='9C0006', bold=True)
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
yellow_font = Font(color='9C6500', bold=True)
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_font = Font(color='006100', bold=True)

for cell in ws[4]:
    cell.fill = purple_fill
    cell.font = white_font
    cell.alignment = center_align
    cell.border = thin_border

# Data
for index, row in merged.iterrows():
    data_row = [
        row['Municipio_Orig'],
        row['Universo 2026'],
        row['% Meta'],
        row['Meta Sect.'],
        row['Cubos Ene-May 25'],
        row['Nominal Jun25-May26*'],
        row['Total Dosis'],
        row['Pendientes'],
        row['Cob. vs Meta (%)'],
        row['Semáforo']
    ]
    ws.append(data_row)
    
    curr_row = ws.max_row
    semaforo_val = row['Semáforo']
    semaforo_cell = ws.cell(row=curr_row, column=10)
    
    if "CRÍTICO" in semaforo_val:
        semaforo_cell.fill = red_fill
        semaforo_cell.font = red_font
    elif "EN RIESGO" in semaforo_val:
        semaforo_cell.fill = yellow_fill
        semaforo_cell.font = yellow_font
    elif "ALCANZADA" in semaforo_val:
        semaforo_cell.fill = green_fill
        semaforo_cell.font = green_font

# Formatting numbers and columns
for r in range(5, ws.max_row + 1):
    ws.cell(row=r, column=2).number_format = '#,##0'
    ws.cell(row=r, column=3).number_format = '0%'
    ws.cell(row=r, column=4).number_format = '#,##0'
    ws.cell(row=r, column=5).number_format = '#,##0'
    ws.cell(row=r, column=6).number_format = '#,##0'
    ws.cell(row=r, column=7).number_format = '#,##0'
    ws.cell(row=r, column=8).number_format = '#,##0'
    ws.cell(row=r, column=9).number_format = '0.0%'
    
    # Border for data
    for c in range(1, 11):
        cell = ws.cell(row=r, column=c)
        cell.border = thin_border
        if c != 10: # Semáforo has its own font style
            cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')

# Column widths
ws.column_dimensions['A'].width = 25
for c in 'BCDEFGHIJ':
    ws.column_dimensions[c].width = 18

wb.save(original_file)
print(f"Successfully added colored '6 Años' sheet to: {original_file}")
