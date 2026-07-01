import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\aicil\.gemini\antigravity-ide\scratch\temp_cobertura.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet = wb['6-11 Meses']

# Let's inspect rows 5, 6, 7 in Excel (which are the first data rows)
print("6-11 Meses Formulas (data_only=False):")
for r in range(5, 9):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 12)]
    print(f"Row {r}: {row_vals}")

print("\n6-11 Meses Values (data_only=True):")
wb_data = openpyxl.load_workbook(file_path, data_only=True)
sheet_data = wb_data['6-11 Meses']
for r in range(5, 9):
    row_vals = [sheet_data.cell(row=r, column=c).value for c in range(1, 12)]
    print(f"Row {r}: {row_vals}")
