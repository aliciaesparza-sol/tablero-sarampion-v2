import openpyxl
from openpyxl.styles import Font, Alignment

target_file = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\FORMATOS\FORMATOS CENSIA\BLOQUEO VACUNAL\Formato_Concentrado_Vacunacion_Sarampion_ACTUALIZADO.xlsx"

def update_sheet(sheet):
    # Find the current last column (we want to add AFTER the ones we added before, 
    # OR better, find where we started and insert there)
    
    # Let's just append these 3 new specific geographics at the end of what we have
    last_col = sheet.max_column
    for c in range(last_col, 1, -1):
        if sheet.cell(row=2, column=c).value is not None:
            last_col = c
            break
            
    start_col = last_col + 1
    
    new_rubros = [
        "DOMICILIO / DIRECCI\u00d3N", "AGEB", "MANZANAS CUBIERTAS"
    ]
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for i, name in enumerate(new_rubros):
        target_col = start_col + i
        cell = sheet.cell(row=2, column=target_col)
        cell.value = name
        cell.font = bold_font
        cell.alignment = center_align

try:
    wb = openpyxl.load_workbook(target_file)
    for sname in ["Formato bloqueo barrido", "Concentrado"]:
        if sname in wb.sheetnames:
            print(f"Adding extra fields to {sname}...")
            update_sheet(wb[sname])

    wb.save(target_file)
    print("Added geographic fields successfully.")
except Exception as e:
    print(f"Error: {e}")
