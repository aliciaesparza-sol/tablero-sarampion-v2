import openpyxl
import pandas as pd
import shutil
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

input_excel = r'c:\Users\aicil\OneDrive\Escritorio\PVU\VPH\CAMPAÑA VPH 2025\TABLERO VPH 2025\TABLERO_VPH_05-05-2026_3.xlsx'
output_excel = r'c:\Users\aicil\OneDrive\Escritorio\PVU\VPH\CAMPAÑA VPH 2025\TABLERO VPH 2025\TABLERO_VPH_10-05-2026.xlsx'
csv_file = r'c:\Descargas_VPH\VPH 10-05-2026 05-36-43.csv'

shutil.copy2(input_excel, output_excel)

df = pd.read_csv(csv_file)

# ---------------------------------------------------------
# 1. AVANCE SHEET (Filtered by Escolarizada, No Escolarizada, Rezagadas)
# ---------------------------------------------------------
target_pops = ['ESCOLARIZADA', 'NO ESCOLARIZADA', 'REZAGADAS']
df_avance = df[(df['POBLACION OBJETIVO'].isin(target_pops)) & (df['DOSIS'] != 'RECHAZO')]

app_counts = df_avance.groupby(['INSTITUCION', 'JURISDICCION'])['CANT APLICACIONES'].sum().to_dict()
app_inst_counts = df_avance.groupby('INSTITUCION')['CANT APLICACIONES'].sum().to_dict()
schools_visited = df_avance.groupby(['INSTITUCION', 'JURISDICCION'])['CLAVE ESCUELA'].nunique().to_dict()
schools_inst_visited = df_avance.groupby('INSTITUCION')['CLAVE ESCUELA'].nunique().to_dict()

wb = openpyxl.load_workbook(output_excel)
ws = wb['AVANCE AL 05-MAY-2026']
ws.title = 'AVANCE AL 10-MAY-2026'

ws['A1'] = '💉 AVANCE DE VACUNACIÓN VPH — CORTE AL 10 DE MAYO 2026'
ws['A2'] = 'Fuente: Sistema nominal VPH · Dosis aplicadas confirmadas · Corte al 10 de mayo 2026'

def get_semaforo(cob):
    if cob < 0.5: return '🔴 CRÍTICO'
    if cob < 0.75: return '🟡 EN PROCESO'
    if cob < 0.95: return '🟢 EN META'
    return '🟢 EXCELENTE'

row_map = {
    10: ('SSA', 'DURANGO'),
    11: ('SSA', 'GOMEZ PALACIO'),
    12: ('SSA', 'SANTIAGO PAPASQUIARO'),
    13: ('SSA', 'RODEO'),
    14: ('IMSS', None),
    15: ('IMSS Bienestar', None),
    16: ('ISSSTE', None),
    17: ('SEDENA', None)
}

for r, (inst, jur) in row_map.items():
    if jur:
        apps = app_counts.get((inst, jur), 0)
        schools = schools_visited.get((inst, jur), 0)
    else:
        apps = app_inst_counts.get(inst, 0)
        schools = schools_inst_visited.get(inst, 0)
    
    ws.cell(row=r, column=5).value = int(apps)
    ws.cell(row=r, column=11).value = int(schools)
    
    meta = ws.cell(row=r, column=4).value
    if isinstance(meta, (int, float)):
        faltantes = max(0, meta - apps)
        ws.cell(row=r, column=7).value = int(faltantes)
        if meta > 0:
            cob = apps / meta
            ws.cell(row=r, column=8).value = float(cob)
            ws.cell(row=r, column=9).value = get_semaforo(cob)
            
    asignadas = ws.cell(row=r, column=10).value
    if isinstance(asignadas, (int, float)):
        sin_visita = max(0, asignadas - schools)
        ws.cell(row=r, column=12).value = int(sin_visita)

ssa_apps = app_inst_counts.get('SSA', 0)
ssa_schools = schools_inst_visited.get('SSA', 0)
ws.cell(row=9, column=5).value = int(ssa_apps)
ws.cell(row=9, column=11).value = int(ssa_schools)
meta_ssa = ws.cell(row=9, column=4).value
ws.cell(row=9, column=7).value = max(0, meta_ssa - ssa_apps)
cob_ssa = ssa_apps / meta_ssa
ws.cell(row=9, column=8).value = float(cob_ssa)
ws.cell(row=9, column=9).value = get_semaforo(cob_ssa)
asig_ssa = ws.cell(row=9, column=10).value
ws.cell(row=9, column=12).value = max(0, asig_ssa - ssa_schools)

tot_apps = int(df_avance['CANT APLICACIONES'].sum())
tot_schools = sum(schools_inst_visited.values())
ws.cell(row=18, column=5).value = tot_apps
ws.cell(row=18, column=11).value = int(tot_schools)
meta_tot = ws.cell(row=18, column=4).value
ws.cell(row=18, column=7).value = max(0, meta_tot - tot_apps)
cob_tot = tot_apps / meta_tot
ws.cell(row=18, column=8).value = float(cob_tot)
ws.cell(row=18, column=9).value = get_semaforo(cob_tot)
asig_tot = ws.cell(row=18, column=10).value
ws.cell(row=18, column=12).value = max(0, asig_tot - tot_schools)

# Update cell D4 (or wherever 24816 was) with new total
for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
    for cell in row:
        if cell.value == 24816 or cell.value == 24947:
            cell.value = tot_apps

# ---------------------------------------------------------
# 2. JURISDICCIONES SHEET (SSA Only, All Populations)
# ---------------------------------------------------------
if 'JURISDICCIONES' in wb.sheetnames:
    del wb['JURISDICCIONES']
ws_jur = wb.create_sheet('JURISDICCIONES')

# Filter for JURISDICCIONES
df_ssa = df[(df['INSTITUCION'] == 'SSA') & (df['DOSIS'] != 'RECHAZO')]

ws_jur['A1'] = 'Dosis aplicadas en total por Jurisdicción y Grupo (CENSIA)'
crosstab = pd.crosstab(df_ssa['JURISDICCION'], df_ssa['POBLACION OBJETIVO'], values=df_ssa['CANT APLICACIONES'], aggfunc='sum', margins=True, margins_name='Total').fillna(0)

# Reorder correctly to not miss any group
original_columns = ['ESCOLARIZADA', 'HOMBRES CIS QUE VIVEN CON VIH', 'HOMBRES TRANS QUE VIVEN CON VIH',
                    'MUJERES CIS QUE VIVEN CON VIH', 'MUJERES TRANS QUE VIVEN CON VIH',
                    'MUJERES VICTIMAS DE VIOLACION', 'NO ESCOLARIZADA', 'REZAGADAS']
for col in original_columns:
    if col not in crosstab.columns:
        crosstab[col] = 0
crosstab = crosstab[original_columns + ['Total']]

# Write Table 1 Headers
ws_jur.cell(row=3, column=1, value='JURISDICCIÓN')
for col_idx, col_name in enumerate(crosstab.columns, start=2):
    ws_jur.cell(row=3, column=col_idx, value=col_name)

# Write Table 1 Data
row_map_jur = {'DURANGO': 4, 'GOMEZ PALACIO': 5, 'RODEO': 6, 'SANTIAGO PAPASQUIARO': 7, 'Total': 8}
for idx, row_data in crosstab.iterrows():
    r = row_map_jur.get(idx)
    if r:
        ws_jur.cell(row=r, column=1, value=idx)
        for c_idx, col_name in enumerate(crosstab.columns, start=2):
            ws_jur.cell(row=r, column=c_idx, value=int(row_data[col_name]))

# Write Table 2 Headers
start_row_t2 = 11
ws_jur.cell(row=start_row_t2-2, column=1, value='Dosis recibidas por CENSIA y Existencias')

table_data = [
    ['Jurisdicción', 'Exist. Inicial', 'Aplicaciones', 'Exist. Final'],
    ['N° 1', '6,560', 0, 0],
    ['N° 2', '4,880', 0, 0],
    ['N° 3', '1,790', 0, 0],
    ['N° 4', '920', 0, 0],
    ['ISSSTE', '110', 0, 0],
    ['Total', '14,260', 0, 0]
]

apps_n1 = int(crosstab.loc['DURANGO', 'Total'])
apps_n2 = int(crosstab.loc['GOMEZ PALACIO', 'Total'])
apps_n3 = int(crosstab.loc['SANTIAGO PAPASQUIARO', 'Total'])
apps_n4 = int(crosstab.loc['RODEO', 'Total'])
apps_issste = int(df[(df['INSTITUCION'] == 'ISSSTE') & (df['DOSIS'] != 'RECHAZO')]['CANT APLICACIONES'].sum())

table_data[1][2] = apps_n1
table_data[2][2] = apps_n2
table_data[3][2] = apps_n3
table_data[4][2] = apps_n4
table_data[5][2] = apps_issste
table_data[6][2] = apps_n1 + apps_n2 + apps_n3 + apps_n4 + apps_issste

for r_idx, row in enumerate(table_data, start=start_row_t2):
    exist_ini_val = int(str(row[1]).replace(',', '')) if r_idx > start_row_t2 else row[1]
    
    if r_idx > start_row_t2:
        row[3] = exist_ini_val - row[2]
        
    for c_idx, val in enumerate(row, start=1):
        if c_idx == 2 and r_idx > start_row_t2:
            val = exist_ini_val
        ws_jur.cell(row=r_idx, column=c_idx, value=val)

# ---------------------------------------------------------
# 3. APPLY FORMATTING TO JURISDICCIONES
# ---------------------------------------------------------
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Table 1 Formats
for col in range(1, 11):
    c = ws_jur.cell(row=3, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center_align
    c.border = thin_border
for row in range(4, 9):
    for col in range(1, 11):
        c = ws_jur.cell(row=row, column=col)
        c.border = thin_border
        if row == 8:
            c.fill = total_fill
            c.font = bold_font
        if col == 1:
            c.alignment = Alignment(horizontal="right" if row == 8 else "left", vertical="center", wrap_text=True)
        else:
            c.alignment = center_align
            c.number_format = '#,##0'

# Table 2 Formats
for col in range(1, 5):
    c = ws_jur.cell(row=start_row_t2, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center_align
    c.border = thin_border
for row in range(start_row_t2+1, start_row_t2+7):
    for col in range(1, 5):
        c = ws_jur.cell(row=row, column=col)
        c.border = thin_border
        if row == start_row_t2+6:
            c.fill = total_fill
            c.font = bold_font
        if col == 1:
            c.alignment = Alignment(horizontal="right" if row == start_row_t2+6 else "left", vertical="center", wrap_text=True)
        else:
            c.alignment = center_align
            c.number_format = '#,##0'

ws_jur['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws_jur.cell(row=start_row_t2-2, column=1).font = Font(bold=True, size=14, color="1F4E78")
ws_jur.sheet_view.showGridLines = False

for col in ws_jur.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except: pass
    ws_jur.column_dimensions[column].width = min((max_length + 2), 30)

wb.save(output_excel)
print("Dashboard rebuilt successfully.")
