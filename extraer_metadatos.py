import openpyxl
import json
import unicodedata

def normalize_text(text):
    if not text: return ""
    text = str(text).strip().upper()
    return unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('utf-8')

wb = openpyxl.load_workbook("temp_excel2.xlsx", data_only=True)

sheet_to_tab = {
    'Resumen General': 'resumen',
    '🍼 6-11 Meses': 'g611',
    '👶 1 Año': 'g1',
    '🧒 18 Meses': 'g18',
    '📚 Rezago 2-12': 'grez',
    '🎓 13-19 Años': 'g1319',
    '🧑 20-39 Años': 'g2039',
    '👩 40-49 Años': 'g4049'
}

datos = {}

for sheet_name, tab_id in sheet_to_tab.items():
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    datos[tab_id] = {}
    
    is_resumen = (sheet_name == 'Resumen General')
    
    for row in range(9, ws.max_row + 1):
        mun_cell = ws.cell(row=row, column=1).value
        mun_name_raw = str(mun_cell).strip()
        if not mun_cell or mun_name_raw.upper() in ['TOTAL GENERAL', 'TOTAL']:
            continue
        
        mun_norm = normalize_text(mun_name_raw)
        
        try:
            universo = int(ws.cell(row=row, column=2).value or 0)
        except: universo = 0
            
        if is_resumen:
            try: meta_sect = int(ws.cell(row=row, column=3).value or 0)
            except: meta_sect = 0
            
            try: cubos = int(ws.cell(row=row, column=4).value or 0)
            except: cubos = 0
                
            pct_meta = f"{round((meta_sect/universo*100))}%" if universo else "0%"
        else:
            pct_meta = str(ws.cell(row=row, column=3).value or "0%")
            if pct_meta.replace(".","").isdigit():
                pct_meta = f"{int(float(pct_meta)*100)}%" if float(pct_meta) < 2 else f"{pct_meta}%"
            try: meta_sect = round(float(ws.cell(row=row, column=4).value or 0))
            except: meta_sect = 0
            try: cubos = int(ws.cell(row=row, column=5).value or 0)
            except: cubos = 0
                
        datos[tab_id][mun_norm] = {
            "name": mun_name_raw, # KEEP ORIGINAL ACCENTED NAME FOR DISPLAY
            "u": universo,
            "p": pct_meta,
            "m": meta_sect,
            "c": cubos
        }

with open("metadatos_completos.py", "w", encoding="utf-8") as f:
    f.write("DATOS_EXCEL = " + json.dumps(datos, ensure_ascii=False, indent=2))
print("Meta extraccion COMPLETADA SIN ACENTOS EN LLAVES.")
