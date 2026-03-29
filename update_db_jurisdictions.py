import pandas as pd
import sys
import re
import os
from openpyxl import load_workbook

def parse_date_from_filename(filename):
    basename = os.path.basename(filename)
    m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', basename)
    if m:
        return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
    return None

def process_daily_file(filepath):
    df = pd.read_excel(filepath, sheet_name="Existencias en puntos")
    df.columns = ['Biológico', 'Almacén', 'Puntos', 'Extra']
    
    results = {}
    
    # Get global totals first (top of the file)
    sr_global = df[df['Biológico'].astype(str).str.contains(r'SR\s*\(Doble viral', na=False, regex=True)].iloc[0]
    srp_global = df[df['Biológico'].astype(str).str.contains(r'SRP\s*\(Triple viral', na=False, regex=True)].iloc[0]
    
    total_sr_almacen = int(sr_global['Almacén']) if pd.notna(sr_global['Almacén']) else 0
    total_sr_puntos = int(sr_global['Puntos']) if pd.notna(sr_global['Puntos']) else 0
    total_srp_almacen = int(srp_global['Almacén']) if pd.notna(srp_global['Almacén']) else 0
    total_srp_puntos = int(srp_global['Puntos']) if pd.notna(srp_global['Puntos']) else 0
    
    results['TOTAL GENERAL'] = {
        'SR ALMACÉN': total_sr_almacen, 'SR PUNTOS': total_sr_puntos, 'TOTAL SR': total_sr_almacen + total_sr_puntos,
        'SRP ALMACÉN': total_srp_almacen, 'SRP PUNTOS': total_srp_puntos, 'TOTAL SRP': total_srp_almacen + total_srp_puntos,
        'GRAN TOTAL': total_sr_almacen + total_sr_puntos + total_srp_almacen + total_srp_puntos
    }
    
    # Find jurisdiction rows
    current_jurisdiction = None
    
    for _, row in df.iterrows():
        val = str(row['Biológico']).strip()
        if 'JURISDICCIÓN' in val.upper() or 'JURISDICCION' in val.upper():
            if '1' in val: current_jurisdiction = 'JURISDICCIÓN No. 1'
            elif '2' in val: current_jurisdiction = 'JURISDICCIÓN No. 2'
            elif '3' in val: current_jurisdiction = 'JURISDICCIÓN No. 3'
            elif '4' in val: current_jurisdiction = 'JURISDICCIÓN No. 4'
            else: current_jurisdiction = val
            
            if current_jurisdiction not in results:
                results[current_jurisdiction] = {
                    'SR ALMACÉN': 0, 'SR PUNTOS': 0, 'TOTAL SR': 0,
                    'SRP ALMACÉN': 0, 'SRP PUNTOS': 0, 'TOTAL SRP': 0
                }
        elif pd.notna(current_jurisdiction):
            if re.search(r'SR\s*\(Doble viral', val, flags=re.IGNORECASE):
                almacen = int(row['Almacén']) if pd.notna(row['Almacén']) else 0
                puntos = int(row['Puntos']) if pd.notna(row['Puntos']) else 0
                results[current_jurisdiction]['SR ALMACÉN'] = almacen
                results[current_jurisdiction]['SR PUNTOS'] = puntos
                results[current_jurisdiction]['TOTAL SR'] = almacen + puntos
            elif re.search(r'SRP\s*\(Triple viral', val, flags=re.IGNORECASE):
                almacen = int(row['Almacén']) if pd.notna(row['Almacén']) else 0
                puntos = int(row['Puntos']) if pd.notna(row['Puntos']) else 0
                results[current_jurisdiction]['SRP ALMACÉN'] = almacen
                results[current_jurisdiction]['SRP PUNTOS'] = puntos
                results[current_jurisdiction]['TOTAL SRP'] = almacen + puntos
                
    return results

def update_db(db_path, daily_files):
    # Read the current DB logic to avoid duplicates
    wb = load_workbook(db_path)
    if 'Resumen ' in wb.sheetnames:
        ws_resumen = wb['Resumen ']
    else:
        print("Sheet 'Resumen ' not found!")
        return
        
    if 'Detalle por Jurisdicción' in wb.sheetnames:
        ws_detalle = wb['Detalle por Jurisdicción']
    else:
        print("Sheet 'Detalle por Jurisdicción' not found!")
        return

    def fmt_date(d):
        if pd.isna(d): return ""
        if isinstance(d, str):
            m = re.search(r'(\d{2})[/-](\d{2})[/-](\d{4})', d)
            if m: return f"{m.group(1)}/{m.group(2)}/{m.group(3)}"
            elif "00:00:00" in d:
               dt = pd.to_datetime(d, errors='coerce')
               if pd.notna(dt): return dt.strftime('%d/%m/%Y')
            return d.strip()
        else:
            return pd.to_datetime(d).strftime('%d/%m/%Y')

    df_res = pd.read_excel(db_path, sheet_name='Resumen ')
    existing_dates_res = set(df_res['FECHA'].apply(fmt_date)) if 'FECHA' in df_res.columns else set()

    df_det = pd.read_excel(db_path, sheet_name='Detalle por Jurisdicción')
    if 'FECHA' in df_det.columns and 'SECCIÓN' in df_det.columns:
        # Keep track of (date, section) tuples to avoid duplicate jurisdiction rows
        existing_det = set(zip(df_det['FECHA'].apply(fmt_date), df_det['SECCIÓN'].astype(str).str.strip()))
    else:
        existing_det = set()

    for fp in daily_files:
        date_str = parse_date_from_filename(fp)
        if not date_str:
            print(f"Could not parse date from {fp}")
            continue
            
        print(f"Processing date: {date_str} from {fp}")
        try:
            data = process_daily_file(fp)
        except Exception as e:
            print(f"  Error processing {fp}: {e}")
            continue

        if 'TOTAL GENERAL' in data:
            d_tot = data['TOTAL GENERAL']
            if date_str in existing_dates_res:
                 print(f"  Date {date_str} already in Resumen. Skipping.")
            else:
                 ws_resumen.append([date_str, d_tot['SR ALMACÉN'], d_tot['SR PUNTOS'], d_tot['TOTAL SR'], d_tot['SRP ALMACÉN'], d_tot['SRP PUNTOS'], d_tot['TOTAL SRP'], d_tot['GRAN TOTAL']])
                 print(f"  Added {date_str} to Resumen.")
                 existing_dates_res.add(date_str)
        
        # Append all mapped sections (Jurisdictions + Total General)
        for section, d_sec in data.items():
            if (date_str, section) in existing_det:
                print(f"  Row {(date_str, section)} already in Detalle por Jurisdicción. Skipping.")
            else:
                ws_detalle.append([date_str, section, d_sec['SR ALMACÉN'], d_sec['SR PUNTOS'], d_sec['SRP ALMACÉN'], d_sec['SRP PUNTOS'], d_sec['TOTAL SR'], d_sec['TOTAL SRP']])
                print(f"  Added {(date_str, section)} to Detalle por Jurisdicción.")
                existing_det.add((date_str, section))
            
    wb.save(db_path)
    print("Database saved.")

if __name__ == "__main__":
    db = sys.argv[1]
    files = sys.argv[2:]
    update_db(db, files)
