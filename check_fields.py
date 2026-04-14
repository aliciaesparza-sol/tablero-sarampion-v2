import openpyxl

path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\FORMATOS\FORMATOS CENSIA\BLOQUEO VACUNAL\Formato_Concentrado_Vacunacion_Sarampion_ACTUALIZADO.xlsx"

try:
    wb = openpyxl.load_workbook(path)
    for name in ["Concentrado", "Formato bloqueo barrido"]:
        if name in wb.sheetnames:
            ws = wb[name]
            print(f"Sheet: {name}")
            # Check row 2 (where I added things) and row 5 (sample data row)
            for r in [2, 3, 5]:
                row_vals = [str(c.value).lower() if c.value is not None else "" for c in ws[r]]
                for query in ["ageb", "domicilio", "direcc", "manzana"]:
                    if any(query in v for v in row_vals):
                        print(f"  Found '{query}' in Row {r}")
except Exception as e:
    print(f"Error: {e}")
