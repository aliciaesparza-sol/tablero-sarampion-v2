import openpyxl
import json

wb = openpyxl.load_workbook("temp_excel2.xlsx", data_only=True)

# Mapeo de nombres de hojas a los IDs de los apartados en el HTML
sheet_to_tab = {
    'Resumen General': 'resumen',
    '🍼 6-11 Meses': 'g611',
    '👶 1 Año': 'g1',
    '🧒 18 Meses': 'g18',
    '📚 Rezago 2-12': 'grez',
    '🎓 13-19 Años': 'g1319',
    '🧑 20-39 Años': 'g2039',
    '👩 40-49 Años': 'g4049'
}

metas = {}

for sheet_name, tab_id in sheet_to_tab.items():
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    metas[tab_id] = {}
    
    # Header is at row 8, data starts at row 9
    for row in range(9, ws.max_row + 1):
        mun_cell = ws.cell(row=row, column=1).value
        if not mun_cell or str(mun_cell).strip().upper() in ['TOTAL GENERAL', 'TOTAL']:
            continue
        
        mun_norm = str(mun_cell).strip().upper()
        # Columna 3 is 'Meta Sect.'
        meta_sect = ws.cell(row=row, column=3).value
        try:
            meta_sect = int(meta_sect)
        except:
            meta_sect = 0
            
        metas[tab_id][mun_norm] = meta_sect

with open("metas_dict.py", "w", encoding="utf-8") as f:
    f.write("METAS_MUNICIPIOS = " + json.dumps(metas, ensure_ascii=False, indent=2))
print("Extracted metas successfully.")
