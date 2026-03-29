import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import os

def fmt_date(d):
    try:
        if pd.isna(d): return ""
        if isinstance(d, str):
            try:
                dt = pd.to_datetime(d, dayfirst=True)
                return dt.strftime('%d/%m/%Y')
            except:
                pass
            return d.strip()
        else:
            return pd.to_datetime(d).strftime('%d/%m/%Y')
    except:
        return ""

def main():
    # Load applied doses
    aplicaciones_file = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\REPORTE_SRP-SR-CENSIA\10. DURANGO.xlsx"
    df_aplicaciones = pd.read_excel(aplicaciones_file, sheet_name="Aplicación diaria")
    
    dosis_map = {}
    for _, row in df_aplicaciones.iterrows():
        d = fmt_date(row['Día de aplicación'])
        if d:
            sr_ap = row.iloc[1] # Aplicaciones SR
            if pd.notna(sr_ap):
                dosis_map[d] = float(sr_ap)

    # Load existences
    existences_file = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\EXISTENCIAS SRP Y SR\EXISTENCIAS_SRP_SR_POR_DIA.xlsx"
    df_exist = pd.read_excel(existences_file, sheet_name="Resumen ")
    
    exist_map = {}
    for _, row in df_exist.iterrows():
        d = fmt_date(row['FECHA'])
        if d:
            sr_pts = row.get('SR PUNTOS')
            if pd.notna(sr_pts):
                exist_map[d] = int(sr_pts)

    # Load target testing (as a base)
    target_file = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\EXISTENCIAS SRP Y SR\CONGRUENCIA_SR_PARA_LLENAR_LLENADO.xlsx"
    backup_file = r"C:\Users\aicil\.gemini\antigravity\scratch\CONGRUENCIA_SR_PARA_LLENAR_TEST.xlsx"
    
    wb = openpyxl.load_workbook(backup_file)
    ws = wb.active # Llenado de Dosis Aplicadas
    
    existing_dates = []
    max_row = ws.max_row
    
    for r in range(3, max_row + 1):
        d_val = ws.cell(row=r, column=1).value
        # If openpyxl reads it as datetime it will be a datetime object
        d_str = fmt_date(d_val)
        if not d_str:
            continue
            
        existing_dates.append(d_str)
        
        # Fill Dosis (Column D is 4)
        if d_str in dosis_map:
            ws.cell(row=r, column=4).value = dosis_map[d_str]
        else:
            ws.cell(row=r, column=4).value = 0 # or leave as is
            
    # Now append new dates from exist_map that are not in existing_dates
    # Sort dates by parsing them back
    all_dates = sorted(list(exist_map.keys()), key=lambda x: pd.to_datetime(x, format='%d/%m/%Y'))
    
    current_row = ws.max_row
    if not ws.cell(row=current_row, column=1).value:
        # fallback if last row is empty
        while current_row > 1 and not ws.cell(row=current_row, column=1).value:
            current_row -= 1

    last_appended_row = current_row
    
    for d in all_dates:
        if d not in existing_dates:
            new_row = last_appended_row + 1
            ws.cell(row=new_row, column=1).value = d
            ws.cell(row=new_row, column=2).value = exist_map[d]
            
            # Formula C: B{prev} - B{curr}
            ws.cell(row=new_row, column=3).value = f"=B{new_row-1}-B{new_row}"
            
            # Dosis
            doses = dosis_map.get(d, 0)
            ws.cell(row=new_row, column=4).value = doses
            
            # Formula E: C{curr} - D{curr}
            ws.cell(row=new_row, column=5).value = f"=C{new_row}-D{new_row}"
            
            last_appended_row = new_row
            existing_dates.append(d)

    # Save to the the new file target in the original dir
    wb.save(target_file)
    print("Done. Filled dosis aplicadas and appended missing rows to", target_file)

if __name__ == "__main__":
    main()
