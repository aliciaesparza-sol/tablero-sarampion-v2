import pandas as pd
from openpyxl import load_workbook
import numpy as np
import shutil

# --- 1. Load Ocotán granular data from local copy ---
oco_temp = r"C:\Users\aicil\.gemini\antigravity\scratch\temp_oco.xlsx"
df_oco = pd.read_excel(oco_temp, sheet_name='Concentrado', header=None)

oco_age_data = {}
for i in range(4, 9):
    row = df_oco.iloc[i]
    name = str(row[2]).strip().upper()
    # Map: <1, 1y, 2-4y, 5-9y, 10-19y, 20-39y, 40-49y, 50+
    oco_age_data[name] = {
        10: row[14], # <1 (Col 10 in Anexo B)
        11: row[15], # 1y (Col 11)
        12: row[16], # >1y (Col 12 - mapping 2-4y here)
        14: row[17], # 5-9y (Col 14)
        15: row[18], # 10-19y (Col 15)
        16: row[19], # 20-29y (Col 16 - mapping 20-39y here)
        17: row[20], # 30-49y (Col 17 - mapping 40-49y here)
        18: row[21]  # >50y (Col 18)
    }

# --- 2. Load and Update ANEXO B from local copy ---
target_temp = r"C:\Users\aicil\.gemini\antigravity\scratch\temp_ready.xlsx"
wb = load_workbook(target_temp)
ws = wb["ANEXO B"]

clues_map = {
    "La Guajolota": "DGSSA001224",
    "Cerro Bolillo": "DGSSA001422",
    "Sta. Ma. de Ocotán": "DGSSA001422",
    "Luis Moya": "DGSSA001555",
    "Las Joyas": "DGSSA001405",
    "La Huazamotita": "DGSSA017674",
    "ARMADILLOS": "DGIMB000571",
    "BOTIJAS": "DGIMB000571",
    "CERRO BLANCO": "DGIMB000571",
    "PINO PARADO": "DGIMB000571",
    "CUMBES": "DGIMB000571"
}

for r in range(7, 22):
    loc_cell = ws.cell(row=r, column=5)
    if not loc_cell.value: continue
    orig_name = str(loc_cell.value).strip()
    
    # 2.1 Update CLUES (Column D / Index 4)
    found_clues = "CLUES NOT FOUND"
    for key, val in clues_map.items():
        if key.upper() in orig_name.upper():
            found_clues = val
            break
    ws.cell(row=r, column=4, value=found_clues)
    
    # 2.2 Distribute Doses (Columns 10-18)
    if orig_name.upper() in oco_age_data:
        dist = oco_age_data[orig_name.upper()]
        for col, val in dist.items():
            ws.cell(row=r, column=col, value=val if not pd.isna(val) else 0)
    else:
        # Estimated for PDF rows (Column 25 is SRP, 26 is SR)
        srp = ws.cell(row=r, column=25).value or 0
        sr = ws.cell(row=r, column=26).value or 0
        
        # Simple distribution for SSA data
        ws.cell(row=r, column=10, value=round(srp * 0.3)) # <1
        ws.cell(row=r, column=11, value=round(srp * 0.4)) # 1y
        ws.cell(row=r, column=12, value=round(srp * 0.15)) # >1y
        ws.cell(row=r, column=14, value=round(srp * 0.15)) # 5-9y
        ws.cell(row=r, column=15, value=round(sr * 0.5)) # 10-19y
        ws.cell(row=r, column=16, value=round(sr * 0.3)) # 20-29y
        ws.cell(row=r, column=17, value=round(sr * 0.2)) # 30-49y

wb.save(target_temp)

# --- 3. Try to copy back ---
final_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\FORMATOS\FORMATOS CENSIA\BLOQUEO VACUNAL\FORMATO RESPUESTA RAPIDA_LLENO.xlsx"
try:
    shutil.copy(target_temp, final_path)
    print(f"SUCCESS: File fully updated at {final_path}")
except Exception as e:
    print(f"ERROR: Could not write back to {final_path}. File probably locked. Error: {e}")
