import openpyxl
import pandas as pd
import shutil
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define file paths
input_excel = r'c:\Users\aicil\OneDrive\Escritorio\PVU\VPH\CAMPAÑA VPH 2025\TABLERO VPH 2025\TABLERO_VPH_05-05-2026_3.xlsx'
output_excel = r'c:\Users\aicil\OneDrive\Escritorio\PVU\VPH\CAMPAÑA VPH 2025\TABLERO VPH 2025\TABLERO_VPH_10-05-2026.xlsx'
csv_file = r'c:\Descargas_VPH\VPH 10-05-2026 05-36-43.csv'

# Make a copy
shutil.copy2(input_excel, output_excel)

# Process CSV
df = pd.read_csv(csv_file)

# Calculate applications by institution / jurisdiction
app_counts = df.groupby(['INSTITUCION', 'JURISDICCION'])['CANT APLICACIONES'].sum().to_dict()
app_inst_counts = df.groupby('INSTITUCION')['CANT APLICACIONES'].sum().to_dict()

# Calculate schools visited
schools_visited = df.groupby(['INSTITUCION', 'JURISDICCION'])['CLAVE ESCUELA'].nunique().to_dict()
schools_inst_visited = df.groupby('INSTITUCION')['CLAVE ESCUELA'].nunique().to_dict()

# Mappings for Excel rows
# Row -> (Institution, Jurisdiction)
row_map = {
    10: ('SSA', 'DURANGO'),
    11: ('SSA', 'GOMEZ PALACIO'),
    12: ('SSA', 'SANTIAGO PAPASQUIARO'),
    13: ('SSA', 'RODEO'),
    14: ('IMSS', None),
    15: ('IMSS Bienestar', None),
    16: ('ISSSTE', None),
    17: ('SEDENA', None)
}

# Open workbook
wb = openpyxl.load_workbook(output_excel)

# --- 1. UPDATE AVANCE SHEET ---
ws = wb['AVANCE AL 05-MAY-2026']
ws.title = 'AVANCE AL 10-MAY-2026'

# Update headers
ws['A1'] = '💉 AVANCE DE VACUNACIÓN VPH — CORTE AL 10 DE MAYO 2026'
ws['A2'] = 'Fuente: Sistema nominal VPH · Dosis aplicadas confirmadas · Corte al 10 de mayo 2026'

# Helper function for semaforo
def get_semaforo(cob):
    if cob < 0.5: return '🔴 CRÍTICO'
    if cob < 0.75: return '🟡 EN PROCESO'
    if cob < 0.95: return '🟢 EN META'
    return '🟢 EXCELENTE'

for r, (inst, jur) in row_map.items():
    # Get values
    if jur:
        apps = app_counts.get((inst, jur), 0)
        schools = schools_visited.get((inst, jur), 0)
    else:
        apps = app_inst_counts.get(inst, 0)
        schools = schools_inst_visited.get(inst, 0)
    
    # Col E: Aplicadas
    ws.cell(row=r, column=5).value = int(apps)
    
    # Col K: Escuelas Visitadas
    ws.cell(row=r, column=11).value = int(schools)
    
    # Recalculate Dosis Faltantes (G) = Meta Propia (D) - Aplicadas (E)
    meta = ws.cell(row=r, column=4).value
    if isinstance(meta, (int, float)):
        faltantes = max(0, meta - apps)
        ws.cell(row=r, column=7).value = int(faltantes)
        
        # Recalculate Cobertura (H)
        if meta > 0:
            cob = apps / meta
            ws.cell(row=r, column=8).value = float(cob)
            ws.cell(row=r, column=9).value = get_semaforo(cob)
            
    # Recalculate Escuelas sin visita (L) = Asignadas (J) - Visitadas (K)
    asignadas = ws.cell(row=r, column=10).value
    if isinstance(asignadas, (int, float)):
        sin_visita = max(0, asignadas - schools)
        ws.cell(row=r, column=12).value = int(sin_visita)

# Update SSA Total (Row 9)
ssa_apps = app_inst_counts.get('SSA', 0)
ssa_schools = schools_inst_visited.get('SSA', 0)
ws.cell(row=9, column=5).value = int(ssa_apps)
ws.cell(row=9, column=11).value = int(ssa_schools)
meta_ssa = ws.cell(row=9, column=4).value
ws.cell(row=9, column=7).value = max(0, meta_ssa - ssa_apps)
cob_ssa = ssa_apps / meta_ssa
ws.cell(row=9, column=8).value = float(cob_ssa)
ws.cell(row=9, column=9).value = get_semaforo(cob_ssa)
asig_ssa = ws.cell(row=9, column=10).value
ws.cell(row=9, column=12).value = max(0, asig_ssa - ssa_schools)

# Update Total Estatal (Row 18)
tot_apps = int(df['CANT APLICACIONES'].sum())
tot_schools = sum(schools_inst_visited.values()) # Approx
ws.cell(row=18, column=5).value = tot_apps
ws.cell(row=18, column=11).value = int(tot_schools)
meta_tot = ws.cell(row=18, column=4).value
ws.cell(row=18, column=7).value = max(0, meta_tot - tot_apps)
cob_tot = tot_apps / meta_tot
ws.cell(row=18, column=8).value = float(cob_tot)
ws.cell(row=18, column=9).value = get_semaforo(cob_tot)
asig_tot = ws.cell(row=18, column=10).value
ws.cell(row=18, column=12).value = max(0, asig_tot - tot_schools)

# Update top summary (row 3/4 area, maybe it's D4? Let's assume there is a cell with 24816)
for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
    for cell in row:
        if cell.value == 24816:
            cell.value = tot_apps
        if cell.value == 12042:
            cell.value = max(0, meta_tot - tot_apps)

# --- 2. NEW SHEET: JURISDICCIONES ---
if 'JURISDICCIONES' in wb.sheetnames:
    del wb['JURISDICCIONES']
ws_jur = wb.create_sheet('JURISDICCIONES')

# Add table 1: Poblacion Objetivo by Jurisdiccion
ws_jur['A1'] = 'Dosis aplicadas en total por Jurisdicción y Grupo (CENSIA)'
ws_jur['A1'].font = Font(bold=True, size=14)

crosstab = pd.crosstab(df['JURISDICCION'], df['POBLACION OBJETIVO'], values=df['CANT APLICACIONES'], aggfunc='sum', margins=True, margins_name='Total').fillna(0)

# Write headers
ws_jur.cell(row=3, column=1, value='JURISDICCIÓN').font = Font(bold=True)
for col_idx, col_name in enumerate(crosstab.columns, start=2):
    c = ws_jur.cell(row=3, column=col_idx, value=col_name)
    c.font = Font(bold=True)
    c.alignment = Alignment(wrap_text=True)

# Write data
for row_idx, (idx, row) in enumerate(crosstab.iterrows(), start=4):
    ws_jur.cell(row=row_idx, column=1, value=idx).font = Font(bold=True if idx == 'Total' else False)
    for col_idx, val in enumerate(row, start=2):
        ws_jur.cell(row=row_idx, column=col_idx, value=int(val))

# Add table 2: Dosis Recibidas CENSIA vs Aplicaciones
start_row = 4 + len(crosstab) + 3
ws_jur.cell(row=start_row, column=1, value='Dosis recibidas por CENSIA y Existencias').font = Font(bold=True, size=14)

table_data = [
    ['Jurisdicción', 'Exist. Inicial', 'Aplicaciones', 'Exist. Final'],
    ['N° 1', '6,560', '5,054', '1,506'],
    ['N° 2', '4,880', '4,697', '183'],
    ['N° 3', '1,790', '667', '1,123'],
    ['N° 4', '920', '665 (11 desp.)', '244'],
    ['ISSSTE', '110', '110', '0'],
    ['Total', '14,260', '11,204', '3,056']
]

for r_idx, row in enumerate(table_data, start=start_row+2):
    for c_idx, val in enumerate(row, start=1):
        c = ws_jur.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == start_row+2 or r_idx == start_row+2+len(table_data)-1:
            c.font = Font(bold=True)
            
# Adjust column widths
for col in ws_jur.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_jur.column_dimensions[column].width = min(adjusted_width, 30)

wb.save(output_excel)
print("Dashboard updated successfully!")
