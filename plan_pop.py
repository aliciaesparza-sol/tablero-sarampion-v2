import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\COBERTURA DE VACUNACIÓN\TABLERO 3\Poblacion_municipio_edad_simple_y_sexo_Mexico_2026_CENJSIA_EGM.xlsx", data_only=True)
ws = wb['Durango']

# Identify columns for Durango
mun_col = None
for col in range(2, ws.max_column + 1):
    val = ws.cell(row=4, column=col).value
    if val and str(val).strip().upper() == "DURANGO":
        mun_col = col
        break

if not mun_col:
    print("No encontro columna DURANGO")
    exit()

pop_hombres = {}
pop_mujeres = {}

# Parse rows
current_sex = None
for row in range(5, ws.max_row + 1):
    edad_val = ws.cell(row=row, column=1).value
    if edad_val == "Hombres":
        current_sex = "H"
        continue
    elif edad_val == "Mujeres":
        current_sex = "M"
        continue
        
    if isinstance(edad_val, int):
        pop_val = ws.cell(row=row, column=mun_col).value or 0
        if current_sex == "H":
            pop_hombres[edad_val] = pop_val
        elif current_sex == "M":
            pop_mujeres[edad_val] = pop_val

pop_total = {edad: pop_hombres.get(edad, 0) + pop_mujeres.get(edad, 0) for edad in range(110)}

def sum_pop(start, end):
    return sum(pop_total.get(e, 0) for e in range(start, end+1))

# Let's calculate for Durango:
# 6 a 11 meses (50% de menores de 1 año)
g611_uni = sum_pop(0, 0)
g611_meta = round(g611_uni * 0.5)

# 1 año (100%)
g1_uni = sum_pop(1, 1)
g1_meta = round(g1_uni * 1.0)

# 18 meses (100%)
# Assuming it means 100% of 1 year olds as shown in image where both are 12255
g18_uni = sum_pop(1, 1)
g18_meta = round(g18_uni * 1.0)

# Rezagados 2 a 12 años (50%)
grez_uni = sum_pop(2, 12)
grez_meta = round(grez_uni * 0.5)

# 13 a 19 años (50%)
g1319_uni = sum_pop(13, 19)
g1319_meta = round(g1319_uni * 0.5)

# 20 a 39 años (50%)
g2039_uni = sum_pop(20, 39)
g2039_meta = round(g2039_uni * 0.5)

# 40 a 49 años (50%)
g4049_uni = sum_pop(40, 49)
g4049_meta = round(g4049_uni * 0.5)

total_meta = g1_meta + g18_meta + grez_meta + g1319_meta + g2039_meta + g4049_meta

print("Resultados Durango:")
print(f"6-11 meses: Univ={g611_uni}, Meta={g611_meta}")
print(f"1 año: Univ={g1_uni}, Meta={g1_meta}")
print(f"18 meses: Univ={g18_uni}, Meta={g18_meta}")
print(f"Rezagados 2-12: Univ={grez_uni}, Meta={grez_meta}")
print(f"13-19: Univ={g1319_uni}, Meta={g1319_meta}")
print(f"20-39: Univ={g2039_uni}, Meta={g2039_meta}")
print(f"40-49: Univ={g4049_uni}, Meta={g4049_meta}")
print(f"TOTAL META: {total_meta}")
