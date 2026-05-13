import openpyxl
import pandas as pd

excel_file = r'c:\Users\aicil\OneDrive\Escritorio\PVU\VPH\CAMPAÑA VPH 2025\TABLERO VPH 2025\TABLERO_VPH_10-05-2026.xlsx'
csv_file = r'c:\Descargas_VPH\VPH 10-05-2026 05-36-43.csv'

# Calculate totals from CSV
df = pd.read_csv(csv_file)
ssa_df = df[df['INSTITUCION'] == 'SSA']
ssa_jur = ssa_df.groupby('JURISDICCION')['CANT APLICACIONES'].sum()
issste_total = df[df['INSTITUCION'] == 'ISSSTE']['CANT APLICACIONES'].sum()

# Load Excel
wb = openpyxl.load_workbook(excel_file)
ws_jur = wb['JURISDICCIONES']

# Find Table 2
start_row_t2 = 0
for r in range(8, 20):
    if ws_jur.cell(row=r, column=1).value == 'Jurisdicción' and ws_jur.cell(row=r, column=2).value == 'Exist. Inicial':
        start_row_t2 = r
        break

if start_row_t2 > 0:
    apps_n1 = int(ssa_jur.get('DURANGO', 0))
    apps_n2 = int(ssa_jur.get('GOMEZ PALACIO', 0))
    apps_n3 = int(ssa_jur.get('SANTIAGO PAPASQUIARO', 0))
    apps_n4 = int(ssa_jur.get('RODEO', 0))
    apps_issste = int(issste_total)
    
    # Update Applications (Col 3)
    ws_jur.cell(row=start_row_t2+1, column=3).value = apps_n1
    ws_jur.cell(row=start_row_t2+2, column=3).value = apps_n2
    ws_jur.cell(row=start_row_t2+3, column=3).value = apps_n3
    ws_jur.cell(row=start_row_t2+4, column=3).value = apps_n4
    ws_jur.cell(row=start_row_t2+5, column=3).value = apps_issste
    ws_jur.cell(row=start_row_t2+6, column=3).value = apps_n1 + apps_n2 + apps_n3 + apps_n4 + apps_issste
    
    # Format numbers and Recalculate Exist. Final
    for r in range(start_row_t2+1, start_row_t2+7):
        ws_jur.cell(row=r, column=3).number_format = '#,##0'
        exist_ini_cell = ws_jur.cell(row=r, column=2)
        exist_ini_val = exist_ini_cell.value
        
        if isinstance(exist_ini_val, str):
            exist_ini_val = int(exist_ini_val.replace(',', '').replace(' ', ''))
            
        apps = ws_jur.cell(row=r, column=3).value
        exist_final = exist_ini_val - apps
        
        ws_jur.cell(row=r, column=4).value = exist_final
        ws_jur.cell(row=r, column=4).number_format = '#,##0'

wb.save(excel_file)
print("Updated table successfully.")
