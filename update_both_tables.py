import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

excel_file = r'c:\Users\aicil\OneDrive\Escritorio\PVU\VPH\CAMPAÑA VPH 2025\TABLERO VPH 2025\TABLERO_VPH_10-05-2026.xlsx'
csv_file = r'c:\Descargas_VPH\VPH 10-05-2026 05-36-43.csv'

# Load CSV
df = pd.read_csv(csv_file)

# Filter for SSA only and exclude RECHAZO
df_ssa = df[(df['INSTITUCION'] == 'SSA') & (df['DOSIS'] != 'RECHAZO')]

# 1. Update Table 1: Crosstab
crosstab = pd.crosstab(df_ssa['JURISDICCION'], df_ssa['POBLACION OBJETIVO'], values=df_ssa['CANT APLICACIONES'], aggfunc='sum', margins=True, margins_name='Total').fillna(0)

# Ensure all original columns are present so we don't mess up the headers
original_columns = ['ESCOLARIZADA', 'HOMBRES CIS QUE VIVEN CON VIH', 'HOMBRES TRANS QUE VIVEN CON VIH',
                    'MUJERES CIS QUE VIVEN CON VIH', 'MUJERES TRANS QUE VIVEN CON VIH',
                    'MUJERES VICTIMAS DE VIOLACION', 'NO ESCOLARIZADA', 'REZAGADAS']
for col in original_columns:
    if col not in crosstab.columns:
        crosstab[col] = 0

# Reorder columns to match original + Total
crosstab = crosstab[original_columns + ['Total']]

# Load Excel
wb = openpyxl.load_workbook(excel_file)
ws_jur = wb['JURISDICCIONES']

# Write Table 1 data (rows 4 to 8, cols 2 to 10)
row_map = {
    'DURANGO': 4,
    'GOMEZ PALACIO': 5,
    'RODEO': 6,
    'SANTIAGO PAPASQUIARO': 7,
    'Total': 8
}

for idx, row_data in crosstab.iterrows():
    r = row_map.get(idx)
    if r:
        for c_idx, col_name in enumerate(crosstab.columns, start=2):
            val = int(row_data[col_name])
            cell = ws_jur.cell(row=r, column=c_idx)
            cell.value = val
            cell.number_format = '#,##0'

# 2. Update Table 2: Aplicaciones & Exist. Final
# Find Table 2
start_row_t2 = 0
for r in range(8, 20):
    if ws_jur.cell(row=r, column=1).value == 'Jurisdicción' and ws_jur.cell(row=r, column=2).value == 'Exist. Inicial':
        start_row_t2 = r
        break

if start_row_t2 > 0:
    # Applications from the crosstab 'Total' column
    apps_n1 = int(crosstab.loc['DURANGO', 'Total'])
    apps_n2 = int(crosstab.loc['GOMEZ PALACIO', 'Total'])
    apps_n3 = int(crosstab.loc['SANTIAGO PAPASQUIARO', 'Total'])
    apps_n4 = int(crosstab.loc['RODEO', 'Total'])
    
    # ISSSTE: recalculate excluding RECHAZO
    apps_issste = int(df[(df['INSTITUCION'] == 'ISSSTE') & (df['DOSIS'] != 'RECHAZO')]['CANT APLICACIONES'].sum())
    
    # Update Applications (Col 3)
    ws_jur.cell(row=start_row_t2+1, column=3).value = apps_n1
    ws_jur.cell(row=start_row_t2+2, column=3).value = apps_n2
    ws_jur.cell(row=start_row_t2+3, column=3).value = apps_n3
    ws_jur.cell(row=start_row_t2+4, column=3).value = apps_n4
    ws_jur.cell(row=start_row_t2+5, column=3).value = apps_issste
    ws_jur.cell(row=start_row_t2+6, column=3).value = apps_n1 + apps_n2 + apps_n3 + apps_n4 + apps_issste
    
    # Recalculate Exist. Final (Col 4) = Exist. Inicial (Col 2) - Aplicaciones (Col 3)
    for r in range(start_row_t2+1, start_row_t2+7):
        ws_jur.cell(row=r, column=3).number_format = '#,##0'
        exist_ini_cell = ws_jur.cell(row=r, column=2)
        exist_ini_val = exist_ini_cell.value
        
        if isinstance(exist_ini_val, str):
            exist_ini_val = int(exist_ini_val.replace(',', '').replace(' ', ''))
            
        apps = ws_jur.cell(row=r, column=3).value
        exist_final = exist_ini_val - apps
        
        ws_jur.cell(row=r, column=4).value = exist_final
        ws_jur.cell(row=r, column=4).number_format = '#,##0'

wb.save(excel_file)
print("Updated both tables successfully.")
