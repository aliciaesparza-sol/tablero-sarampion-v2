import openpyxl
from openpyxl.styles import Font, Alignment

target_file = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\FORMATOS\FORMATOS CENSIA\BLOQUEO VACUNAL\Formato_Concentrado_Vacunacion_Sarampion_ACTUALIZADO.xlsx"

def update_sheet(sheet):
    # Find a safe starting column
    # We'll just look for the first column that has nothing in rows 2-10
    start_col = 1
    for c in range(sheet.max_column, 200): # Check up to 200 columns
        is_empty = True
        for r in range(1, 15):
            if sheet.cell(row=r, column=c).value is not None:
                is_empty = False
                break
        if is_empty:
            # Also check if it's merged
            is_merged = False
            for merged_range in sheet.merged_cells.ranges:
                if c >= merged_range.min_col and c <= merged_range.max_col:
                    is_merged = True
                    break
            if not is_merged:
                start_col = c
                break
    
    print(f"Adding columns starting at index {start_col}")
    
    new_rubros = [
        "TIPO DE LOCALIDAD", "URBANA", "SEMIURBANA", "RURAL",
        "POBLACI\u00d3N ESPECIAL", "PERSONAL DE SALUD", "DOCENTE",
        "EDADES ADICIONALES", "20 A 29 A\u00d1OS", "30 A 49 A\u00d1OS", ">= 50 A\u00d1OS",
        "T\u00c1CTICA", "BLOQUEO < 72H", "BARRIDO DOC", "BRIGADA"
    ]
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for i, name in enumerate(new_rubros):
        target_col = start_col + i
        cell = sheet.cell(row=2, column=target_col)
        cell.value = name
        cell.font = bold_font
        cell.alignment = center_align

try:
    wb = openpyxl.load_workbook(target_file)
    for sname in ["Formato bloqueo barrido", "Concentrado"]:
        if sname in wb.sheetnames:
            print(f"Updating {sname}...")
            update_sheet(wb[sname])

    wb.save(target_file)
    print("Updated successfully.")
except Exception as e:
    import traceback
    print(f"Error: {e}")
    traceback.print_exc()
