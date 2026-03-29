import pandas as pd
from openpyxl import load_workbook

# Paths
csv_path = r'c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\COBERTURA DE VACUNACIÓN\TABLERO SARAMPION\SRP-SR-2025_21-03-2026 04-30-33.csv'
excel_path = r'c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\CAMPAÑA SARAMPIÓN 10 SEMANAS\Cobertura_SRP-SR_SSA_Durango_2026.xlsx'

# Load CSV
df = pd.read_csv(csv_path)
dur = df[df['ESTADO'] == 'DURANGO']

def get_sum(query):
    # Sum all columns matching query for Durango (All institutions)
    return dur.filter(like=query).fillna(0).sum().sum()

# Calculations
new_doses = {
    "6 a 11 meses": int(get_sum('6 A 11 MESES')),
    "1 año": int(get_sum('1 ANIO')),
    "18 meses": int(get_sum('18 MESES')),
    "Rezagados 2 a 12 años": int(get_sum('2 A 5 ANIOS') + get_sum('6 ANIOS') + get_sum('7 A 9 ANIOS') + get_sum('10 A 12 ANIOS')),
    "13 a 19 años": int(get_sum('13 A 19 ANIOS')),
    "20 a 39 años": int(get_sum('20 A 29 ANIOS') + get_sum('30 A 39 ANIOS')),
    "40 a 49 años": int(get_sum('40 A 49 ANIOS'))
}

# Update Excel
wb = load_workbook(excel_path)
sheet = wb.active

# Row indices (1-based)
mapping = {
    "6 a 11 meses": 4,
    "1 año": 5,
    "18 meses": 6,
    "Rezagados 2 a 12 años": 7,
    "13 a 19 años": 8,
    "20 a 39 años": 9,
    "40 a 49 años": 10
}

# Columns: C=Meta (3), D=Dosis (4), E=Susceptible (5), F=Cobertura (6)
meta_col = 3
dosis_col = 4
susc_col = 5
cob_col = 6

total_meta = 0
total_dosis = 0

print("Category | Meta | New Doses | %")
for category, row_idx in mapping.items():
    meta = sheet.cell(row=row_idx, column=meta_col).value or 0
    dosis = new_doses[category]
    susc = max(0, meta - dosis)
    cob = (dosis / meta * 100) if meta > 0 else 0
    
    sheet.cell(row=row_idx, column=dosis_col).value = dosis
    sheet.cell(row=row_idx, column=susc_col).value = susc
    sheet.cell(row=row_idx, column=cob_col).value = round(cob, 2)
    
    total_meta += meta
    total_dosis += dosis
    print(f"{category:25} | {meta:8} | {dosis:8} | {cob:6.2f}%")

# Update TOTAL DURANGO row (Row 11)
sheet.cell(row=11, column=meta_col).value = total_meta
sheet.cell(row=11, column=dosis_col).value = total_dosis
sheet.cell(row=11, column=susc_col).value = max(0, total_meta - total_dosis)
sheet.cell(row=11, column=cob_col).value = round((total_dosis / total_meta * 100), 2) if total_meta > 0 else 0

# Update Source in Row 2
source_val = f"Fuente dosis aplicadas: SRP-SR-2025_21-03-2026.csv (SRP-SR Todas Instituciones)  |  Corte: 21 de marzo 2026"
sheet.cell(row=2, column=1).value = source_val

wb.save(excel_path)
print(f"\nExcel file fully updated and saved.")
