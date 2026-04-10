import pandas as pd
import openpyxl
import shutil
import os

excel_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\COBERTURA DE VACUNACIÓN\TABLERO 3\Cobertura_Sarampion_Durango_Municipio_04abril2026.xlsx"
csv_path = r"c:\Descargas_SRP\SRP-SR-2025_08-04-2026 02-31-03.csv"
output_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\COBERTURA DE VACUNACIÓN\TABLERO 3\DATOS_CORREGIDOS_TABLERO_2.xlsx"

# 1. Copiar y cargar
shutil.copy2(excel_path, "temp_wb.xlsx")
wb = openpyxl.load_workbook("temp_wb.xlsx")

# 2. Cargar CSV
df_csv = pd.read_csv(csv_path, sep=";", encoding="latin1", low_memory=False)
df_csv["MUNICIPIO_NORM"] = df_csv["MUNICIPIO"].astype(str).str.strip().str.upper()

# 3. Definir las bolsas de columnas para cada hoja
def get_cols(sheet_name):
    if sheet_name == 'Resumen General':
        return ["SRP  PRIMERA TOTAL", "SRP SEGUNDA TOTAL", "SR PRIMERA TOTAL", "SR SEGUNDA TOTAL"]
    elif '6-11' in sheet_name:
        return ["SRP 6 A 11 MESES PRIMERA", "SR 6 A 11 MESES PRIMERA"]
    elif '1 Año' in sheet_name:
        return ["SRP 1 ANIO  PRIMERA", "SR 1 ANIO PRIMERA", "SRP 18 MESES SEGUNDA", "SR 18 MESES SEGUNDA"]
    elif '18 Meses' in sheet_name:
        return ["SRP 18 MESES SEGUNDA", "SR 18 MESES SEGUNDA"]
    elif '2-12' in sheet_name:
        return ["SRP 2 A 5 ANIOS PRIMERA","SRP 6 ANIOS PRIMERA","SRP 7 A 9 ANIOS PRIMERA",
                "SRP 10 A 12 ANIOS PRIMERA","SR 2 A 5 ANIOS PRIMERA","SR 6 ANIOS PRIMERA",
                "SR 7 A 9 ANIOS PRIMERA","SR 10 A 12 ANIOS PRIMERA",
                "SRP 2 A 5 ANIOS SEGUNDA","SRP 6 ANIOS SEGUNDA","SRP 7 A 9 ANIOS SEGUNDA",
                "SRP 10 A 12 ANIOS SEGUNDA","SR 2 A 5 ANIOS SEGUNDA","SR 6 ANIOS SEGUNDA",
                "SR 7 A 9 ANIOS SEGUNDA","SR 10 A 12 ANIOS SEGUNDA"]
    elif '13-19' in sheet_name:
        return ["SRP 13 A 19 ANIOS PRIMERA","SR 13 A 19 ANIOS PRIMERA", "SRP 13 A 19 ANIOS SEGUNDA","SR 13 A 19 ANIOS SEGUNDA"]
    elif '20-39' in sheet_name:
        return ["SRP 20 A 29 ANIOS PRIMERA","SRP 30 A 39 ANIOS PRIMERA", "SR 20 A 29 ANIOS PRIMERA","SR 30 A 39 ANIOS PRIMERA",
                "SRP 20 A 29 ANIOS SEGUNDA","SRP 30 A 39 ANIOS SEGUNDA", "SR 20 A 29 ANIOS SEGUNDA","SR 30 A 39 ANIOS SEGUNDA"]
    elif '40-49' in sheet_name:
        return ["SRP 40 A 49 ANIOS PRIMERA","SR 40 A 49 ANIOS PRIMERA", "SRP 40 A 49 ANIOS SEGUNDA","SR 40 A 49 ANIOS SEGUNDA"]
    return []

def get_semaforo(val):
    if val >= 0.95: return "✅ META ALCANZADA"
    elif val >= 0.80: return "🟢 BUENA COBERTURA"
    elif val >= 0.50: return "🟡 EN PROCESO"
    elif val >= 0.25: return "🟠 BAJA COBERTURA"
    else: return "🔴 CRÍTICO"

# 4. Procesar hojas
for sheet_name in wb.sheetnames:
    if sheet_name == 'Dosis Semanales': continue
    
    cols_to_sum = get_cols(sheet_name)
    if not cols_to_sum: continue
    cols_exist = [c for c in cols_to_sum if c in df_csv.columns]
    
    df_csv["_tmp_sum"] = df_csv[cols_exist].sum(axis=1) if cols_exist else 0
    dosis_nom = df_csv.groupby("MUNICIPIO_NORM")["_tmp_sum"].sum().to_dict()
    # Aliases
    dosis_nom['PUEBLO NUEVO'] = dosis_nom.get('PUEBLO NUEVO DGO', 0)
    
    ws = wb[sheet_name]
    
    # Cambiar nombre del encabezado en fila 8 col E a "Nominal Jun-Abril"
    if ws.cell(row=8, column=5).value is not None:
        ws.cell(row=8, column=5).value = "Nominal\nJun-Abril"
    if ws.cell(row=8, column=6).value is not None:
        ws.cell(row=8, column=6).value = "Nominal\nOpcional"
        
    for row in range(9, ws.max_row + 1):
        mun_cell = ws.cell(row=row, column=1).value
        if not mun_cell or str(mun_cell).strip() in ['Total general', 'TOTAL']: continue
        
        mun_norm = str(mun_cell).strip().upper()
        
        # Meta y Cubos
        meta_sect = ws.cell(row=row, column=3).value
        try: meta_sect = int(meta_sect)
        except: meta_sect = 0
        
        cubos = ws.cell(row=row, column=4).value
        try: cubos = int(cubos)
        except: cubos = 0
            
        nominal = dosis_nom.get(mun_norm, 0)
        
        total_dosis = cubos + nominal
        pendientes = max(0, meta_sect - total_dosis)
        cob_pct = (total_dosis / meta_sect) if meta_sect > 0 else 0
        
        ws.cell(row=row, column=5).value = nominal
        # limpiar E/F si es necesario o no
        ws.cell(row=row, column=6).value = 0 # clear F just in case it's 2026 data col
        
        ws.cell(row=row, column=7).value = total_dosis
        ws.cell(row=row, column=8).value = pendientes
        ws.cell(row=row, column=9).value = cob_pct
        ws.cell(row=row, column=11).value = get_semaforo(cob_pct)

wb.save(output_path)
print(f"Hecho: {output_path}")
