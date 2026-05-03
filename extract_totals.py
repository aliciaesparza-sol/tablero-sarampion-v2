import openpyxl
import os

# Columnas TOTAL por biológico
biologicos = {
    'BCG': 'R',
    'HEPATITIS B': 'Z',
    'HEXAVALENTE': 'AI',
    'DPT': 'AL',
    'ROTAVIRUS': 'AQ',
    'NEUMOCOCICA 13V': 'AY',
    'NEUMOCOCICA 20V': 'BG',
    'SRP': 'BL',
    'SR': 'BU',
    'VPH': 'CC',
    'VARICELA': 'CF',
    'HEPATITIS A': 'CJ',
    'Td EMBARAZADAS': 'CS',
    'Td 1ra DOSIS': 'CW',
    'Td 2da DOSIS': 'DA',
    'Td 3ra DOSIS': 'DE',
    'Td REFUERZO': 'DI',
    'Tdpa': 'DJ',
    'VSR': 'DM',
    'COVID 5a11': 'DP',
    'COVID 5a59 FR': 'DS',
    'COVID 60mas': 'DV',
    'COVID EMBARAZADAS': 'DY',
    'COVID PERSONAL SALUD': 'EB',
    'COVID 6a11m': 'EE',
    'COVID 12a59m': 'EH',
    'INFLUENZA': 'EX',
    'INFLUENZA RIESGO': 'GD',
}

path = r'C:\Users\aicil\OneDrive\Escritorio\PVU\SEMANA NACIONAL DE SALUD PUBLICA'
files = [
    'H I NUEVO IDEAL FORMATO _REPORTE_DIARIO_SNV_27 04 2026(1).xlsx',
    'FORMATO _REPORTE_DIARIO_SNV_2026 OJITO.xlsx',
    'FORMATO _REPORTE_DIARIO_SNV_2026 HIC.xlsx',
    'CS LAJAS REPORTE_DIARIO_SNV_2026.xlsx',
    'CS DTBG REPORTE_DIARIO_SNV 2026.xlsx',
    'FORMATO _REPORTE_DIARIO_SNV_2026 JUR 4.xlsx',
    'SNV 2026 JSN2 TERCERO. REPORTE.xlsx',
    'FORMATO _REPORTE_DIARIO_SNV_2026 JS1 DURANGO 28-04-26.xlsx',
]

totales = {bio: 0 for bio in biologicos}
detalle = {}

for fname in files:
    fpath = os.path.join(path, fname)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    sheet_name = None
    for sn in wb.sheetnames:
        if 'REGISTRO' in sn.upper():
            sheet_name = sn
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[-1]
    ws = wb[sheet_name]
    
    # Buscar la fila que dice TOTAL
    total_row = None
    for r in range(9, 15):
        cell_val = ws.cell(row=r, column=1).value
        if cell_val and 'TOTAL' in str(cell_val).upper():
            total_row = r
            break
    
    if total_row is None:
        print(f'WARN: No se encontro fila TOTAL en {fname}')
        continue
    
    arch_totales = {}
    for bio, col in biologicos.items():
        val = ws[f'{col}{total_row}'].value
        if val is not None and isinstance(val, (int, float)):
            arch_totales[bio] = int(val)
            totales[bio] += int(val)
        else:
            arch_totales[bio] = 0
    
    short = fname.replace('.xlsx', '')
    detalle[short] = arch_totales
    print(f'OK: {short}')

# Agrupar
td_keys = ['Td EMBARAZADAS', 'Td 1ra DOSIS', 'Td 2da DOSIS', 'Td 3ra DOSIS', 'Td REFUERZO']
td_total = sum(totales[k] for k in td_keys)

covid_keys = ['COVID 5a11', 'COVID 5a59 FR', 'COVID 60mas', 'COVID EMBARAZADAS', 'COVID PERSONAL SALUD', 'COVID 6a11m', 'COVID 12a59m']
covid_total = sum(totales[k] for k in covid_keys)

influenza_total = totales['INFLUENZA'] + totales['INFLUENZA RIESGO']

print()
print('=' * 55)
print('  DOSIS TOTALES POR BIOLOGICO - SNV 2026')
print('  (8 archivos consolidados)')
print('=' * 55)

resumen = [
    ('BCG', totales['BCG']),
    ('HEPATITIS B', totales['HEPATITIS B']),
    ('HEXAVALENTE', totales['HEXAVALENTE']),
    ('DPT', totales['DPT']),
    ('ROTAVIRUS', totales['ROTAVIRUS']),
    ('NEUMOCOCICA 13V', totales['NEUMOCOCICA 13V']),
    ('NEUMOCOCICA 20V', totales['NEUMOCOCICA 20V']),
    ('SRP', totales['SRP']),
    ('SR', totales['SR']),
    ('VPH', totales['VPH']),
    ('VARICELA', totales['VARICELA']),
    ('HEPATITIS A', totales['HEPATITIS A']),
    ('Td (TOTAL)', td_total),
    ('Tdpa', totales['Tdpa']),
    ('VSR', totales['VSR']),
    ('COVID-19 (TOTAL)', covid_total),
    ('INFLUENZA (TOTAL)', influenza_total),
]

gran_total = 0
for bio, val in resumen:
    print(f'  {bio:25s} {val:>8,}')
    gran_total += val

print('  ' + '-' * 35)
print(f'  {"GRAN TOTAL":25s} {gran_total:>8,}')

print()
print('  --- Detalle Td ---')
for k in td_keys:
    print(f'    {k:20s} {totales[k]:>6,}')

print()
print('  --- Detalle COVID-19 ---')
for k in covid_keys:
    print(f'    {k:25s} {totales[k]:>6,}')

print()
print('  --- Detalle Influenza ---')
print(f'    {"Poblacion Blanco":25s} {totales["INFLUENZA"]:>6,}')
print(f'    {"Riesgo 5-59":25s} {totales["INFLUENZA RIESGO"]:>6,}')

# Detalle por archivo
print()
print('=' * 55)
print('  DETALLE POR ARCHIVO')
print('=' * 55)
for archivo, dat in detalle.items():
    short_name = archivo.split('SNV')[0].strip().rstrip('_').rstrip()
    if not short_name:
        short_name = archivo[:40]
    total_archivo = sum(dat.values())
    print(f'\n  {archivo[:50]}')
    print(f'  Total dosis: {total_archivo:,}')
