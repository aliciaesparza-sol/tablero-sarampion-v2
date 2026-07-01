
# -*- coding: utf-8 -*-
"""
Revisión detallada: para cada pestaña y cada municipio,
muestra exactamente qué dice el CSV (columna por columna)
y qué quedó escrito en el Excel, con fechas reales del CSV.
"""
import pandas as pd
import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

CSV_PATH  = r'C:\SRP\SRP-SR-2025_21-06-2026 08-32-01.csv'
XLSX_PATH = r'C:\Users\aicil\.gemini\antigravity-ide\scratch\COBERTURA_SARAMPION_21JUN2026_ACTUALIZADO.xlsx'

# ── Leer CSV completo ─────────────────────────────────────────────────────────
df = pd.read_csv(CSV_PATH, encoding='latin-1', low_memory=False)
dur = df[df['ESTADO'].str.upper().str.contains('DURANGO', na=False)].copy()

# Revisar rango real de fechas del CSV
dur['Fecha'] = pd.to_datetime(dur['Fecha de registro'], errors='coerce')
print('=== RANGO DE FECHAS EN EL CSV ===')
print(f'  Fecha mínima: {dur["Fecha"].min()}')
print(f'  Fecha máxima: {dur["Fecha"].max()}')
print(f'  Registros 2025: {len(dur[dur["Temporada"]==2025])}')
print(f'  Registros 2026: {len(dur[dur["Temporada"]==2026])}')
print(f'  Registros TOTAL: {len(dur)}')

# ── Mapa municipio Excel → CSV ────────────────────────────────────────────────
MAPA = {
    'Peñón Blanco':          'PEÃ\x91ON BLANCO',
    'Santa Clara':           'SANTA CLARA',
    'Guanaceví':             'GUANACEVI',
    'San Bernardo':          'SAN BERNARDO',
    'San Pedro del Gallo':   'SAN PEDRO DEL GALLO',
    'Topia':                 'TOPIA',
    'Tepehuanes':            'TEPEHUANES',
    'San Juan de Guadalupe': 'SAN JUAN DE GUADALUPE',
    'Canelas':               'CANELAS',
    'Súchil':                'SUCHIL',
    'Otáez':                 'OTAEZ',
    'Poanas':                'POANAS',
    'Cuencamé':              'CUENCAME',
    'Nombre de Dios':        'NOMBRE DE DIOS',
    'Indé':                  'INDE',
    'San Dimas':             'SAN DIMAS',
    'Mezquital':             'MEZQUITAL',
    'Pueblo Nuevo':          'PUEBLO NUEVO DGO',
    'General Simón Bolívar': 'GENERAL SIMON BOLIVAR',
    'Hidalgo':               'HIDALGO DGO',
    'Mapimí':                'MAPIMI',
    'Coneto de Comonfort':   'CONETO DE COMONFORT',
    'Nazas':                 'NAZAS',
    'Canatlán':              'CANATLAN',
    'Ocampo':                'OCAMPO DGO',
    'El Oro':                'ORO EL',
    'Rodeo':                 'RODEO',
    'Durango':               'DURANGO',
    'Lerdo':                 'LERDO',
    'Pánuco de Coronado':    'PANUCO DE CORONADO',
    'Gómez Palacio':         'GOMEZ PALACIO',
    'Nuevo Ideal':           'NUEVO IDEAL',
    'Tamazula':              'TAMAZULA',
    'Tlahualilo':            'TLAHUALILO',
    'Guadalupe Victoria':    'GUADALUPE VICTORIA DGO',
    'San Luis del Cordero':  'SAN LUIS DEL CORDERO',
    'San Juan del Río':      'SAN JUAN DEL RIO DGO',
    'Vicente Guerrero':      'VICENTE GUERRERO',
    'Santiago Papasquiaro':  'SANTIAGO PAPASQUIARO',
}

# ── Columnas CSV exactas por grupo etario ─────────────────────────────────────
GRUPOS_CSV = {
    '6-11 Meses':      ['SRP 6 A 11 MESES PRIMERA', 'SR 6 A 11 MESES PRIMERA'],
    '1 Año':           ['SRP 1 ANIO  PRIMERA', 'SR 1 ANIO PRIMERA'],
    '18 Meses':        ['SRP 18 MESES SEGUNDA', 'SR 18 MESES SEGUNDA'],
    '6 Años':          ['SRP 6 ANIOS PRIMERA', 'SRP 6 ANIOS SEGUNDA',
                        'SR 6 ANIOS PRIMERA',  'SR 6 ANIOS SEGUNDA'],
    'Rezag 2-12 Años': ['SRP 2 A 5 ANIOS PRIMERA',  'SRP 2 A 5 ANIOS SEGUNDA',
                        'SRP 7 A 9 ANIOS PRIMERA',   'SRP 7 A 9 ANIOS SEGUNDA',
                        'SRP 10 A 12 ANIOS PRIMERA', 'SRP 10 A 12 ANIOS SEGUNDA',
                        'SR 2 A 5 ANIOS PRIMERA',    'SR 2 A 5 ANIOS SEGUNDA',
                        'SR 7 A 9 ANIOS PRIMERA',    'SR 7 A 9 ANIOS SEGUNDA',
                        'SR 10 A 12 ANIOS PRIMERA',  'SR 10 A 12 ANIOS SEGUNDA'],
    '13-19 Años':      ['SRP 13 A 19 ANIOS PRIMERA', 'SRP 13 A 19 ANIOS SEGUNDA',
                        'SR 13 A 19 ANIOS PRIMERA',   'SR 13 A 19 ANIOS SEGUNDA'],
    '20-39 Años':      ['SRP 20 A 29 ANIOS PRIMERA', 'SRP 20 A 29 ANIOS SEGUNDA',
                        'SRP 30 A 39 ANIOS PRIMERA',  'SRP 30 A 39 ANIOS SEGUNDA',
                        'SR 20 A 29 ANIOS PRIMERA',   'SR 20 A 29 ANIOS SEGUNDA',
                        'SR 30 A 39 ANIOS PRIMERA',   'SR 30 A 39 ANIOS SEGUNDA'],
    '40-49 Años':      ['SRP 40 A 49 ANIOS PRIMERA', 'SRP 40 A 49 ANIOS SEGUNDA',
                        'SR 40 A 49 ANIOS PRIMERA',   'SR 40 A 49 ANIOS SEGUNDA'],
}

HOJA_CFG = {
    '6-11 Meses':      {'col_mun':2, 'col_nominal':7},
    '1 Año':           {'col_mun':2, 'col_nominal':7},
    '18 Meses':        {'col_mun':2, 'col_nominal':7},
    '6 Años':          {'col_mun':1, 'col_nominal':6},
    'Rezag 2-12 Años': {'col_mun':2, 'col_nominal':7},
    '13-19 Años':      {'col_mun':2, 'col_nominal':7},
    '20-39 Años':      {'col_mun':2, 'col_nominal':7},
    '40-49 Años':      {'col_mun':2, 'col_nominal':7},
}

# Pre-procesar columnas numéricas
all_dose_cols = list(set(c for cols in GRUPOS_CSV.values() for c in cols))
for c in all_dose_cols:
    if c in dur.columns:
        dur[c] = pd.to_numeric(dur[c], errors='coerce').fillna(0)

dur['MUN_NORM'] = dur['MUNICIPIO'].str.strip()

# Pre-calcular dosis por grupo y municipio (TODOS los registros del CSV)
dosis_csv_por_grp = {}
for grp, cols in GRUPOS_CSV.items():
    ok = [c for c in cols if c in dur.columns]
    agg = dur.groupby('MUN_NORM')[ok].sum().sum(axis=1)
    dosis_csv_por_grp[grp] = agg

# ── Abrir Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

print()
print('=' * 90)
print('AUDITORÍA COMPLETA: DOSIS EN EXCEL vs CSV  (corte 21/06/2026)')
print('Fuente CSV: Jun 2025 – Jun 2026 (todos los registros del archivo)')
print('=' * 90)

discrepancias = []

for sname, cfg in HOJA_CFG.items():
    if sname not in wb.sheetnames:
        continue

    ws  = wb[sname]
    cm  = cfg['col_mun']
    cn  = cfg['col_nominal']
    agg = dosis_csv_por_grp.get(sname, pd.Series(dtype=float))

    print(f'\n{"━"*90}')
    print(f'  PESTAÑA: {sname}')
    print(f'  Columnas CSV usadas: {", ".join(GRUPOS_CSV[sname])}')
    print(f'{"━"*90}')
    print(f'  {"Municipio":<28} {"En CSV":>10} {"En Excel":>10} {"Diff":>8}  {"Detalle CSV por columna"}')
    print(f'  {"─"*28} {"─"*10} {"─"*10} {"─"*8}  {"─"*30}')

    for row in range(5, ws.max_row + 1):
        mun_val = ws.cell(row=row, column=cm).value
        if not mun_val:
            continue
        mun_str = str(mun_val).strip()
        if 'TOTAL' in mun_str.upper():
            continue

        # Valor en Excel
        xls_raw = ws.cell(row=row, column=cn).value
        try:
            xls_num = int(float(str(xls_raw).replace(',', '')))
        except:
            xls_num = 0

        # Valor del CSV — suma de columnas del grupo
        csv_key = MAPA.get(mun_str)
        if not csv_key:
            print(f'  {mun_str:<28} {"N/A":>10} {xls_num:>10,} {"N/A":>8}  ⚠ SIN MAPEO')
            continue

        mun_df = dur[dur['MUN_NORM'] == csv_key]
        
        # Detalle por columna
        detalle = []
        csv_total = 0
        for col in GRUPOS_CSV[sname]:
            if col in mun_df.columns:
                v = int(mun_df[col].sum())
                csv_total += v
                if v > 0:
                    detalle.append(f'{col.replace("SRP","SRP").replace(" PRIMERA","1a").replace(" SEGUNDA","2a")}={v:,}')

        diff = xls_num - csv_total
        flag = '✅' if diff == 0 else f'❌ DIFF={diff:+,}'
        if diff != 0:
            discrepancias.append((sname, mun_str, csv_total, xls_num, diff))

        det_str = ' | '.join(detalle) if detalle else '(sin datos)'
        print(f'  {mun_str:<28} {csv_total:>10,} {xls_num:>10,} {diff:>+8,}  {flag}  {det_str}')

print()
print('=' * 90)
print(f'TOTAL DISCREPANCIAS: {len(discrepancias)}')
if discrepancias:
    print()
    for sn, mn, csv_v, xls_v, d in discrepancias:
        print(f'  ❌ {sn} | {mn}: CSV={csv_v:,}  Excel={xls_v:,}  Diff={d:+,}')
else:
    print('✅ Sin discrepancias — todos los valores en Excel coinciden con el CSV al 21/06/2026')
print('=' * 90)
