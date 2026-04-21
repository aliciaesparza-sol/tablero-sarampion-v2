import pandas as pd
import numpy as np
from openpyxl import load_workbook
from difflib import get_close_matches

# --- 1. Load PFAM Data ---
pfam_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\mezquital\MICRO PFAM2025.xlsx"
df_pfam = pd.read_excel(pfam_path, sheet_name='MEZQUITAL', header=None)

locality_data = {}
temp_group = []
for i in range(12, len(df_pfam)):
    row = df_pfam.iloc[i]
    name = str(row[1]).strip()
    inegi = str(row[0]).strip() if not pd.isna(row[0]) else "N/A"
    if "TOTAL AREA DE INFLUENCIA" in name:
        pop_val = row[8]
        try:
            pop_val = float(pop_val)
        except:
            pop_val = 0
        for loc_name, loc_inegi in temp_group:
            locality_data[loc_name] = {"pop": pop_val, "inegi": loc_inegi}
        temp_group = []
    elif name != 'nan' and "TOTAL" not in name:
        temp_group.append((name, inegi))

# --- 2. Update Excel File ---
# We'll use a temporary copy to avoid permission issues if user has it open
final_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\FORMATOS\FORMATOS CENSIA\BLOQUEO VACUNAL\FORMATO RESPUESTA RAPIDA_LLENO.xlsx"
temp_path = r"C:\Users\aicil\.gemini\antigravity\scratch\FORMATO_FINAL_TEMP.xlsx"

import shutil
shutil.copy(final_path, temp_path)

wb = load_workbook(temp_path)
ws = wb["ANEXO B"]

pfam_names = list(locality_data.keys())

# Iterate through the rows I added (starting from row 7)
for r in range(7, 30): # Checking up to 30 to be safe
    locality_cell = ws.cell(row=r, column=5) # Locality Name is in Column 5 (E)
    if not locality_cell.value or locality_cell.value == "":
        continue
    
    orig_name = str(locality_cell.value).strip()
    target = orig_name.split('(')[0].split(',')[0].replace("Mezquital", "").strip()
    if "SAN FRANCISCO DE OCOTAN" in orig_name.upper():
        target = "SAN FRANCISCO DEL MEZQUITAL"
    
    matches = get_close_matches(target, pfam_names, n=1, cutoff=0.3)
    if matches:
        match_info = locality_data[matches[0]]
        pop = match_info['pop']
        inegi = match_info['inegi']
        
        # Populate missing fields
        # Column 4: CLUES (Unit)
        ws.cell(row=r, column=4, value=matches[0]) 
        
        # Column 6: AGEB (INEGI)
        ws.cell(row=r, column=6, value=inegi)
        
        # Column 28: Pop Goal
        ws.cell(row=r, column=28, value=int(pop) if pop > 0 else 0)
        
        # Column 29: Coverage (%)
        # Need to read the doses from Column 27 (Index 26)
        doses = ws.cell(row=r, column=27).value
        try:
            doses = float(doses)
        except:
            doses = 0
        reach = (doses / pop * 100) if pop > 0 else 0
        ws.cell(row=r, column=29, value=round(reach, 2))

wb.save(temp_path)
shutil.copy(temp_path, final_path) # Direct overwrite if possible
print(f"File updated successfully: {final_path}")
