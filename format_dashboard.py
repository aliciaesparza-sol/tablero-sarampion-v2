import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

excel_file = r'c:\Users\aicil\OneDrive\Escritorio\PVU\VPH\CAMPAÑA VPH 2025\TABLERO VPH 2025\TABLERO_VPH_10-05-2026.xlsx'
wb = openpyxl.load_workbook(excel_file)
ws_jur = wb['JURISDICCIONES']

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
header_font = Font(color="FFFFFF", bold=True)
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Blue
bold_font = Font(bold=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply styles to Table 1 (Dosis aplicadas en total por Jurisdicción y Grupo)
# The headers for table 1 are in row 3. The data is in rows 4-8. Total row is row 8.
# Find dimensions of table 1
table1_cols = 0
for col in range(1, 20):
    if ws_jur.cell(row=3, column=col).value:
        table1_cols = col
    else:
        break

# Headers for Table 1
for col in range(1, table1_cols + 1):
    cell = ws_jur.cell(row=3, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

# Data for Table 1 (rows 4 to 8)
for row in range(4, 9):
    for col in range(1, table1_cols + 1):
        cell = ws_jur.cell(row=row, column=col)
        cell.border = thin_border
        
        if row == 8: # Total row
            cell.fill = total_fill
            cell.font = bold_font
            
        if col == 1:
            cell.alignment = left_align
            if row == 8:
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        else:
            cell.alignment = center_align

# Number formats for Table 1
for row in range(4, 9):
    for col in range(2, table1_cols + 1):
        ws_jur.cell(row=row, column=col).number_format = '#,##0'

# Apply styles to Table 2 (Dosis Recibidas CENSIA vs Aplicaciones)
# Title is at row 11. Table headers are at row 13.
start_row_t2 = 0
for r in range(8, 20):
    if ws_jur.cell(row=r, column=1).value == 'Jurisdicción' and ws_jur.cell(row=r, column=2).value == 'Exist. Inicial':
        start_row_t2 = r
        break

if start_row_t2 > 0:
    table2_cols = 4
    
    # Headers for Table 2
    for col in range(1, table2_cols + 1):
        cell = ws_jur.cell(row=start_row_t2, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Data for Table 2 (start_row_t2 + 1 to start_row_t2 + 6)
    end_row_t2 = start_row_t2 + 6
    for row in range(start_row_t2 + 1, end_row_t2 + 1):
        for col in range(1, table2_cols + 1):
            cell = ws_jur.cell(row=row, column=col)
            cell.border = thin_border
            
            if row == end_row_t2: # Total row
                cell.fill = total_fill
                cell.font = bold_font
                
            if col == 1:
                cell.alignment = left_align
                if row == end_row_t2:
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                cell.alignment = center_align
                
                # Format strings with commas if possible
                val = cell.value
                if isinstance(val, str):
                    if val.replace(',', '').isdigit():
                        cell.value = int(val.replace(',', ''))
                        cell.number_format = '#,##0'
                elif isinstance(val, (int, float)):
                    cell.number_format = '#,##0'

# Title formatting
ws_jur['A1'].font = Font(bold=True, size=14, color="1F4E78")
ws_jur.cell(row=start_row_t2-2, column=1).font = Font(bold=True, size=14, color="1F4E78")

# Turn off gridlines
ws_jur.sheet_view.showGridLines = False

wb.save(excel_file)
print("Formatting applied successfully.")
