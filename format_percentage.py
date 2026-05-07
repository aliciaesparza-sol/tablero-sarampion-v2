import openpyxl
from openpyxl.styles import NumberFormatDescriptor

file_path = r"C:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\BLOQUEOS VACUNALES\BLOQUEOS VACUNALES 2026\VACUNACIÓN MEZQUITAL 2026\Formato_Concentrado_Mezquital_2026_Con_Datos_Geograficos.xlsx"

try:
    print(f"Loading workbook {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    if "Concentrado" in wb.sheetnames:
        ws = wb["Concentrado"]
        # Column 84 is 'CF' (1-indexed: 84)
        col_idx = 84
        
        print("Formatting cells...")
        for row in range(4, ws.max_row + 1): # Start at row 4 (data)
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value
            try:
                # If value is numeric, convert to percentage format
                if isinstance(val, (int, float)) and val != 0:
                    # Our previous value was like 160.61. For Excel % format, it should be 1.6061
                    cell.value = val / 100
                    cell.number_format = '0.00%'
                elif val == 0:
                    cell.value = 0
                    cell.number_format = '0.00%'
            except:
                pass
        
        wb.save(file_path)
        print("Done!")
    else:
        print("Sheet 'Concentrado' not found.")

except Exception as e:
    print(f"Error: {e}")
