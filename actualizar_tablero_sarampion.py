
# -*- coding: utf-8 -*-
"""
Actualización del Tablero de Sarampión - Corte 21 de junio 2026
Lee el CSV con datos acumulados y actualiza el Excel con:
  - Columna Nominal (Dosis Aplicadas)
  - Total Dosis
  - Pendientes
  - Cobertura vs Meta
  - Semáforo
  - Encabezados "Al corte del 21 de junio"
  - Fila TOTAL DURANGO
  - RESUMEN MUNICIPIOS
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE ARCHIVOS
# ─────────────────────────────────────────────────────────────────────────────
CSV_PATH   = r'C:\SRP\SRP-SR-2025_21-06-2026 08-32-01.csv'
XLSX_IN    = r'C:\Users\aicil\.gemini\antigravity-ide\scratch\temp_cobertura.xlsx'
XLSX_OUT   = r'C:\Users\aicil\.gemini\antigravity-ide\scratch\COBERTURA_SARAMPION_21JUN2026_ACTUALIZADO.xlsx'

# ─────────────────────────────────────────────────────────────────────────────
# MAPA: nombre en Excel → nombre en CSV (sin acentos, mayúsculas)
# ─────────────────────────────────────────────────────────────────────────────
MAPA_MUNICIPIOS = {
    'Peñón Blanco':         'PEÑON BLANCO',          # CSV tiene PEÃ\x91ON BLANCO por encoding, normalizamos
    'Santa Clara':          'SANTA CLARA',
    'Guanaceví':            'GUANACEVI',
    'San Bernardo':         'SAN BERNARDO',
    'San Pedro del Gallo':  'SAN PEDRO DEL GALLO',
    'Topia':                'TOPIA',
    'Tepehuanes':           'TEPEHUANES',
    'San Juan de Guadalupe':'SAN JUAN DE GUADALUPE',
    'Canelas':              'CANELAS',
    'Súchil':               'SUCHIL',
    'Otáez':                'OTAEZ',
    'Poanas':               'POANAS',
    'Cuencamé':             'CUENCAME',
    'Nombre de Dios':       'NOMBRE DE DIOS',
    'Indé':                 'INDE',
    'San Dimas':            'SAN DIMAS',
    'Mezquital':            'MEZQUITAL',
    'Pueblo Nuevo':         'PUEBLO NUEVO DGO',
    'General Simón Bolívar':'GENERAL SIMON BOLIVAR',
    'Hidalgo':              'HIDALGO DGO',
    'Mapimí':               'MAPIMI',
    'Coneto de Comonfort':  'CONETO DE COMONFORT',
    'Nazas':                'NAZAS',
    'Canatlán':             'CANATLAN',
    'Ocampo':               'OCAMPO DGO',
    'El Oro':               'ORO EL',
    'Rodeo':                'RODEO',
    'Durango':              'DURANGO',
    'Lerdo':                'LERDO',
    'Pánuco de Coronado':   'PANUCO DE CORONADO',
    'Gómez Palacio':        'GOMEZ PALACIO',
    'Nuevo Ideal':          'NUEVO IDEAL',
    'Tamazula':             'TAMAZULA',
    'Tlahualilo':           'TLAHUALILO',
    'Guadalupe Victoria':   'GUADALUPE VICTORIA DGO',
    'San Luis del Cordero': 'SAN LUIS DEL CORDERO',
    'San Juan del Río':     'SAN JUAN DEL RIO DGO',
    'Vicente Guerrero':     'VICENTE GUERRERO',
    'Santiago Papasquiaro': 'SANTIAGO PAPASQUIARO',
}

# ─────────────────────────────────────────────────────────────────────────────
# COLUMNAS CSV POR GRUPO ETARIO
# ─────────────────────────────────────────────────────────────────────────────
GRUPOS_CSV = {
    '6-11 Meses':    ['SRP 6 A 11 MESES PRIMERA', 'SR 6 A 11 MESES PRIMERA'],
    '1 Año':         ['SRP 1 ANIO  PRIMERA', 'SR 1 ANIO PRIMERA'],
    '18 Meses':      ['SRP 18 MESES SEGUNDA', 'SR 18 MESES SEGUNDA'],
    '6 Años':        ['SRP 6 ANIOS PRIMERA', 'SRP 6 ANIOS SEGUNDA',
                      'SR 6 ANIOS PRIMERA',  'SR 6 ANIOS SEGUNDA'],
    'Rezag 2-12 Años': ['SRP 2 A 5 ANIOS PRIMERA', 'SRP 2 A 5 ANIOS SEGUNDA',
                        'SRP 7 A 9 ANIOS PRIMERA',  'SRP 7 A 9 ANIOS SEGUNDA',
                        'SRP 10 A 12 ANIOS PRIMERA','SRP 10 A 12 ANIOS SEGUNDA',
                        'SR 2 A 5 ANIOS PRIMERA',   'SR 2 A 5 ANIOS SEGUNDA',
                        'SR 7 A 9 ANIOS PRIMERA',   'SR 7 A 9 ANIOS SEGUNDA',
                        'SR 10 A 12 ANIOS PRIMERA',  'SR 10 A 12 ANIOS SEGUNDA'],
    '13-19 Años':    ['SRP 13 A 19 ANIOS PRIMERA', 'SRP 13 A 19 ANIOS SEGUNDA',
                      'SR 13 A 19 ANIOS PRIMERA',   'SR 13 A 19 ANIOS SEGUNDA'],
    '20-39 Años':    ['SRP 20 A 29 ANIOS PRIMERA', 'SRP 20 A 29 ANIOS SEGUNDA',
                      'SRP 30 A 39 ANIOS PRIMERA', 'SRP 30 A 39 ANIOS SEGUNDA',
                      'SR 20 A 29 ANIOS PRIMERA',  'SR 20 A 29 ANIOS SEGUNDA',
                      'SR 30 A 39 ANIOS PRIMERA',  'SR 30 A 39 ANIOS SEGUNDA'],
    '40-49 Años':    ['SRP 40 A 49 ANIOS PRIMERA', 'SRP 40 A 49 ANIOS SEGUNDA',
                      'SR 40 A 49 ANIOS PRIMERA',  'SR 40 A 49 ANIOS SEGUNDA'],
}

# ─────────────────────────────────────────────────────────────────────────────
# SEMÁFORO
# ─────────────────────────────────────────────────────────────────────────────
def semaforo(cob_pct):
    """Recibe cobertura como decimal (e.g. 0.95) o porcentaje (e.g. 95.0)"""
    if cob_pct is None:
        return '⬛ SIN DATOS'
    # Normalizar a porcentaje
    if cob_pct <= 1.5:
        cob_pct = cob_pct * 100
    if cob_pct >= 85:
        return '✅ META CUMPLIDA'
    elif cob_pct >= 60:
        return '🟡 EN PROCESO'
    else:
        return '🔴 CRÍTICO'

def color_semaforo(texto):
    """Devuelve color de fondo hexadecimal para el semáforo"""
    if '✅' in texto:
        return '00AA44'   # verde oscuro
    elif '🟡' in texto:
        return 'FFC000'   # amarillo
    elif '🔴' in texto:
        return 'FF0000'   # rojo
    else:
        return 'AAAAAA'

def parse_cob(val):
    """Convierte valor de celda (str '95.3%' o float 0.953 o 95.3) a float porcentaje 0-100"""
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip().replace('%', '').replace(',', '.')
        try:
            return float(s)
        except:
            return None
    if isinstance(val, (int, float)):
        if val <= 1.5:
            return val * 100
        return float(val)
    return None

# ─────────────────────────────────────────────────────────────────────────────
# LEER CSV Y CALCULAR DOSIS POR MUNICIPIO/GRUPO
# ─────────────────────────────────────────────────────────────────────────────
print('Leyendo CSV...')
df = pd.read_csv(CSV_PATH, encoding='latin-1', low_memory=False)
dur = df[df['ESTADO'].str.upper().str.contains('DURANGO', na=False)].copy()
print(f'  Registros Durango: {len(dur)}')

# Normalizar municipio en CSV (eliminar caracteres de encoding)
def norm_csv_mun(s):
    if not isinstance(s, str):
        return ''
    s = s.upper().strip()
    # Arreglar Peñón Blanco que aparece como PEÃ\x91ON BLANCO en latin-1 mal leído
    s = s.replace('PEÃ\x91ON', 'PEÑON').replace('PEÃON', 'PEÑON')
    s = s.replace('Ã', 'Ñ').replace('Ã\x91', 'Ñ')
    return s

dur['MUN_NORM'] = dur['MUNICIPIO'].apply(norm_csv_mun)

# Calcular dosis por grupo y municipio
dosis_por_grupo = {}
for grp, cols in GRUPOS_CSV.items():
    existentes = [c for c in cols if c in dur.columns]
    for c in existentes:
        dur[c] = pd.to_numeric(dur[c], errors='coerce').fillna(0)
    agg = dur.groupby('MUN_NORM')[existentes].sum().sum(axis=1)
    dosis_por_grupo[grp] = agg
    print(f'  Grupo {grp}: {int(agg.sum())} dosis totales Durango')

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN AUXILIAR: obtener dosis CSV para un municipio/grupo
# ─────────────────────────────────────────────────────────────────────────────
def get_dosis(mun_excel, grp):
    csv_key = MAPA_MUNICIPIOS.get(mun_excel)
    if csv_key is None:
        print(f'  ⚠️  Municipio no mapeado: {mun_excel}')
        return 0
    agg = dosis_por_grupo.get(grp, pd.Series(dtype=float))
    val = agg.get(csv_key, 0)
    return int(val) if val == val else 0  # NaN → 0

# ─────────────────────────────────────────────────────────────────────────────
# ABRIR EXCEL (preservando formatos)
# ─────────────────────────────────────────────────────────────────────────────
print('\nAbriendo Excel...')
wb = openpyxl.load_workbook(XLSX_IN)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE CADA HOJA:
#  col_mun: columna donde está el municipio (1-based)
#  col_cubos, col_nominal, col_total, col_pend, col_cob, col_sem: columnas de datos
#  titulo_fila2: texto de resumen en fila 2
# ─────────────────────────────────────────────────────────────────────────────
HOJA_CONFIG = {
    '6-11 Meses': {
        'col_mun':4, 'corr':-2,   # ajuste: columna B = 2, pero encabezado dice Municipio en col2
        'col_mun':2, 'col_universo':3, 'col_meta_pct':4, 'col_meta':5,
        'col_cubos':6, 'col_nominal':7, 'col_total':8,
        'col_pend':9, 'col_cob':10, 'col_sem':11,
    },
    '1 Año': {
        'col_mun':2, 'col_universo':3, 'col_meta_pct':4, 'col_meta':5,
        'col_cubos':6, 'col_nominal':7, 'col_total':8,
        'col_pend':9, 'col_cob':10, 'col_sem':11,
    },
    '18 Meses': {
        'col_mun':2, 'col_universo':3, 'col_meta_pct':4, 'col_meta':5,
        'col_cubos':6, 'col_nominal':7, 'col_total':8,
        'col_pend':9, 'col_cob':10, 'col_sem':11,
    },
    '6 Años': {
        'col_mun':1, 'col_universo':2, 'col_meta_pct':3, 'col_meta':4,
        'col_cubos':5, 'col_nominal':6, 'col_total':7,
        'col_pend':8, 'col_cob':9, 'col_sem':10,
    },
    'Rezag 2-12 Años': {
        'col_mun':2, 'col_universo':3, 'col_meta_pct':4, 'col_meta':5,
        'col_cubos':6, 'col_nominal':7, 'col_total':8,
        'col_pend':9, 'col_cob':10, 'col_sem':11,
    },
    '13-19 Años': {
        'col_mun':2, 'col_universo':3, 'col_meta_pct':4, 'col_meta':5,
        'col_cubos':6, 'col_nominal':7, 'col_total':8,
        'col_pend':9, 'col_cob':10, 'col_sem':11,
    },
    '20-39 Años': {
        'col_mun':2, 'col_universo':3, 'col_meta_pct':4, 'col_meta':5,
        'col_cubos':6, 'col_nominal':7, 'col_total':8,
        'col_pend':9, 'col_cob':10, 'col_sem':11,
    },
    '40-49 Años': {
        'col_mun':2, 'col_universo':3, 'col_meta_pct':4, 'col_meta':5,
        'col_cubos':6, 'col_nominal':7, 'col_total':8,
        'col_pend':9, 'col_cob':10, 'col_sem':11,
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# FUNCIÓN PARA PARSEAR NÚMERO DE CELDA (puede ser str con comas o int/float)
# ─────────────────────────────────────────────────────────────────────────────
def parse_num(val, default=0):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        s = val.strip().replace(',', '').replace('%', '').replace(' ', '')
        try:
            return float(s)
        except:
            return default
    return default

# ─────────────────────────────────────────────────────────────────────────────
# ACTUALIZAR CADA HOJA DE GRUPO ETARIO
# ─────────────────────────────────────────────────────────────────────────────
for sname, cfg in HOJA_CONFIG.items():
    if sname not in wb.sheetnames:
        print(f'⚠️  Hoja "{sname}" no encontrada, saltando...')
        continue
    
    ws = wb[sname]
    print(f'\nActualizando hoja: {sname}')
    
    cm = cfg['col_mun']
    cu = cfg['col_universo']
    cmp = cfg['col_meta_pct']
    cmet = cfg['col_meta']
    ccub = cfg['col_cubos']
    cnom = cfg['col_nominal']
    ctot = cfg['col_total']
    cpend = cfg['col_pend']
    ccob = cfg['col_cob']
    csem = cfg['col_sem']
    
    # Acumuladores para fila TOTAL y fila 2
    total_univ = 0
    total_meta = 0
    total_cubos = 0
    total_nominal = 0
    total_total = 0
    
    for row in range(5, ws.max_row + 1):
        mun_val = ws.cell(row=row, column=cm).value
        if mun_val is None:
            continue
        mun_str = str(mun_val).strip()
        
        # Fila de TOTAL DURANGO
        if 'TOTAL' in mun_str.upper():
            # Lo actualizaremos después
            continue
        
        # Obtener datos actuales de la fila
        universo = parse_num(ws.cell(row=row, column=cu).value)
        meta = parse_num(ws.cell(row=row, column=cmet).value)
        cubos = parse_num(ws.cell(row=row, column=ccub).value)
        
        # Nuevas dosis del CSV
        nuevas_dosis = get_dosis(mun_str, sname)
        
        # Calcular totales
        total_dosis = int(cubos) + nuevas_dosis
        
        # Pendientes: si total > meta → SUPER +X, si no → meta - total
        pendientes_val = meta - total_dosis
        if pendientes_val <= 0:
            pend_str = f'SUPER +{int(abs(pendientes_val))}'
        else:
            pend_str = int(pendientes_val)
        
        # Cobertura
        if meta > 0:
            cob_pct = (total_dosis / meta) * 100
        else:
            cob_pct = 0
        
        cob_str = f'{cob_pct:.1f}%'
        sem_str = semaforo(cob_pct)
        
        # Escribir en Excel
        ws.cell(row=row, column=cnom).value = nuevas_dosis
        ws.cell(row=row, column=ctot).value = total_dosis
        ws.cell(row=row, column=cpend).value = pend_str
        ws.cell(row=row, column=ccob).value = cob_str
        ws.cell(row=row, column=csem).value = sem_str
        
        # Colorear semáforo
        color = color_semaforo(sem_str)
        ws.cell(row=row, column=csem).fill = PatternFill(fill_type='solid', fgColor=color)
        ws.cell(row=row, column=csem).font = Font(bold=True, color='FFFFFF')
        ws.cell(row=row, column=csem).alignment = Alignment(horizontal='center')
        
        # Colorear cobertura
        cob_cell = ws.cell(row=row, column=ccob)
        if cob_pct >= 85:
            cob_cell.fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
            cob_cell.font = Font(bold=True, color='276221')
        elif cob_pct >= 60:
            cob_cell.fill = PatternFill(fill_type='solid', fgColor='FFEB9C')
            cob_cell.font = Font(bold=True, color='9C6500')
        else:
            cob_cell.fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
            cob_cell.font = Font(bold=True, color='9C0006')
        
        # Acumular para totales
        total_univ   += universo
        total_meta   += meta
        total_cubos  += cubos
        total_nominal += nuevas_dosis
        total_total  += total_dosis
        
        print(f'  {mun_str}: Nominal={nuevas_dosis}, Total={total_dosis}, Cob={cob_str}, Sem={sem_str}')
    
    # Actualizar fila TOTAL DURANGO
    for row in range(5, ws.max_row + 1):
        mun_val = ws.cell(row=row, column=cm).value
        if mun_val and 'TOTAL' in str(mun_val).upper():
            total_row = row
            ws.cell(row=row, column=cnom).value = int(total_nominal)
            ws.cell(row=row, column=ctot).value = int(total_total)
            
            pend_tot = total_meta - total_total
            if pend_tot <= 0:
                ws.cell(row=row, column=cpend).value = f'SUPER +{int(abs(pend_tot))}'
            else:
                ws.cell(row=row, column=cpend).value = int(pend_tot)
            
            cob_tot = (total_total / total_meta * 100) if total_meta > 0 else 0
            ws.cell(row=row, column=ccob).value = f'{cob_tot:.1f}%'
            sem_tot = semaforo(cob_tot)
            ws.cell(row=row, column=csem).value = sem_tot
            ws.cell(row=row, column=csem).fill = PatternFill(fill_type='solid', fgColor=color_semaforo(sem_tot))
            ws.cell(row=row, column=csem).font = Font(bold=True, color='FFFFFF')
            
            print(f'  TOTAL DURANGO: Nominal={int(total_nominal)}, Total={int(total_total)}, Cob={cob_tot:.1f}%')
            break
    
    # Actualizar fila 1 (título) y fila 2 (resumen)
    # Fila 1: cambiar "Corte 11 de junio" por "Corte 21 de junio 2026"
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value and isinstance(cell.value, str):
            if 'corte' in cell.value.lower() or 'sarampión' in cell.value.lower() or 'sarampion' in cell.value.lower():
                nuevo_titulo = cell.value
                nuevo_titulo = re.sub(r'[Cc]orte\s+\d+\s+de\s+\w+\s*\d*', 'Corte 21 de junio 2026', nuevo_titulo)
                nuevo_titulo = re.sub(r'[Cc]orte\s+\w+\s+\d+,?\s*\d*', 'Corte 21 de junio 2026', nuevo_titulo)
                ws.cell(row=1, column=col).value = nuevo_titulo
    
    # Fila 2: actualizar resumen estadístico
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=2, column=col)
        if cell.value and isinstance(cell.value, str) and ('Univ' in cell.value or 'Meta' in cell.value or 'Cubos' in cell.value or 'Nominal' in cell.value):
            cob_tot = (total_total / total_meta * 100) if total_meta > 0 else 0
            
            # Obtener % meta de la primera fila de datos
            meta_pct_val = ws.cell(row=5, column=cmp).value
            meta_pct_str = str(meta_pct_val) if meta_pct_val else '50%'
            
            nuevo_resumen = (
                f'% Meta: {meta_pct_str}  |  '
                f'Univ: {int(total_univ):,}  |  '
                f'Meta: {int(total_meta):,}  |  '
                f'Cubos: {int(total_cubos):,}  |  '
                f'Nominal: {int(total_nominal):,}  |  '
                f'Total: {int(total_total):,}  |  '
                f'Cob: {cob_tot:.1f}% — Corte 21 de junio 2026'
            )
            ws.cell(row=2, column=col).value = nuevo_resumen
            break

# ─────────────────────────────────────────────────────────────────────────────
# ACTUALIZAR RESUMEN MUNICIPIOS
# ─────────────────────────────────────────────────────────────────────────────
print('\nActualizando RESUMEN MUNICIPIOS...')
ws_res = wb['RESUMEN MUNICIPIOS']

# Fila 2: actualizar corte
for col in range(1, ws_res.max_column + 1):
    cell = ws_res.cell(row=2, column=col)
    if cell.value and isinstance(cell.value, str):
        nuevo = re.sub(r'\d+\s+de\s+\w+\s+\d{4}', '21 de junio 2026', cell.value)
        nuevo = re.sub(r'\d+\w*\s+[Jj]une?\s+\d{4}', '21 de junio 2026', nuevo)
        if nuevo != cell.value:
            ws_res.cell(row=2, column=col).value = nuevo

# Recalcular fila de totales en resumen
# Sumar dosis de todos los grupos para obtener totales globales
total_nominal_global = sum(dosis_por_grupo[g].sum() for g in dosis_por_grupo)
print(f'  Total dosis Durango (todos grupos): {int(total_nominal_global):,}')

# Actualizar la celda de nominal en resumen (col E = 5) y total (col F = 6)
for row in range(3, ws_res.max_row + 1):
    row_vals = [ws_res.cell(row=row, column=c).value for c in range(1, 9)]
    row_text = ' '.join(str(v) for v in row_vals if v is not None).lower()
    # Buscar fila de datos principale (Durango total)
    if any(v is not None for v in row_vals):
        # Actualizar col E (Nominal)
        col_nom_res = 5
        col_tot_res = 6
        col_cob_res = 7
        
        cubos_res = ws_res.cell(row=row, column=4).value
        if cubos_res is not None:
            cubos_num = parse_num(cubos_res)
            meta_res = parse_num(ws_res.cell(row=row, column=3).value)
            univ_res = parse_num(ws_res.cell(row=row, column=2).value)
            
            if meta_res > 0:
                ws_res.cell(row=row, column=col_nom_res).value = f'{int(total_nominal_global):,}'
                total_res = cubos_num + total_nominal_global
                ws_res.cell(row=row, column=col_tot_res).value = int(total_res)
                cob_res = (total_res / meta_res * 100) if meta_res > 0 else 0
                ws_res.cell(row=row, column=col_cob_res).value = f'{cob_res:.1f}%'
                print(f'  Resumen fila {row}: Nominal={int(total_nominal_global):,}, Total={int(total_res):,}, Cob={cob_res:.1f}%')
                break

# ─────────────────────────────────────────────────────────────────────────────
# GUARDAR
# ─────────────────────────────────────────────────────────────────────────────
print(f'\nGuardando como: {XLSX_OUT}')
wb.save(XLSX_OUT)
print('✅ ¡Tablero actualizado correctamente!')
print(f'   Archivo: {XLSX_OUT}')
