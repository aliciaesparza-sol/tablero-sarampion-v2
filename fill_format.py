import pandas as pd
from openpyxl import load_workbook
import datetime

# --- DATA PREPARATION ---

# Data from PDF (Extracted manually from previous view)
pdf_data = [
    {"date": "29 Ene", "locality": "La Guajolota (Las Aguilillas)", "srp": 105, "sr": 327, "inst": "SSA"},
    {"date": "30 Ene", "locality": "La Guajolota (Bajío y Centro)", "srp": 175, "sr": 228, "inst": "SSA"},
    {"date": "03 Feb", "locality": "Cerro Bolillo, Sta. Ma. de Ocotán", "srp": 150, "sr": 445, "inst": "SSA"},
    {"date": "11 Feb", "locality": "Luis Moya (Gpe. Victoria)", "srp": 2, "sr": 151, "inst": "SSA"},
    {"date": "12 Feb", "locality": "Luis Moya (Gpe. Victoria)", "srp": 2, "sr": 164, "inst": "SSA"},
    {"date": "24 Feb", "locality": "Cerro Bolillo, Sta. Ma. de Ocotán", "srp": 120, "sr": 358, "inst": "SSA"},
    {"date": "01 Mar", "locality": "Las Joyas, Mezquital", "srp": 124, "sr": 483, "inst": "SSA"},
    {"date": "26 Mar", "locality": "La Huazamotita", "srp": 0, "sr": 307, "inst": "SSA"},
    {"date": "27 Mar", "locality": "La Huazamotita", "srp": 0, "sr": 281, "inst": "SSA"},
    {"date": "28 Mar", "locality": "Las Joyas, Mezquital", "srp": 95, "sr": 693, "inst": "SSA"},
]

# Data from Excel (Extracting from san fco de ocotan file)
excel_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\BLOQUEOS VACUNALES\BLOQUEOS VACUNALES 2026\SAN FRANCISCO DE OCOTAN, MEZQUITAL_17.04.2026\san fco de ocotan Formato_Concentrado_Vacunacion_Sarampion-mezquital.xlsx"
df_excel = pd.read_excel(excel_path, sheet_name='Concentrado', header=None)

# Localities in Excel start from row 4 (index 4)
# Column indices in Concentrado: 2(Localidad), 3(Fecha), 51(SRP), 52(SR), 23(Total Pop?), 
# Age groups in Concentrado: 14(6-11m), 15(1y), 16(2-4y), 17(5-9y), 18(10-19y), 19(20-39y), 20(40-49y), 21(50+)

excel_entries = []
for i in range(4, 9): # Rows 4 to 8 are the localities
    row = df_excel.iloc[i]
    entry = {
        "locality": row[2],
        "date": str(row[3]),
        "inst": "IMSS-C", # Based on previous check
        "srp": row[51],
        "sr": row[52],
        "ages": {
            "<1": row[14],
            "1y": row[15],
            "2-4y": row[16],
            "5-9y": row[17],
            "10-19y": row[18],
            "20-39y": row[19] + row[20], # Combine for 20-49 if needed
            "50+": row[21]
        },
        "total_applied": row[51] + row[52]
    }
    excel_entries.append(entry)

# --- WRITING TO TARGET ---

target_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\FORMATOS\FORMATOS CENSIA\BLOQUEO VACUNAL\FORMATO RESPUESTA RAPIDA.xlsx"
output_path = r"c:\Users\aicil\OneDrive\Escritorio\PVU\SARAMPIÓN\FORMATOS\FORMATOS CENSIA\BLOQUEO VACUNAL\FORMATO RESPUESTA RAPIDA_LLENO.xlsx"

wb = load_workbook(target_path)
ws = wb["ANEXO B"]

current_row = 7 # Starting below headers

# Add PDF entries
for entry in pdf_data:
    ws.cell(row=current_row, column=2, value="RESUMEN LOCALIDAD") # Patient Name
    ws.cell(row=current_row, column=3, value=entry["inst"])
    ws.cell(row=current_row, column=5, value=entry["locality"])
    ws.cell(row=current_row, column=9, value="X") # Rural
    ws.cell(row=current_row, column=22, value=entry["date"]) # Barrido column (index 21+1)
    ws.cell(row=current_row, column=25, value=entry["srp"]) # SRP column (index 24+1)
    ws.cell(row=current_row, column=26, value=entry["sr"]) # SR column (index 25+1)
    ws.cell(row=current_row, column=27, value=entry["srp"] + entry["sr"]) # Total
    ws.cell(row=current_row, column=30, value="Información de Informe Mezquital")
    current_row += 1

# Add Excel entries
for entry in excel_entries:
    ws.cell(row=current_row, column=2, value="RESUMEN LOCALIDAD")
    ws.cell(row=current_row, column=3, value=entry["inst"])
    ws.cell(row=current_row, column=5, value=entry["locality"])
    ws.cell(row=current_row, column=9, value="X") # Rural
    
    # Ages
    ws.cell(row=current_row, column=10, value=entry["ages"]["<1"])
    ws.cell(row=current_row, column=11, value=entry["ages"]["1y"])
    ws.cell(row=current_row, column=12, value=entry["ages"]["2-4y"]) # index 11+1
    ws.cell(row=current_row, column=14, value=entry["ages"]["5-9y"]) # index 13+1
    ws.cell(row=current_row, column=15, value=entry["ages"]["10-19y"]) # index 14+1
    ws.cell(row=current_row, column=16, value=entry["ages"]["20-39y"]) # index 15+1
    ws.cell(row=current_row, column=18, value=entry["ages"]["50+"]) # index 17+1

    ws.cell(row=current_row, column=22, value=entry["date"]) # Barrido
    ws.cell(row=current_row, column=25, value=entry["srp"])
    ws.cell(row=current_row, column=26, value=entry["sr"])
    ws.cell(row=current_row, column=27, value=entry["total_applied"])
    ws.cell(row=current_row, column=30, value="Información de Concentrado San Fco de Ocotán")
    current_row += 1

wb.save(output_path)
print(f"File saved to: {output_path}")
print(f"Total rows added: {len(pdf_data) + len(excel_entries)}")
